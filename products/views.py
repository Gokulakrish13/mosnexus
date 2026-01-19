from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from .models import UserRole, UserStreamAccess
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_http_methods
from django.db import models
from django.db.models import Q
from django.core.files.storage import default_storage
from .models import Product, ProductHistory, Category, Participant, Location, Stream, SystemStatusHistory, UserDataVersion, SubLevel, SubLevelHistory, SubLevelTool, SubLevelToolHistory, LegacyExcelUpload, ZenitionProduct, UsageTracking, SystemStatus, UserSession, SystemMetrics, Note, NoteAttachment, SharedNote, CustomUser, UserRole, UserStreamAccess, SystemTag, HolisticSystem, HolisticWeeklyData, HolisticSystemHistory, SystemDowntime, SystemDowntimeMetrics, BuildServer, BuildServerHistory, BuildServerMaintenanceLog, Floor
from .utils import get_stream_or_404
import qrcode
from io import BytesIO
import base64
import json
from django.http import HttpResponse, FileResponse, JsonResponse
from django.template.loader import render_to_string
from django.views.decorators.http import require_GET
from django.contrib.auth import get_user_model
from PIL import Image as PILImage, ImageDraw, ImageFont
import openpyxl
import os
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from datetime import datetime, date, timedelta, timezone
import pandas as pd
from django.utils.html import escape
import re
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
import json
import logging

logger = logging.getLogger(__name__)
from .models import SystemAllocation
from django.views.decorators.csrf import csrf_exempt
import json
from django.utils.timezone import localtime, make_aware
from .models import System
from django.contrib.auth.decorators import user_passes_test
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.db import IntegrityError
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
import io
from django.db.models import Q
from django.db.models import Count, Avg, F
import numpy as np
from .models import ZenitionProduct, ProductEntry
from openpyxl import Workbook
from .models import Communication, CommunicationAttachment
import subprocess
import sys
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.views import PasswordChangeView, PasswordChangeDoneView
from .models import CustomUser, Notification
from django.contrib.auth.hashers import check_password
from django.contrib.auth import update_session_auth_hash
from django.utils import timezone
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import mm
from django.views.decorators.http import require_POST
from django.db.models import F
from datetime import datetime, date, timedelta
import pickle
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import csv
from django.http import HttpResponse
import logging

logger = logging.getLogger(__name__)

class CustomPasswordChangeView(PasswordChangeView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['selected_stream'] = 'HIC'  # Default stream, adjust if needed
        return context

class CustomPasswordChangeDoneView(PasswordChangeDoneView):
    template_name = 'products/password_change_done.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['selected_stream'] = 'HIC'
        return context

def is_admin(user):
    if hasattr(user, 'custom_profile'):
        return user.custom_profile.is_admin()
    return user.is_superuser

def is_super_admin(user):
    if hasattr(user, 'custom_profile'):
        return user.custom_profile.is_super_admin()
    return user.is_superuser

def is_lab_incharge(user):
    if hasattr(user, 'custom_profile'):
        return user.custom_profile.is_lab_incharge()
    return user.is_superuser

def can_manage_users(user):
    if hasattr(user, 'custom_profile'):
        return user.custom_profile.can_manage_users()
    return user.is_superuser

def can_manage_system_allocation(user):
    return user.is_authenticated

def can_edit_products(user):
    if hasattr(user, 'custom_profile'):
        return user.custom_profile.can_edit_products()
    return user.is_superuser

def can_delete_products(user):
    if hasattr(user, 'custom_profile'):
        return user.custom_profile.can_delete_products()
    return user.is_superuser

def can_view_analytics(user):
    if hasattr(user, 'custom_profile'):
        return user.custom_profile.can_view_analytics()
    return user.is_superuser

# Utility functions for access control
def check_user_access(request, stream=None):
    """
    Check if user has access to the application and specific stream.
    Returns (has_access, error_message, custom_profile)
    """
    custom_profile, created = CustomUser.objects.get_or_create(user=request.user)

    # Check if user has any roles (except superusers)
    if not request.user.is_superuser and not custom_profile.user_roles.exists():
        error_message = 'Access denied. You have no assigned roles. Please contact an administrator.'
        return False, error_message, custom_profile

    # If stream is specified, check stream access
    if stream and not request.user.is_superuser and not custom_profile.can_access_stream(stream):
        error_message = f'Access denied. You do not have permission to access the {stream} stream.'
        return False, error_message, custom_profile

    return True, None, custom_profile

def home(request):
    return render(request, 'products/home.html')

@login_required
def product_list(request, stream=None):
    category_id = request.GET.get('category')
    q = request.GET.get('q', '').strip()
    stream = stream or request.GET.get('stream', 'HIC')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    location_id = request.GET.get('location')

    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')

    # Get the Stream object for database queries (404 if not found)
    stream_obj = get_stream_or_404(stream)

    if category_id:
        products = Product.objects.filter(category_id=category_id, stream=stream_obj)
    else:
        products = Product.objects.filter(stream=stream_obj)
    if q:
        from django.db.models import Q
        products = products.filter(
            Q(name__icontains=q) |
            Q(serial_number__icontains=q) |
            Q(description__icontains=q) |
            Q(twelve_nc__icontains=q) |
            Q(location__name__icontains=q) |
            Q(category__serial_number__icontains=q)
        )
    # Date range filtering using created_at date (works directly with YYYY-MM-DD strings from the form)
    if start_date and end_date:
        # Filter inclusively between the two dates
        products = products.filter(created_at__date__range=(start_date, end_date))
    else:
        if start_date:
            products = products.filter(created_at__date__gte=start_date)
        if end_date:
            products = products.filter(created_at__date__lte=end_date)
    if location_id:
        products = products.filter(location_id=location_id).exclude(location=None)
    products = products.prefetch_related('system_tags__system').order_by('location__name', '-created_at')
    locations = Location.objects.filter(stream=stream_obj).order_by('name')
    
    # Calculate product counts
    total_count = products.count()
    status_counts = {
        'Active': products.filter(status='Active').count(),
        'Not Active': products.filter(status='Not Active').count(),
        'Scraped': products.filter(status='Scraped').count(),
        'Hand-Overed': products.filter(status='Hand-Overed').count(),
        'Issue': products.filter(status='Issue').count(),
    }
    
    return render(request, 'products/product_list.html', {
        'products': products,
        'selected_stream': stream,
        'category_id': category_id,
        'product_list_url': f'/stream/{stream}/products/?category={category_id}' if category_id else f'/stream/{stream}/products/',
        'locations': locations,
        'total_count': total_count,
        'status_counts': status_counts,
    })

@login_required
def product_create(request, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    category_id = request.GET.get('category')
    # Always get the Stream object from name
    # Always get the Stream object from name (404 if not found)
    stream_obj = get_stream_or_404(stream, default='HIC')
    categories = Category.objects.filter(stream=stream_obj).order_by('name')
    locations = Location.objects.filter(stream=stream_obj).order_by('name')
    selected_category = None
    next_serial = ''
    if category_id:
        try:
            selected_category = Category.objects.get(id=category_id, stream=stream_obj)
            # Find the max serial_number for this category, increment for next
            last_product = Product.objects.filter(category_id=category_id, stream=stream_obj).order_by('-serial_number').first()
            if last_product and last_product.serial_number.isdigit():
                next_serial = str(int(last_product.serial_number) + 1)
            else:
                # If no products, start from category serial + 1
                if selected_category.serial_number.isdigit():
                    next_serial = str(int(selected_category.serial_number) + 1)
        except Category.DoesNotExist:
            selected_category = None
    if request.method == 'POST':
        name = request.POST.get('name', '')
        serial_number = request.POST.get('serial_number', '')
        description = request.POST.get('description', '')
        issue_description = request.POST.get('issue_description', '')
        twelve_nc = request.POST.get('twelve_nc', '')
        category_id = request.POST.get('category', '')
        status = request.POST.get('status', 'Active')
        handover_team_type = request.POST.get('handover_team_type', '')
        handover_external_team = request.POST.get('handover_external_team', '')
        handover_owner = request.POST.get('handover_owner', '')
        location_id = request.POST.get('location', '')
        location = Location.objects.filter(id=location_id, stream=stream_obj).first() if location_id else None
        # If category_id is missing or not a number, show warning and preserve form
        if not category_id.isdigit():
            messages.warning(request, 'Please select a valid category before submitting the form.')
            return render(request, 'products/product_form.html', {
                'categories': categories,
                'selected_category': None,
                'category': None,
                'next_serial': next_serial,
                'locations': locations,
                'stream': stream,
                'selected_stream': stream,
                'product': {
                    'name': name,
                    'serial_number': serial_number,
                    'description': description,
                    'issue_description': issue_description,
                    'twelve_nc': twelve_nc,
                    'status': status,
                    'handover_team_type': handover_team_type,
                    'handover_external_team': handover_external_team,
                    'handover_owner': handover_owner,
                    'location': location,
                },
                'edit': False
            })
        # Try to get category, handle missing gracefully
        try:
            category = Category.objects.get(pk=category_id, stream=stream_obj)
        except Category.DoesNotExist:
            messages.warning(request, 'No category matches the given query. Please select a valid category.')
            return render(request, 'products/product_form.html', {
                'categories': categories,
                'selected_category': None,
                'category': None,
                'next_serial': next_serial,
                'locations': locations,
                'stream': stream,
                'selected_stream': stream,
                'product': {
                    'name': name,
                    'serial_number': serial_number,
                    'description': description,
                    'issue_description': issue_description,
                    'twelve_nc': twelve_nc,
                    'status': status,
                    'handover_team_type': handover_team_type,
                    'handover_external_team': handover_external_team,
                    'handover_owner': handover_owner,
                    'location': location,
                },
                'edit': False
            })
        # Check for duplicate serial number in this stream
        if Product.objects.filter(serial_number=serial_number, stream=stream_obj).exists():
            messages.error(request, f"A product with serial number '{serial_number}' already exists in this stream.")
            return render(request, 'products/product_form.html', {
                'categories': categories,
                'selected_category': selected_category,
                'category': selected_category,
                'next_serial': next_serial,
                'locations': locations,
                'stream': stream,
                'selected_stream': stream,
                'product': {
                    'name': name,
                    'serial_number': serial_number,
                    'description': description,
                    'issue_description': issue_description,
                    'twelve_nc': twelve_nc,
                    'status': status,
                    'handover_team_type': handover_team_type,
                    'handover_external_team': handover_external_team,
                    'handover_owner': handover_owner,
                    'location': location,
                },
                'edit': False
            })
        product = Product.objects.create(
            name=name,
            serial_number=serial_number,
            description=description,
            issue_description=issue_description,
            twelve_nc=twelve_nc,
            category=category,
            status=status,
            handover_team_type=handover_team_type if status == 'Hand-Overed' else '',
            handover_external_team=handover_external_team if status == 'Hand-Overed' and handover_team_type == 'External' else '',
            handover_owner=handover_owner if status == 'Hand-Overed' else '',
            location=location,
            created_by=request.user,
            updated_by=request.user,
            stream=stream_obj
        )
        # Handle multiple image uploads
        images = request.FILES.getlist('images')
        from .models import ProductImage
        for img in images:
            ProductImage.objects.create(product=product, image=img)
        ProductHistory.objects.create(
            product=product,
            action='created',
            user=request.user,
            details=f'Product created with name: {name}, serial: {serial_number}'
        )
        # Redirect to product list for the stream, optionally with category as a query param
        url = reverse('product_list_stream', kwargs={'stream': product.stream})
        if product.category:
            url += f'?category={product.category.pk}'
        return redirect(url)
    return render(request, 'products/product_form.html', {
        'categories': categories,
        'selected_category': selected_category,
        'category': selected_category,  # ensure template always gets 'category'
        'next_serial': next_serial,
        'locations': locations,
        'stream': stream,
        'selected_stream': stream,
        'product': None,
        'edit': False
    })

@login_required
def product_detail(request, stream, pk):
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Resolve stream name to Stream object (404 if not found)
    stream_obj = get_stream_or_404(stream)

    product = get_object_or_404(Product, pk=pk, stream=stream_obj)
    # Generate QR code for the product detail URL
    from django.urls import reverse
    # Build absolute URL to product detail page
    product_url = request.build_absolute_uri(reverse('product_detail_stream', args=[stream, product.pk]))
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(product_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_image = base64.b64encode(buffer.getvalue()).decode()
    history = product.history.order_by('-timestamp')
    return render(request, 'products/product_detail.html', {'product': product, 'qr_image': qr_image, 'history': history, 'stream': stream, 'selected_stream': stream})

def user_register(request):
    # Get available streams for selection
    available_streams = Stream.objects.filter(is_active=True, allow_public_registration=True)
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password')
        selected_streams = request.POST.getlist('streams')  # Multiple streams can be selected
        
        if not email.endswith('@philips.com'):
            messages.error(request, 'Email must be a @philips.com address.')
            return render(request, 'products/register.html', {'available_streams': available_streams})
            
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'products/register.html', {'available_streams': available_streams})
            
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered.')
            return render(request, 'products/register.html', {'available_streams': available_streams})
            
        if not selected_streams:
            messages.error(request, 'Please select at least one stream.')
            return render(request, 'products/register.html', {'available_streams': available_streams})
        try:
            # Create user as inactive (pending approval)
            user = User.objects.create_user(username=username, email=email, password=password)
            user.is_active = False
            user.save()
            
            # Create custom user profile and add requested streams
            custom_user = CustomUser.objects.create(user=user)
            for stream_id in selected_streams:
                try:
                    stream = Stream.objects.get(id=stream_id)
                    custom_user.requested_streams.add(stream)
                except Stream.DoesNotExist:
                    continue
        except Exception as e:
            messages.error(request, f'Error during registration: {str(e)}')
            return render(request, 'products/register.html', {'available_streams': available_streams})
        
        messages.success(request, 'Registration successful! Please wait for admin approval.')
        return redirect('login')
        
    return render(request, 'products/register.html', {'available_streams': available_streams})

def user_login(request):
    access_error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if not user.is_active:
                messages.error(request, 'Your account is pending admin approval.')
                return render(request, 'products/login.html')
            # Check access using the user object, not request.user
            custom_profile, created = CustomUser.objects.get_or_create(user=user)
            if not user.is_superuser and not custom_profile.user_roles.exists():
                access_error = 'Access denied. You have no assigned roles. Please contact an administrator.'
                return render(request, 'products/login.html', {'access_error': access_error})
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
            return render(request, 'products/login.html')
    return render(request, 'products/login.html')

def user_logout(request):
    # Clear chat history from session before logout
    if 'chat_history' in request.session:
        request.session.modified = True
    
    logout(request)
    return redirect('home')

@login_required
def promote_user(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'Only admins can promote users.')
        return redirect('product_list')
    user = get_object_or_404(User, id=user_id)
    user.is_staff = True
    user.is_superuser = True
    user.save()
    messages.success(request, f'{user.username} changed to admin.')
    return redirect('user_list')

@login_required
def depromote_user(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'Only super admins can depromote users.')
        return redirect('user_list')
    User = get_user_model()
    user = get_object_or_404(User, id=user_id)
    if user == request.user:
        messages.error(request, 'You cannot depromote yourself.')
        return redirect('user_list')
    user.is_superuser = False    
    user.is_staff = False
    user.save()
    messages.success(request, f'User {user.username} has been depromoted from admin.')
    return redirect('user_list')

@login_required
@require_POST
def change_user_role(request, user_id):
    if not can_manage_users(request.user):
        messages.error(request, 'Access denied. You need admin privileges to change user roles.')
        return redirect('user_list')
    
    user = get_object_or_404(User, id=user_id)
    new_role = request.POST.get('role')
    stream = request.POST.get('stream', 'HIC')
    
    # Prevent users from changing their own role (except super admins)
    if user == request.user and not is_super_admin(request.user):
        messages.error(request, 'You cannot change your own role.')
        return redirect('user_list')
    
    # Only super admins can create other super admins
    if new_role == 'super_admin' and not is_super_admin(request.user):
        messages.error(request, 'Only Super Admins can assign Super Admin role.')
        return redirect('user_list')
    
    # Get or create custom profile
    custom_profile, created = CustomUser.objects.get_or_create(user=user)
    old_role = custom_profile.role
    
    # Update role and stream
    custom_profile.role = new_role
    custom_profile.stream = stream
    custom_profile.save()
    
    # Update Django's built-in permissions for backwards compatibility
    if new_role in ['admin', 'super_admin']:
        user.is_staff = True
        if new_role == 'super_admin':
            user.is_superuser = True
    else:
        user.is_staff = False
        user.is_superuser = False
    
    user.save()
    
    role_display = dict(CustomUser.ROLE_CHOICES).get(new_role, new_role)
    messages.success(request, f'{user.username} role changed from {dict(CustomUser.ROLE_CHOICES).get(old_role, old_role)} to {role_display}.')
    return redirect('user_list')

@login_required
def manage_user_roles_and_streams(request, user_id):
    """View for super admins to manage user roles and stream access"""
    if not is_super_admin(request.user):
        messages.error(request, 'Access denied. Only Super Admins can manage user roles and streams.')
        return redirect('user_list')
    
    target_user = get_object_or_404(User, id=user_id)
    custom_profile, created = CustomUser.objects.get_or_create(user=target_user)
    
    if request.method == 'POST':
        # Handle role updates
        selected_roles = request.POST.getlist('roles')
        if target_user != request.user:  # Prevent self-modification unless super admin
            custom_profile.roles.clear()
            for role_name in selected_roles:
                try:
                    role = UserRole.objects.get(name=role_name)
                    custom_profile.roles.add(role)
                except UserRole.DoesNotExist:
                    pass
        
        # Handle stream access updates
        stream_ids = request.POST.getlist('streams')
        can_write_streams = request.POST.getlist('can_write')
        can_delete_streams = request.POST.getlist('can_delete')
        
        # Clear existing stream access
        UserStreamAccess.objects.filter(user=target_user).delete()
        
        # Add new stream access
        for stream_id in stream_ids:
            try:
                stream = Stream.objects.get(id=stream_id)
                UserStreamAccess.objects.create(
                    user=target_user,
                    stream=stream,
                    can_read=True,
                    can_write=stream_id in can_write_streams,
                    can_delete=stream_id in can_delete_streams,
                    granted_by=request.user
                )
            except Stream.DoesNotExist:
                pass
        
        messages.success(request, f'Updated roles and stream access for {target_user.username}')
        return redirect('user_list')
    
    # Get all available roles and streams
    all_roles = UserRole.objects.all()
    all_streams = Stream.objects.filter(is_active=True)
    user_roles = custom_profile.roles.all()
    user_stream_access = UserStreamAccess.objects.filter(user=target_user)
    
    return render(request, 'products/manage_user_access.html', {
        'target_user': target_user,
        'custom_profile': custom_profile,
        'all_roles': all_roles,
        'all_streams': all_streams,
        'user_roles': user_roles,
        'user_stream_access': user_stream_access,
    })

@login_required  
def user_registration_request(request):
    """Handle user registration with stream access requests"""
    if request.method == 'POST':
        # Update user's requested streams
        custom_profile, created = CustomUser.objects.get_or_create(user=request.user)
        requested_stream_ids = request.POST.getlist('requested_streams')
        
        custom_profile.requested_streams.clear()
        for stream_id in requested_stream_ids:
            try:
                stream = Stream.objects.get(id=stream_id, is_active=True, allow_public_registration=True)
                custom_profile.requested_streams.add(stream)
            except Stream.DoesNotExist:
                pass
        
        messages.success(request, 'Your stream access requests have been submitted for approval.')
        return redirect('dashboard')
    
    available_streams = Stream.objects.filter(is_active=True, allow_public_registration=True)
    custom_profile, created = CustomUser.objects.get_or_create(user=request.user)
    
    return render(request, 'products/request_stream_access.html', {
        'available_streams': available_streams,
        'custom_profile': custom_profile,
    })

@login_required
def analytics_dashboard(request):
    if not request.user.is_authenticated:
        return render(request, 'products/please_login.html')
    # Get selected stream from GET params
    selected_stream = request.GET.get('stream', '')
    
    # Handle stream object conversion
    stream_obj = None
    if selected_stream:
        stream_obj = get_stream_or_404(selected_stream, default='HIC')
    
    # 1. Category usage (most/least used) - filter by stream if selected
    if stream_obj:
        category_usage = (
            Category.objects.filter(stream=stream_obj)
            .annotate(product_count=Count('products'))
            .order_by('-product_count')
        )
        total_products = Product.objects.filter(stream=stream_obj).count()
    else:
        category_usage = (
            Category.objects.annotate(product_count=Count('products'))
            .order_by('-product_count')
        )
        total_products = Product.objects.count()
    most_used = category_usage.first()
    least_used = category_usage.last()    # 2. Product growth over time (monthly trend) - filter by stream
    if stream_obj:
        products_by_month = (
            Product.objects.filter(stream=stream_obj)
            .extra({'month': "strftime('%%Y-%%m', created_at)"})
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )
    else:
        products_by_month = (
            Product.objects.extra({'month': "strftime('%%Y-%%m', created_at)"})
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )
    months = [row['month'] for row in products_by_month]
    counts = [row['count'] for row in products_by_month]
    growth_trend = {'months': months, 'counts': counts}

    # 3. Recommendations for merging/splitting categories
    counts_list = [cat.product_count for cat in category_usage]
    if counts_list:
        mean = np.mean(counts_list)
        std = np.std(counts_list)
        merge_candidates = [cat for cat in category_usage if cat.product_count < mean - std]
        split_candidates = [cat for cat in category_usage if cat.product_count > mean + std]
    else:
        merge_candidates = []
        split_candidates = []

    recommendations = {
        'merge': [cat.name for cat in merge_candidates],
        'split': [cat.name for cat in split_candidates],
    }

    # Streams for filter dropdown
    streams = list(Stream.objects.values_list('name', flat=True).order_by('name'))

    # Prepare JSON for charts
    category_usage_json = json.dumps([{'name': cat.name, 'product_count': cat.product_count} for cat in category_usage])
    months_json = json.dumps(months)
    counts_json = json.dumps(counts)    # --- Sub Level Distribution (Products) ---
    def get_sublevel_product_count(sub):
        return (sub.in_stock or 0) + (sub.in_use or 0) + (sub.scraped or 0)
    if stream_obj:
        sublevel_dist = SubLevel.objects.filter(stream=stream_obj)
    else:
        sublevel_dist = SubLevel.objects.all()
    sublevel_dist_json = json.dumps([
        {'name': sub.name, 'product_count': get_sublevel_product_count(sub)} for sub in sublevel_dist
    ])    # --- Sub Level Growth Trend ---
    from collections import Counter
    if stream_obj:
        sublevel_histories = SubLevelHistory.objects.filter(sublevel__stream=stream_obj)
    else:
        sublevel_histories = SubLevelHistory.objects.all()
    created_histories = [h for h in sublevel_histories if h.action == 'Created']
    def format_month(dt):
        return dt.strftime('%Y-%m')
    month_counts = Counter(format_month(h.at) for h in created_histories)
    months_sorted = sorted(month_counts.keys())
    sublevel_growth_months_json = json.dumps(months_sorted)
    sublevel_growth_counts_json = json.dumps([month_counts[m] for m in months_sorted])

    return render(request, 'products/analytics_dashboard.html', {
        'category_usage': category_usage,
        'most_used': most_used,
        'least_used': least_used,
        'growth_trend': growth_trend,
        'recommendations': recommendations,
        'total_products': total_products,
        'streams': streams,
        'selected_stream': selected_stream,
        'category_usage_json': category_usage_json,
        'months_json': months_json,
        'counts_json': counts_json,
        'sublevel_dist_json': sublevel_dist_json,
        'sublevel_growth_months_json': sublevel_growth_months_json,
        'sublevel_growth_counts_json': sublevel_growth_counts_json,
    })

@login_required
def user_list(request):
    if not can_manage_users(request.user):
        messages.error(request, 'Access denied. You need admin privileges to view user list.')
        return redirect('dashboard')
    users = User.objects.filter(is_active=True)
    pending_users = User.objects.filter(is_active=False)
    participants = Participant.objects.all()
    streams = Stream.objects.all()
    selected_stream = request.GET.get('stream', 'HIC')
    # Get backups for UI
    user_backups = UserDataVersion.objects.order_by('-created_at')
    from .models import StreamDeletionHistory
    stream_deletion_history = StreamDeletionHistory.objects.select_related('deleted_by').order_by('-deleted_at')[:100]
    
    # Ensure all users (including pending) have custom profiles and get their requested streams
    for user in list(users) + list(pending_users):
        custom_profile, created = CustomUser.objects.get_or_create(user=user)
    
    # Get the requested streams for pending users
    pending_users_data = []
    for user in pending_users:
        custom_profile = CustomUser.objects.get(user=user)
        requested_streams = ", ".join([stream.name for stream in custom_profile.requested_streams.all()])
        pending_users_data.append({
            'user': user,
            'requested_streams': requested_streams
        })
    
    return render(request, 'products/user_list.html', {
        'users': users,
        'pending_users_data': pending_users_data,
        'participants': participants,
        'streams': streams,
        'selected_stream': selected_stream,
        'stream': selected_stream,
        'user_backups': user_backups,
        'stream_deletion_history': stream_deletion_history,
    })

@login_required
@require_POST
def approve_user(request, user_id):
    """Activate a pending user account."""
    # Permission check
    if not can_manage_users(request.user):
        messages.error(request, 'Access denied. You need admin privileges to approve users.')
        return redirect('user_list')
    
    user = get_object_or_404(User, pk=user_id)
    if user.is_active:
        messages.warning(request, f'User {user.username} is already active.')
        return redirect('user_list')
        
    try:
        # Get or create custom user profile
        custom_user, created = CustomUser.objects.get_or_create(user=user)
        
        # Get requested streams before changes
        requested_streams = list(custom_user.requested_streams.all())
        
        # Create basic user role
        UserRole.objects.get_or_create(custom_user=custom_user, role='user')
        
        # Grant access to requested streams
        for stream in requested_streams:
            UserStreamAccess.objects.get_or_create(
                custom_user=custom_user,
                stream=stream
            )
        
        # Clear requested streams after granting access
        custom_user.requested_streams.clear()
        
        # Activate the user
        user.is_active = True
        user.save()
        
        # Log success
        messages.success(request, f'User "{user.username}" has been approved and activated with access to their requested streams.')
        logger.info('User approved: %s by %s with streams: %s', 
                   user.username, request.user.username, 
                   ', '.join(s.name for s in requested_streams))
        
    except Exception as e:
        logger.exception('Error approving user %s: %s', user_id, e)
        messages.error(request, f'An error occurred while approving the user: {str(e)}')
    
    return redirect('user_list')


@login_required
@require_POST
def decline_user(request, user_id):
    """Decline a pending user registration by removing the user record.
    Only applies to users who are not yet active.
    """
    # Permission check
    if not can_manage_users(request.user):
        messages.error(request, 'Permission denied. You do not have rights to decline users.')
        return redirect('user_list')

    user = get_object_or_404(User, pk=user_id)
    if user.is_active:
        messages.error(request, 'Cannot decline an already active user.')
        return redirect('user_list')

    username = user.username
    try:
        user.delete()
        messages.success(request, f'Pending user "{username}" has been declined and removed.')
        logger.info('User declined and deleted: %s by %s', username, request.user.username)
    except Exception as e:
        logger.exception('Error declining user %s: %s', user_id, e)
        messages.error(request, 'An error occurred while declining the user. Please try again.')

    return redirect('user_list')

@login_required
def product_edit(request, pk, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    product = get_object_or_404(Product, pk=pk, stream=stream_obj)
    categories = Category.objects.filter(stream=product.stream).order_by('name')
    locations = Location.objects.filter(stream=product.stream).order_by('name')
    if request.method == 'POST':        
        old_name = product.name
        old_serial = product.serial_number
        old_description = product.description
        old_category = product.category
        old_status = getattr(product, 'status', 'Active')
        old_handover_team_type = getattr(product, 'handover_team_type', '')
        old_handover_external_team = getattr(product, 'handover_external_team', '')
        old_handover_owner = getattr(product, 'handover_owner', '')
        old_location = product.location
        old_issue_description = getattr(product, 'issue_description', '')
        old_twelve_nc = getattr(product, 'twelve_nc', '')

        issue_description = request.POST.get('issue_description', '')
        twelve_nc = request.POST.get('twelve_nc', '')

        # Update product fields
        product.name = request.POST.get('name')
        product.serial_number = request.POST.get('serial_number')
        product.description = request.POST.get('description')
        product.issue_description = issue_description
        product.twelve_nc = twelve_nc
        category_id = request.POST.get('category')
        try:
            product.category = Category.objects.get(pk=category_id, stream=product.stream) if category_id else None
        except Exception:
            product.category = None

        product.status = request.POST.get('status', 'Active')
        product.handover_team_type = request.POST.get('handover_team_type', '') if product.status == 'Hand-Overed' else ''
        product.handover_external_team = request.POST.get('handover_external_team', '') if product.status == 'Hand-Overed' and request.POST.get('handover_team_type', '') == 'External' else ''
        product.handover_owner = request.POST.get('handover_owner', '') if product.status == 'Hand-Overed' else ''
        location_id = request.POST.get('location')
        product.location = Location.objects.filter(id=location_id, stream=product.stream).first() if location_id else None
        product.updated_by = request.user

        # Prepare changed fields only
        changes = []
        if old_name != product.name:
            changes.append(f"Name: '{old_name}' → '{product.name}'")
        if old_serial != product.serial_number:
            changes.append(f"Serial NO: '{old_serial}' → '{product.serial_number}'")
        if old_description != product.description:
            changes.append(f"Desc: '{old_description}' → '{product.description}'")
        if old_issue_description != issue_description:
            changes.append(f"Issue Desc: '{old_issue_description}' → '{issue_description}'")
        if old_twelve_nc != product.twelve_nc:
            changes.append(f"12NC: '{old_twelve_nc}' → '{product.twelve_nc}'")
        if old_category != product.category:
            changes.append(f"Category: '{old_category}' → '{product.category}'")
        if old_status != product.status:
            changes.append(f"Status: '{old_status}' → '{product.status}'")
        if old_handover_team_type != product.handover_team_type:
            changes.append(f"Handover_team_type: '{old_handover_team_type}' → '{product.handover_team_type}'")
        if old_handover_external_team != product.handover_external_team:
            changes.append(f"Handover_external_team: '{old_handover_external_team}' → '{product.handover_external_team}'")
        if old_handover_owner != product.handover_owner:
            changes.append(f"Handover_owner: '{old_handover_owner}' → '{product.handover_owner}'")
        if (old_location.name if old_location else None) != (product.location.name if product.location else None):
            changes.append(f"Location: '{old_location.name if old_location else 'None'}' → '{product.location.name if product.location else 'None'}'")

        product.save()

        if changes:
            ProductHistory.objects.create(
                product=product,
                action='edited',
                user=request.user,
                details="; ".join(changes)
            )

        # Handle multiple image uploads on edit (mirror create behavior)
        images = request.FILES.getlist('images')
        from .models import ProductImage
        for img in images:
            ProductImage.objects.create(product=product, image=img)

        # Redirect to product list for the stream, optionally with category as a query param
        url = reverse('product_list_stream', kwargs={'stream': stream or 'HIC'})
        if product.category:
            url += f'?category={product.category.pk}'
        return redirect(url)
    return render(request, 'products/product_form.html', {
        'product': product,
        'edit': True,
        'categories': categories,
        'selected_category': product.category,  # ensure category is pre-selected and locked
        'next_serial': product.serial_number,    # provide current serial for edit form
        'locations': locations,
        'stream': product.stream,
        'selected_stream': product.stream.name if hasattr(product.stream, 'name') else product.stream
    })

@login_required
def product_delete(request, pk, stream=None):
    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    if not can_delete_products(request.user):
        messages.error(request, 'Access denied. You need admin privileges to delete products.')
        return redirect('product_list_stream', stream=stream or 'HIC')
    
    try:
        product = Product.objects.get(pk=pk, stream=stream_obj)
    except Product.DoesNotExist:
        messages.warning(request, 'Product already deleted.')
        url = reverse('product_list_stream', kwargs={'stream': stream or 'HIC'})
        query_params = request.GET.urlencode()
        if query_params:
            url += '?' + query_params
        return redirect(url)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully')
        url = reverse('product_list_stream', kwargs={'stream': stream or 'HIC'})
        query_params = request.POST.get('query_params', '')
        if query_params:
            url += '?' + query_params
        return redirect(url)
    # GET: render confirmation page with query params
    query_params = request.GET.urlencode()
    return render(request, 'products/product_confirm_delete.html', {'product': product, 'stream': stream, 'query_params': query_params})

def parse_history_details(details):
    """
    Parse the details string and return a list of changed fields with old and new values.
    """
    # Example details string:
    # Edited from name: 23232, serial: 1002, desc: 32323, category: Test (1001), status: Active, handover_team_type: , handover_external_team: , handover_owner: to name: 23232, serial: 1002, desc: 32323, category: Test (1001), status: Scraped, handover_team_type: , handover_external_team: , handover_owner:
    match = re.match(r"Edited from (.+) to (.+)", details)
    if not match:
        return []
    from_part, to_part = match.groups()
    def parse_part(part):
        # Split by comma, then by colon
        fields = {}
        for item in part.split(','):
            if ':' in item:
                key, value = item.split(':', 1)
                fields[key.strip()] = value.strip()
        return fields
    from_fields = parse_part(from_part)
    to_fields = parse_part(to_part)
    changes = []
    for key in from_fields:
        old = from_fields[key]
        new = to_fields.get(key, '')
        if old != new:
            changes.append({'field': key, 'old': old, 'new': new})
    return changes

@login_required
def product_history_ajax(request, stream=None, pk=None):
    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    product = get_object_or_404(Product, pk=pk, stream=stream_obj)
    history = product.history.order_by('-timestamp')[:30]  # Limit to 30 most recent entries
    # Attach changes for edited actions
    for h in history:
        if h.action == 'edited':
            h.changes = parse_history_details(h.details)
        else:
            h.changes = None
    html = render_to_string('products/history_snippet.html', {'history': history})
    return HttpResponse(html)

@login_required
def category_list(request, stream=None):
    q = request.GET.get('q', '').strip()
    sort = request.GET.get('sort', '-created_at')
    page_number = request.GET.get('page', 1)
    stream = stream or request.GET.get('stream', 'HIC')
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'PIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Get the Stream object for database queries (404 if not found)
    stream_obj = get_stream_or_404(stream, default='HIC')

    # Base queryset for the selected stream
    categories_qs = Category.objects.filter(stream=stream_obj)

    # Apply search (name or serial_number)
    if q:
        categories_qs = categories_qs.filter(
            Q(name__icontains=q) | Q(serial_number__icontains=q)
        )

    # Annotate with product count for badges and sorting
    categories_qs = categories_qs.annotate(product_count=Count('products'))

    # Sorting options
    if sort == 'name':
        categories_qs = categories_qs.order_by('name')
    elif sort == 'product_count':
        categories_qs = categories_qs.order_by('-product_count')
    else:
        categories_qs = categories_qs.order_by('-created_at')

    # Pagination
    paginator = Paginator(categories_qs, 24)  # 24 cards per page
    page_obj = paginator.get_page(page_number)

    streams = list(Stream.objects.values_list('name', flat=True).order_by('name'))

    return render(request, 'products/category_list.html', {
        'categories': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'streams': streams,
        'selected_stream': stream,
        'selected_sort': sort,
        'q': q,
    })

@login_required
def category_create(request, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    if not request.user.is_superuser:
        return redirect('category_list_stream', stream=stream or 'HIC')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        serial_number = request.POST.get('serial_number', '').strip()
        if not name or not serial_number:
            messages.warning(request, 'Both Category Name and Serial Number are required.')
            return render(request, 'products/category_form.html', {'category': {'name': name, 'serial_number': serial_number}, 'stream': stream, 'selected_stream': stream})
        # Get Stream object from name (404 if not found). Default to 'HIC' when missing.
        try:
            stream_obj = get_stream_or_404(stream, default='HIC')
        except Exception:
            messages.warning(request, 'Invalid stream specified.')
            return render(request, 'products/category_form.html', {'category': {'name': name, 'serial_number': serial_number}, 'stream': stream, 'selected_stream': stream})
        if Category.objects.filter(name=name, stream=stream_obj).exists() or Category.objects.filter(serial_number=serial_number, stream=stream_obj).exists():
            messages.warning(request, 'Category with this name or serial number already exists in this stream.')
            return render(request, 'products/category_form.html', {'category': {'name': name, 'serial_number': serial_number}, 'stream': stream, 'selected_stream': stream})
        try:
            Category.objects.create(
                name=name,
                serial_number=serial_number,
                created_by=request.user,
                stream=stream_obj
            )
            return redirect('category_list_stream', stream=stream or 'HIC')
        except IntegrityError:
            messages.warning(request, 'Category with this serial number already exists. Please use a unique serial number.')
            return render(request, 'products/category_form.html', {'category': {'name': name, 'serial_number': serial_number}, 'stream': stream, 'selected_stream': stream})
    return render(request, 'products/category_form.html', {'stream': stream, 'selected_stream': stream})

@login_required
def category_edit(request, pk, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Resolve stream param to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    category = get_object_or_404(Category, pk=pk, stream=stream_obj)
    if request.method == 'POST':
        name = request.POST.get('name')
        serial_number = request.POST.get('serial_number')
        if Category.objects.filter(name=name, stream=stream_obj).exclude(pk=pk).exists() or Category.objects.filter(serial_number=serial_number, stream=stream_obj).exclude(pk=pk).exists():
            messages.error(request, 'Category with this name or serial number already exists.')
            return render(request, 'products/category_form.html', {'category': category, 'edit': True, 'stream': stream, 'selected_stream': stream})
        category.name = name
        category.serial_number = serial_number
        try:
            category.save()
        except IntegrityError:
            messages.error(request, 'A category with this name or serial number already exists (database constraint).')
            return render(request, 'products/category_form.html', {'category': category, 'edit': True, 'stream': stream, 'selected_stream': stream})
        return redirect('category_list_stream', stream=stream or 'HIC')
    return render(request, 'products/category_form.html', {'category': category, 'edit': True, 'stream': stream, 'selected_stream': stream})

@login_required
def category_delete(request, pk, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    try:
        category = Category.objects.get(pk=pk, stream=stream_obj)
    except Category.DoesNotExist:
        messages.warning(request, 'Category already deleted.')
        return redirect('category_list_stream', stream=stream or 'HIC')
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Category deleted successfully')
        return redirect('category_list_stream', stream=stream or 'HIC')
    return render(request, 'products/category_confirm_delete.html', {'category': category, 'stream': stream})

@login_required
def post_login_landing(request):
    """
    After login, show only the category management page (no products).
    """
    categories = Category.objects.all().order_by('-created_at')
    return render(request, 'products/category_list.html', {'categories': categories})

@require_POST
@login_required
def remove_user(request, user_id):
    if not request.user.is_superuser:
        messages.error(request, 'Only super admins can remove users.')
        return redirect('user_list')
    User = get_user_model()
    user = get_object_or_404(User, id=user_id)
    if user == request.user:
        messages.error(request, 'You cannot remove yourself.')
        return redirect('user_list')
    username = user.username
    user.delete()
    messages.success(request, f'User removed: {username} has been successfully removed from the system.')
    return redirect('user_list')

def download_qr_with_details(request, stream, pk):
    from django.urls import reverse
    from io import BytesIO
    from PIL import Image as PILImage, ImageDraw, ImageFont
    import qrcode
    from django.http import FileResponse
    from django.shortcuts import get_object_or_404
    from django.utils.text import slugify
    import os
    from django.conf import settings

    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    product = get_object_or_404(Product, pk=pk, stream=stream_obj)

    # Generate QR code for product detail URL
    product_url = request.build_absolute_uri(
        reverse('product_detail_stream', args=[stream, product.pk])
    )

    qr = qrcode.QRCode(box_size=4, border=1)
    qr.add_data(product_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    # Resize QR code to a very small size
    target_size = (80, 80)
    qr_img = qr_img.resize(target_size, PILImage.LANCZOS)

    # Texts
    header_text = "MoS Nexus"
    # Prefer showing the Category serial number; fall back to category name if missing
    cat_name = product.category.name if product.category and getattr(product.category, 'name', '') else ''
    cat_serial = product.category.serial_number if product.category and getattr(product.category, 'serial_number', '') else ''
    details = f"Category: {cat_name}{(' (' + cat_serial + ')') if cat_name and cat_serial else ((' (' + cat_serial + ')') if cat_serial else '')}\nProduct Name: {product.name}"
    serial_text = f"{product.serial_number}"

    # Fonts
    try:
        header_font = ImageFont.truetype("arial.ttf", 20)
        font = ImageFont.truetype("arial.ttf", 14)
        serial_font = ImageFont.truetype("arial.ttf", 13.4)
    except:
        header_font = ImageFont.load_default()
        font = ImageFont.load_default()
        serial_font = ImageFont.load_default()

    # Measure text
    temp_draw = ImageDraw.Draw(qr_img)
    try:
        header_bbox = temp_draw.textbbox((0, 0), header_text, font=header_font)
        header_width = header_bbox[2] - header_bbox[0]
        header_height = header_bbox[3] - header_bbox[1]

        details_bbox = temp_draw.multiline_textbbox((0, 0), details, font=font)
        details_width = details_bbox[2] - details_bbox[0]
        details_height = details_bbox[3] - details_bbox[1]

        serial_bbox = temp_draw.textbbox((0, 0), serial_text, font=serial_font)
        serial_width = serial_bbox[2] - serial_bbox[0]
        serial_height = serial_bbox[3] - serial_bbox[1]
    except AttributeError:
        header_width, header_height = temp_draw.textsize(header_text, font=header_font)
        details_width, details_height = temp_draw.textsize(details, font=font)
        serial_width, serial_height = temp_draw.textsize(serial_text, font=serial_font)

    # Layout
    padding = 20
    total_width = qr_img.width + details_width + 3 * padding
    total_height = max(
        qr_img.height + serial_height + 2 * padding,
        header_height + details_height + 4 * padding,
    )

    new_img = PILImage.new("RGB", (total_width, total_height), "white")
    draw = ImageDraw.Draw(new_img)

    # Header text (top-left)
    header_x = padding
    header_y = padding // 2
    draw.text((header_x, header_y), header_text, fill=(0, 0, 0), font=header_font)

    # QR code (right side, tiny)
    qr_x = total_width - qr_img.width - padding
    qr_y = (total_height - qr_img.height) // 2 - serial_height // 2
    new_img.paste(qr_img, (qr_x, qr_y))

    # Serial number (below tiny QR)
    serial_x = qr_x + (qr_img.width - serial_width) // 2
    serial_y = qr_y + qr_img.height + 4
    draw.text((serial_x, serial_y), serial_text, fill=(0, 0, 0), font=serial_font)

    # Product details (left side, below header)
    details_x = padding + 5
    details_y = header_y + header_height + 25
    draw.multiline_text(
        (details_x, details_y),
        details,
        fill=(34, 34, 34),
        font=font,
        spacing=8,
    )

    # Save to buffer
    buffer = BytesIO()
    new_img.save(buffer, format="PNG")
    buffer.seek(0)

    filename = f"qr_{slugify(product.serial_number)}.png"
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type="image/png",
    )

def get_user_product_data():
    # Helper to get the required data for export
    from .models import Product, Category
    data = []
    products = Product.objects.select_related('category', 'created_by', 'location').all().order_by('category__name', 'name')
    for idx, product in enumerate(products, 1):
        # Ensure stream is a plain string (openpyxl cannot write model instances)
        stream_val = ''
        if hasattr(product, 'stream') and product.stream is not None:
            # Prefer a .name attribute if Stream is a model with that field
            stream_val = getattr(product.stream, 'name', str(product.stream))
        data.append([
            idx,
            product.category.name if product.category else '',
            product.category.serial_number if product.category else '',
            product.name,
            product.serial_number,
            product.description,
            product.created_at.strftime('%Y-%m-%d %H:%M'),
            product.created_by.username if product.created_by else '',
            stream_val,
            product.location.name if product.location else '',
            product.location.address if product.location and hasattr(product.location, 'address') else '',
        ])
    return data

@login_required
def download_users_excel(request):
    if not request.user.is_superuser:
        return HttpResponse('Unauthorized', status=401)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory Data'
    headers = ['S. NO', 'Category name', 'Category Serial Number', 'Product Name', 'Product Serial Number', 'Product Description', 'Product Added date', 'Product added by', 'Stream', 'Location Name', 'Location Address']
    ws.append(headers)
    for row in get_user_product_data():
        ws.append(row)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_MoS_Inventory list.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def download_users_pdf(request):
    if not request.user.is_superuser:
        return HttpResponse('Unauthorized', status=401)
    from django.contrib.staticfiles import finders

    buffer = BytesIO()
    page_width, page_height = landscape(A3)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(page_width, page_height),
        leftMargin=30,
        rightMargin=30,
        topMargin=60,
        bottomMargin=30,
    )

    # Prepare data
    headers = [
        'S. NO', 'Category name', 'Category Serial Number', 'Product Name',
        'Product Serial Number', 'Product Description', 'Product Added date',
        'Product added by', 'Stream', 'Location Name', 'Location Address'
    ]
    data = [headers] + get_user_product_data()
    table = Table(data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005fa3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ])
    table.setStyle(style)

    # Styles for titles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.fontName = 'Helvetica-Bold'
    title_style.fontSize = 22
    title_style.alignment = 1  # Center

    subtitle_style = styles['Heading2']
    subtitle_style.fontName = 'Helvetica-Bold'
    subtitle_style.fontSize = 18
    subtitle_style.alignment = 1

    subsubtitle_style = styles['Heading3']
    subsubtitle_style.fontName = 'Helvetica-Bold'
    subsubtitle_style.fontSize = 16
    subsubtitle_style.alignment = 1

    note_style = styles['Italic']
    note_style.fontSize = 12
    note_style.alignment = 1

    # Logo path
    logo_path = finders.find('products/philips.png')

    def draw_header(canvas, doc):
        # Draw logo
        if logo_path:
            logo_width = 50
            logo_height = 50
            x_logo = page_width - logo_width - 40
            y_logo = page_height - logo_height - 20
            canvas.drawImage(logo_path, x_logo, y_logo, width=logo_width, height=logo_height, mask='auto')
        # Draw titles
        canvas.setFont("Helvetica-Bold", 22)
        canvas.drawCentredString(page_width / 2, page_height - 60, "Image Guided Therapy (IGT)")
        canvas.setFont("Helvetica-Bold", 18)
        canvas.drawCentredString(page_width / 2, page_height - 90, "Mobile Surgery (MoS)")
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(page_width / 2, page_height - 120, "Inventory Data")
        canvas.setFont("Helvetica-Oblique", 12)
        canvas.drawCentredString(page_width / 2, page_height - 145, "(Note: Automated data output. Verification recommended to ensure reliability and compliance with organizational protocols.)")

    elements = [Spacer(1, 120), table]  # Reduced gap above the table

    # --- SubLevel Data Table ---
    from .models import SubLevel
    sublevel_headers = ['Item Name', 'In stock', 'In use', 'Scrapped', 'Stream', 'Last modified by', 'Last modified date']
    sublevel_data = [sublevel_headers]
    sublevels = SubLevel.objects.all()
    for sub in sublevels:
        last_history = sub.history.order_by('-at').first()
        last_by = last_history.by if last_history else ''
        last_at = last_history.at.strftime('%Y-%m-%d %H:%M:%S') if last_history else ''
        sublevel_data.append([
            sub.name,
            sub.in_stock,
            sub.in_use,
            sub.scraped,
            sub.stream or '',
            last_by,
            last_at
        ])
    if len(sublevel_data) > 1:
        sublevel_table = Table(sublevel_data, repeatRows=1)
        sublevel_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005fa3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ])
        sublevel_table.setStyle(sublevel_style)
        # Set column widths for SubLevel Data table
        col_widths = [130, 90, 90, 90, 100, 120, 140] 
        sublevel_table._argW = col_widths
        elements.append(Spacer(1, 40))
        from reportlab.platypus import Paragraph
        elements.append(Paragraph('SubLevel Data', title_style))
        elements.append(Spacer(1, 10))
        elements.append(sublevel_table)

    doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header)

    pdf = buffer.getvalue()
    buffer.close()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_MoS_Inventory list.pdf"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response

@login_required
@require_POST
def upload_products_excel(request):
    if not request.user.is_superuser:
        return HttpResponse('Unauthorized', status=401)
    # --- BACKUP LOGIC ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory Data'
    headers = ['S. NO', 'Category name', 'Category Serial Number', 'Product Name', 'Product Serial Number', 'Product Description', 'Product Added date', 'Product added by', 'Stream']
    ws.append(headers)
    for idx, product in enumerate(Product.objects.all(), start=1):
        # Ensure we write a plain string for stream (don't pass Stream model instances to openpyxl)
        stream_val = ''
        if hasattr(product, 'stream') and product.stream is not None:
            try:
                stream_val = product.stream.name
            except Exception:
                stream_val = str(product.stream)
        ws.append([
            idx,
            product.category.name if product.category else '',
            product.category.serial_number if product.category else '',
            product.name,
            product.serial_number,
            product.description,
            product.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            product.created_by.username if product.created_by else '',
            stream_val,
        ])
    backup_stream = BytesIO()
    wb.save(backup_stream)
    backup_stream.seek(0)
    # --- Version number logic ---
    last_backup = UserDataVersion.objects.order_by('-created_at').first()
    if last_backup and hasattr(last_backup, 'version_str'):
        # Parse previous version_str
        import re
        m = re.match(r'v(\d+)\.(\d+)\.(\d+)\.(\d+)', last_backup.version_str)
        if m:
            major, minor, patch, build = map(int, m.groups())
            build += 1
            if build > 3:
                patch += 1
                build = 1
            if patch > 9:
                minor += 1
                patch = 0
            if minor > 9:
                major += 1
                minor = 0
            version_str = f"v{major}.{minor}.{patch}.{build}"
        else:
            version_str = "v1.0.0.0"
    else:
        version_str = "v1.0.0.0"
    backup_file = ContentFile(backup_stream.read(), name=f"user_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    UserDataVersion.objects.create(
        created_by=request.user,
        description=f"Backup before upload on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        data_file=backup_file,
        version_str=version_str
    )
    # --- EXISTING UPLOAD LOGIC ---
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'No file uploaded.')
        return redirect('user_list')
    try:
        df = pd.read_excel(excel_file)
    except Exception as e:
        messages.error(request, f'Error reading Excel file: {e}')
        return redirect('user_list')
    required_columns = [
        'S. NO', 'Category name', 'Category Serial Number', 'Product Name', 'Product Serial Number', 'Product Description', 'Product Added date', 'Product added by', 'Stream'
    ]
    for col in required_columns:
        if col not in df.columns:
            messages.error(request, f'Missing column: {col}')
            return redirect('user_list')
    mode = request.GET.get('mode', 'overwrite')
    imported_count = 0
    if mode == 'overwrite':
        Product.objects.all().delete()
        Category.objects.all().delete()
    elif mode == 'refresh':
        Product.objects.all().delete()
        Category.objects.all().delete()
        Location.objects.all().delete()
        Participant.objects.all().delete()
    for _, row in df.iterrows():
        cat_name = str(row['Category name']).strip()
        cat_serial = str(row['Category Serial Number']).strip()
        prod_name = str(row['Product Name']).strip()
        prod_serial = str(row['Product Serial Number']).strip()
        prod_desc = str(row['Product Description']).strip()
        prod_added_by = str(row['Product added by']).strip()
        # Resolve stream name from uploaded data to a Stream instance (avoid storing raw strings into FK fields)
        prod_stream_name = 'HIC'
        if 'Stream' in df.columns and pd.notna(row['Stream']) and str(row['Stream']).strip():
            prod_stream_name = str(row['Stream']).strip()
        prod_stream_obj, _ = Stream.objects.get_or_create(name=prod_stream_name)
        category, created_cat = Category.objects.get_or_create(
            name=cat_name,
            defaults={'serial_number': cat_serial, 'created_by': request.user, 'stream': prod_stream_obj}
        )
        # If category exists, update serial_number if needed (avoid IntegrityError)
        if not created_cat:
            if category.serial_number != cat_serial:
                # Show warning if serial_number is different and already exists for another category
                if Category.objects.filter(serial_number=cat_serial).exclude(pk=category.pk).exists():
                    messages.warning(request, f"Category '{cat_name}' already exists with a different serial number. Skipped row due to duplicate serial number.")
                    continue
                category.serial_number = cat_serial
                category.save()
        if hasattr(category, 'stream') and category.stream != prod_stream_obj:
            category.stream = prod_stream_obj
            category.save()
        user_obj = User.objects.filter(username=prod_added_by).first() or request.user
        product, created = Product.objects.get_or_create(
            serial_number=prod_serial,
            defaults={
                'name': prod_name,
                'category': category,
                'description': prod_desc,
                'created_by': user_obj,
                'updated_by': user_obj,
                'stream': prod_stream_obj
            }
        )
        if not created and mode == 'append':
            product.name = prod_name
            product.category = category
            product.description = prod_desc
            product.updated_by = user_obj
            product.stream = prod_stream_obj
            product.save()
        imported_count += 1
    messages.success(request, f'Successfully imported {imported_count} products.')
    return redirect('user_list')

@login_required
@require_POST
def restore_user_backup(request, backup_id):
    if not request.user.is_superuser:
        return HttpResponse('Unauthorized', status=401)
    backup = get_object_or_404(UserDataVersion, id=backup_id)
    # Restore logic: read backup Excel and overwrite current data
    try:
        df = pd.read_excel(backup.data_file)
    except Exception as e:
        messages.error(request, f'Error reading backup file: {e}')
        return redirect('user_list')
    required_columns = [
        'S. NO', 'Category name', 'Category Serial Number', 'Product Name', 'Product Serial Number', 'Product Description', 'Product Added date', 'Product added by', 'Stream'
    ]
    for col in required_columns:
        if col not in df.columns:
            messages.error(request, f'Missing column in backup: {col}')
            return redirect('user_list')
    # Optionally clear current data
    Product.objects.all().delete()
    Category.objects.all().delete()
    imported_count = 0
    for _, row in df.iterrows():
        cat_name = str(row['Category name']).strip()
        cat_serial = str(row['Category Serial Number']).strip()
        prod_name = str(row['Product Name']).strip()
        prod_serial = str(row['Product Serial Number']).strip()
        prod_desc = str(row['Product Description']).strip()
        prod_added_by = str(row['Product added by']).strip()
        # Normalize stream value and convert to Stream instance
        prod_stream_name = 'HIC'
        if 'Stream' in df.columns and pd.notna(row['Stream']) and str(row['Stream']).strip():
            prod_stream_name = str(row['Stream']).strip()
        prod_stream_obj, _ = Stream.objects.get_or_create(name=prod_stream_name)
        category, created_cat = Category.objects.get_or_create(
            name=cat_name,
            defaults={'serial_number': cat_serial, 'created_by': request.user, 'stream': prod_stream_obj}
        )
        # If category exists, update serial_number if needed (avoid IntegrityError)
        if not created_cat:
            if category.serial_number != cat_serial:
                if Category.objects.filter(serial_number=cat_serial).exclude(pk=category.pk).exists():
                    messages.warning(request, f"Category '{cat_name}' already exists with a different serial number. Skipped row due to duplicate serial number.")
                    continue
                category.serial_number = cat_serial
                category.save()
        user_obj = User.objects.filter(username=prod_added_by).first() or request.user
        product, created = Product.objects.get_or_create(
            serial_number=prod_serial,
            defaults={
                'name': prod_name,
                'category': category,
                'description': prod_desc,
                'created_by': user_obj,
                'updated_by': user_obj,
                'stream': prod_stream_obj
            }
        )
        if not created:
            product.name = prod_name
            product.category = category
            product.description = prod_desc
            product.updated_by = user_obj
            product.stream = prod_stream_obj
            product.save()
        imported_count += 1
    messages.success(request, f'Restored {imported_count} products from backup.')
    return redirect('user_list')

@login_required
def system_allocation(request, stream=None):
    if not request.user.is_authenticated:
        return render(request, 'products/please_login.html')
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    if not can_manage_system_allocation(request.user):
        messages.error(request, 'Access denied. You need Lab Incharge or higher privileges to access system allocation.')
        return redirect('dashboard')
    
    combined_participants = {}
    if can_manage_users(request.user):
        users = list(User.objects.filter(is_active=True).values('id', 'username', 'email'))
        participants = list(Participant.objects.all().values('id', 'name', 'email'))
        combined_participants = json.dumps({
            'users': users,
            'participants': participants
        })
    
    # Get the view date from query parameters, default to today
    from django.utils import timezone
    from datetime import datetime
    view_date_str = request.GET.get('view_date')
    if view_date_str:
        try:
            view_date = datetime.strptime(view_date_str, '%Y-%m-%d').date()
        except ValueError:
            view_date = timezone.now().date()
    else:
        view_date = timezone.now().date()
    
    # Get systems with their metrics and current allocations
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    systems = System.objects.filter(stream=stream_obj).order_by('name')
    
    # Calculate utilization for each system and get historical status
    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)
    
    for system in systems:
        # Get historical status for the view_date
        # Convert view_date to timezone-aware datetime at the END of the day
        # This ensures we include all status changes that happened on view_date
        view_datetime_end = timezone.make_aware(
            datetime.combine(view_date, datetime.max.time())
        )
        
        # Find the most recent status change that occurred on or before the view_date
        # We compare the DATE part of updated_at with view_date
        historical_status = SystemStatusHistory.objects.filter(
            system=system,
            updated_at__isnull=False,  # Ensure updated_at is not None
            updated_at__date__lte=view_date  # Compare only the date part
        ).order_by('-updated_at').first()
        
        if historical_status:
            # Only use historical status if it was set on or before the view_date
            system.historical_status = historical_status.status
            system.historical_status_display = historical_status.get_status_display()
            system.historical_description = historical_status.description
            system.historical_assignee = historical_status.assignee
            system.historical_updated_by = historical_status.updated_by
            system.historical_updated_at = historical_status.updated_at
        else:
            # No historical record found before view_date, assume Active
            system.historical_status = 'Active'
            system.historical_status_display = 'Active'
            system.historical_description = ''
            system.historical_assignee = ''
            system.historical_updated_by = ''
            system.historical_updated_at = None
    
    for system in systems:
        # Calculate utilization based on recent allocations
        recent_allocations = SystemAllocation.objects.filter(
            system_type=system.name,
            start_date__gte=thirty_days_ago
        )
        
        # Calculate utilization based on recent allocations
        recent_allocations = SystemAllocation.objects.filter(
            system_type=system.name,
            start_date__gte=thirty_days_ago
        )
        
        total_hours = 0
        for allocation in recent_allocations:
            # Calculate hours for each allocation
            start = max(allocation.start_date, thirty_days_ago)
            end = min(allocation.end_date, now)
            if end > start:
                duration = end - start
                total_hours += duration.total_seconds() / 3600
        
        # Calculate utilization percentage based on business hours (8 hours/day, 5 days/week)
        # More realistic than 24/7 for most business systems
        business_days = 22  # Average business days per month
        business_hours_per_day = 8
        max_business_hours = business_days * business_hours_per_day  # ~176 hours per month
        
        # Calculate both business hours and 24/7 utilization for flexibility
        business_utilization = (total_hours / max_business_hours) * 100 if max_business_hours > 0 else 0
        full_time_utilization = (total_hours / (24 * 30)) * 100 if (24 * 30) > 0 else 0
        
        # Use business hours utilization as primary metric (more realistic)
        # Cap at 100% to prevent unrealistic values
        utilization = min(business_utilization, 100.0)
        
        # Update system utilization
        system.utilization_percentage = round(utilization, 1)
        system.save(update_fields=['utilization_percentage'])
        
        # Set health based on utilization and historical status
        if system.historical_status == 'Active':
            if utilization > 100:
                system.health = 'Critical'
            elif utilization > 80:
                system.health = 'Warning'
            elif utilization > 95:
                system.health = 'Warning'
            else:
                system.health = 'Excellent' if utilization < 50 else 'Good'
        else:
            system.health = 'Critical' if system.historical_status in ['Issue', 'Removed'] else 'Warning'
        
        system.save(update_fields=['health'])
    
    # Add downtime metrics for each system
    for system in systems:
        # Get current downtime status
        current_downtime = system.get_current_downtime()
        system.current_downtime = current_downtime
        system.is_currently_down = system.is_currently_down()
        
        # Get downtime metrics for the last 30 days
        downtime_metrics = system.get_downtime_metrics(30)
        system.downtime_metrics_data = downtime_metrics
    
    return render(request, 'products/system_allocation.html', {
        'participants': combined_participants,
        'systems': systems,
        'stream': stream or 'HIC',
        'selected_stream': stream or 'HIC',
        'view_date': view_date,
        'view_date_str': view_date.strftime('%Y-%m-%d'),
        'can_manage_downtime': can_manage_system_allocation(request.user)
    })

@login_required
@csrf_exempt
@require_POST
def allocate_system(request, stream=None):
    if request.method == 'POST':
        logger.info('Received POST data: %s', request.POST)
    system_type = request.POST.get('system_type')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    participant_id = request.POST.get('participant_id')
    # Parse datetime-local input robustly
    def parse_dt(dt_str):
        if not dt_str:
            return None
        for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
            try:
                return make_aware(datetime.strptime(dt_str, fmt))
            except ValueError:
                continue
        return None
    start_dt = parse_dt(start_date)
    end_dt = parse_dt(end_date)
    if not (system_type and start_dt and end_dt):
        return JsonResponse({'success': False, 'error': 'Missing or invalid data.'}, status=400)
    user = request.user
    blocked_for_participant = None
    if request.user.is_superuser and participant_id:
        if participant_id.startswith('user_'):
            try:
                user_id = int(participant_id.split('_')[1])
                user = User.objects.get(id=user_id)
            except Exception:
                return JsonResponse({'success': False, 'error': 'User not found.'}, status=404)
        elif participant_id.startswith('participant_'):
            try:
                part_id = int(participant_id.split('_')[1])
                participant = Participant.objects.get(id=part_id)
                blocked_for_participant = participant
                # Try to find a user by email or name
                user = User.objects.filter(email=participant.email).first()
                if not user:
                    user = User.objects.filter(username=participant.name).first()
                if not user:
                    user = request.user  # fallback to admin
            except Participant.DoesNotExist:
                user = request.user
        else:
            # fallback: treat as participant id (legacy)
            try:
                participant = Participant.objects.get(id=participant_id)
                blocked_for_participant = participant
                user = request.user
            except Participant.DoesNotExist:
                user = request.user
    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    # Check for overlap (block for all users, including admin)
    overlap = SystemAllocation.objects.filter(
        system_type=system_type,
        end_date__gte=start_dt,
        start_date__lte=end_dt,
        stream=stream_obj
    ).exists()
    if overlap:
        return JsonResponse({'success': False, 'error': 'System already blocked for this period.'}, status=409)
    allocation = SystemAllocation.objects.create(
        system_type=system_type,
        user=user,
        start_date=start_dt,
        end_date=end_dt,
        blocked_for_participant=blocked_for_participant,
        stream=stream_obj
    )
    return JsonResponse({'success': True, 'allocation_id': allocation.id})

@login_required
@require_GET
def get_blocked_systems(request, stream=None):
    # Remove expired allocations before returning
    today = date.today()
    expired = SystemAllocation.objects.filter(end_date__lt=today)
    expired.delete()
    # Get month/year from query params if present
    month = request.GET.get('month')
    year = request.GET.get('year')
    allocations = SystemAllocation.objects.filter(end_date__gte=today)
    if stream:
        # Handle empty stream parameter
        if not stream or stream.strip() == '':
            stream = 'HIC'
        
        # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        
        allocations = allocations.filter(stream=stream_obj)
    if month and year:
        try:
            month = int(month)
            year = int(year)
            # Get all allocations that overlap any day in the selected month
            from calendar import monthrange
            from django.utils.timezone import make_aware
            month_start = make_aware(datetime(year, month, 1))
            month_end = make_aware(datetime(year, month, monthrange(year, month)[1], 23, 59, 59))
            allocations = allocations.filter(
                start_date__lte=month_end,
                end_date__gte=month_start
            )
        except Exception:
            pass
    # Return all allocations (not just latest per system)
    data = []
    for a in allocations.order_by('system_type', 'start_date'):
        blocked_for = None
        if a.blocked_for_participant:
            blocked_for = f"{a.blocked_for_participant.name} ({a.blocked_for_participant.email})"
        # Convert to local time before formatting
        start_date_local = localtime(a.start_date)
        end_date_local = localtime(a.end_date)
        start_date_str = start_date_local.strftime('%Y-%m-%d %H:%M')
        end_date_str = end_date_local.strftime('%Y-%m-%d %H:%M')
        data.append({
            'id': a.id,
            'system_type': a.system_type,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'blocked_by': a.user.username,
            'blocked_for': blocked_for
        })
    return JsonResponse({'allocations': data})

@login_required
@csrf_exempt
def release_system(request, stream=None):
    if request.method == 'POST':
        system_type = request.POST.get('system_type')
        username = request.POST.get('username')
        allocation_id = request.POST.get('allocation_id')  # Get allocation ID

        allocation = None
        user = None

        # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')

        # Only release if allocation_id is a valid integer
        if allocation_id and allocation_id.isdigit():
            allocation = SystemAllocation.objects.filter(id=int(allocation_id), stream=stream_obj).first()
            # Optionally, you can check system_type and user match for extra safety
        else:
            # Do not release anything if allocation_id is missing or invalid
            return JsonResponse({'success': False, 'error': 'Invalid or missing allocation ID.'}, status=400)

        if allocation:
            # Send notification to user if admin is releasing someone else's allocation
            if request.user.is_superuser and username:
                user = User.objects.filter(username=username).first()
                if user and user != request.user:
                    from .models import Notification
                    Notification.objects.create(
                        user=user,
                        message=f"Your system allocation for {system_type} was released by admin.",
                        notification_type='allocation'
                    )
            allocation.delete()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'No active allocation found for this user and system.'})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@login_required
@require_POST
def add_participant(request):
    if not request.user.is_superuser:
        return redirect('user_list')
    name = request.POST.get('name', '').strip()
    email = request.POST.get('email', '').strip()
    if not name or not email:
        messages.error(request, 'Name and email are required.')
        return redirect('user_list')
    if Participant.objects.filter(email=email).exists():
        messages.error(request, 'A participant with this email already exists.')
        return redirect('user_list')
    Participant.objects.create(name=name, email=email)
    messages.success(request, f'Participant {name} added.')
    return redirect('user_list')

@login_required
@require_POST
def remove_participant(request, participant_id):
    if not request.user.is_superuser:
        messages.error(request, 'Only admins can remove participants.')
        return redirect('user_list')
    participant = get_object_or_404(Participant, id=participant_id)
    participant.delete()
    messages.success(request, 'Participant removed.')
    return redirect('user_list')

@login_required
@csrf_exempt
@require_POST
def extend_system(request, stream=None):
    system_type = request.POST.get('system_type')
    username = request.POST.get('username')
    new_end_date = request.POST.get('new_end_date')
    allocation_id = request.POST.get('allocation_id')  # <-- Add this line

    def parse_dt(dt_str):
        if not dt_str:
            return None
        for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
            try:
                return make_aware(datetime.strptime(dt_str, fmt))
            except ValueError:
                continue
        return None

    new_end_dt = parse_dt(new_end_date)
    if not (system_type and username and new_end_dt and allocation_id):
        return JsonResponse({'success': False, 'error': 'Missing or invalid data.'}, status=400)

    user = User.objects.filter(username=username).first()
    if not user:
        return JsonResponse({'success': False, 'error': 'User not found.'}, status=404)
    if request.user != user and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    # Ensure allocation_id is an integer
    try:
        alloc_id_int = int(allocation_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid allocation ID.'}, status=400)

    allocation = SystemAllocation.objects.filter(id=alloc_id_int, system_type=system_type, user=user, stream=stream_obj).first()
    if not allocation:
        return JsonResponse({'success': False, 'error': 'No active allocation found.'}, status=404)

    overlap = SystemAllocation.objects.filter(
        system_type=system_type,
        start_date__lt=new_end_dt,
        end_date__gt=allocation.end_date,
        stream=stream_obj
    ).exclude(id=allocation.id).exists()
    if overlap:
        return JsonResponse({'success': False, 'error': 'System already blocked for this extended period.'}, status=409)

    allocation.end_date = new_end_dt
    allocation.save()
    return JsonResponse({'success': True})

@login_required
@csrf_exempt
@require_POST
def opt_change_system(request, stream=None):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    system_type = request.POST.get('system_type')
    old_username = request.POST.get('old_username')
    participant_id = request.POST.get('participant_id')
    new_start_date = request.POST.get('new_start_date')  # <-- Expect this from frontend for split
    new_end_date = request.POST.get('new_end_date')
    def parse_dt(dt_str):
        for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
            try:
                return make_aware(datetime.strptime(dt_str, fmt))
            except Exception:
                continue
        return None
    new_start_dt = parse_dt(new_start_date) if new_start_date else None
    new_end_dt = parse_dt(new_end_date)
    if not (system_type and old_username and participant_id and new_end_dt):
        return JsonResponse({'success': False, 'error': 'Missing or invalid data.'}, status=400)
    old_user = User.objects.filter(username=old_username).first()
    if not old_user:
        return JsonResponse({'success': False, 'error': 'Old user not found.'}, status=404)
    # Resolve stream to a Stream object and find the original allocation (the one to split)
    stream_obj = get_stream_or_404(stream, default='HIC')
    allocation = SystemAllocation.objects.filter(system_type=system_type, user=old_user, stream=stream_obj).order_by('-end_date').first()
    if not allocation:
        return JsonResponse({'success': False, 'error': 'No active allocation found.'}, status=404)
    # The new allocation must start after the original allocation ends, or at a specified new_start_dt
    orig_end = allocation.end_date
    split_start = new_start_dt if new_start_dt and new_start_dt > allocation.start_date else orig_end
    # Removed validation: if new_end_dt <= split_start:
    #     return JsonResponse({'success': False, 'error': 'New end date/time must be after new start date/time.'}, status=400)
    # Determine if participant_id is user_ or participant_
    new_user = None
    participant = None
    if participant_id.startswith('user_'):
        try:
            user_id = int(participant_id.split('_')[1])
            new_user = User.objects.get(id=user_id)
        except Exception:
            return JsonResponse({'success': False, 'error': 'User not found.'}, status=404)
    elif participant_id.startswith('participant_'):
        try:
            part_id = int(participant_id.split('_')[1])
            participant = Participant.objects.get(id=part_id)
            # Try to find a user by email or name
            new_user = User.objects.filter(email=participant.email).first()
            if not new_user:
                new_user = User.objects.filter(username=participant.name).first()
            if not new_user:
                # Use admin as user, but set blocked_for_participant
                new_user = request.user
            # Always set participant for blocked_for_participant
        except Participant.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Participant not found.'}, status=404)
    else:
        return JsonResponse({'success': False, 'error': 'Invalid participant selection.'}, status=400)
    # Check for overlap for new user (for the new period only)
    overlap = SystemAllocation.objects.filter(
        system_type=system_type,
        start_date__lt=new_end_dt,
        end_date__gt=split_start,
        user=new_user,
        stream=stream_obj
    ).exists()
    if overlap:
        return JsonResponse({'success': False, 'error': 'System already blocked for this period for the new user.'}, status=409)
    # Create new allocation for the new user/participant
    new_alloc = SystemAllocation.objects.create(
        system_type=system_type,
        user=new_user,
        start_date=split_start,
        end_date=new_end_dt,
        blocked_for_participant=participant,
        stream=stream_obj
    )
    # Notify old user if allocation was taken over by admin
    if old_user != new_user:
        from .models import Notification
        Notification.objects.create(
            user=old_user,
            message=f"Your system allocation for {system_type} was released and re-allocated to another user by admin.",
            notification_type='allocation'
        )
    # Only delete the original allocation if the new allocation exactly matches the original's start and end
    if split_start == allocation.start_date and new_end_dt == allocation.end_date:
        allocation.delete()
    return JsonResponse({'success': True, 'new_allocation_id': new_alloc.id})

@login_required
def location_list(request, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'PIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Get the Stream object for database queries (404 if not found). Default to 'PIC' when missing.
    stream_obj = get_stream_or_404(stream, default='PIC')
    
    locations = Location.objects.filter(stream=stream_obj).order_by('-created_at')
    return render(request, 'products/location_list.html', {'locations': locations, 'stream': stream, 'selected_stream': stream})

@login_required
def location_create(request, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    if not request.user.is_superuser:
        messages.error(request, 'Only admins can add locations.')
        return redirect('location_list_stream', stream=stream or 'HIC')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        address = request.POST.get('address', '').strip()
        # Get Stream object from name (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        if name:
            Location.objects.create(name=name, address=address, stream=stream_obj)
            messages.success(request, 'Location added successfully.')
            return redirect('location_list_stream', stream=stream or 'HIC')
        else:
            messages.error(request, 'Location name is required.')
    return render(request, 'products/location_form.html', {'stream': stream, 'selected_stream': stream})

@login_required
def location_edit(request, pk, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    location = get_object_or_404(Location, pk=pk, stream=stream_obj)
    if not request.user.is_superuser:
        messages.error(request, 'Only admins can modify locations.')
        return redirect('location_list_stream', stream=stream)
    if request.method == 'POST':
        location.name = request.POST.get('name', '').strip()
        location.address = request.POST.get('address', '').strip()
        location.save()
        messages.success(request, 'Location updated successfully.')
        return redirect('location_list_stream', stream=stream)
    return render(request, 'products/location_form.html', {'location': location, 'edit': True, 'stream': stream, 'selected_stream': stream})

@login_required
def location_delete(request, pk, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    try:
        location = Location.objects.get(pk=pk, stream=stream_obj)
    except Location.DoesNotExist:
        messages.warning(request, 'Location already deleted.')
        return redirect('location_list_stream', stream=stream)
    if not request.user.is_superuser:
        messages.error(request, 'Only admins can remove locations.')
        return redirect('location_list_stream', stream=stream)
    if request.method == 'POST':
        location.delete()
        messages.success(request, 'Location removed successfully.')
        return redirect('location_list_stream', stream=stream)
    return render(request, 'products/location_confirm_delete.html', {'location': location, 'stream': stream, 'selected_stream': stream})

@login_required
@csrf_exempt
@require_POST
def add_system(request, stream=None):
    try:
        name = request.POST.get('name', '').strip()
        if not name:
            print("DEBUG: System name missing")
            return JsonResponse({'success': False, 'error': 'System name required.'}, status=400)
        
        # Handle empty stream parameter
        if not stream or stream.strip() == '':
            stream = 'HIC'
        
        # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        
        # Ensure system is unique within the stream
        if System.objects.filter(name=name, stream=stream_obj).exists():
            return JsonResponse({'success': False, 'error': 'System already exists.'}, status=409)
        system = System.objects.create(name=name, stream=stream_obj)
        return JsonResponse({
            'success': True,
            'system_id': system.id,
            'name': system.name,
            'status': system.status,
            'description': system.description
        })
    except Exception as e:
        import traceback
        print(f"DEBUG: Exception occurred: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'}, status=500)

@user_passes_test(is_admin)
@login_required
@csrf_exempt
@require_POST
def delete_system(request, stream=None):
    system_id = request.POST.get('id')
    if not system_id:
        return JsonResponse({'success': False, 'error': 'System ID required.'}, status=400)
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    try:
        system = System.objects.get(id=system_id, stream=stream_obj)
        system.delete()
        return JsonResponse({'success': True})
    except System.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'System not found.'}, status=404)

@user_passes_test(is_admin)
@login_required
@csrf_exempt
@require_POST
def update_system(request, stream=None):
    system_id = request.POST.get('id')
    status = request.POST.get('status')
    description = request.POST.get('description', '').strip()
    participant_id = request.POST.get('participant_id')
    status_date = request.POST.get('status_date', '')  # Get the date when status changed
    assignee = None

    # Basic validation
    if not system_id or not status:
        return JsonResponse({'success': False, 'error': 'Missing system id or status.'}, status=400)

    # Ensure system_id is an integer
    try:
        system_id_int = int(system_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid system id.'}, status=400)

    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    # Fetch the system using the numeric id and Stream instance
    try:
        system = System.objects.get(id=system_id_int, stream=stream_obj)
    except System.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'System not found.'}, status=404)

    if participant_id:
        if participant_id.startswith('user_'):
            try:
                user_id = int(participant_id.split('_')[1])
                user_obj = User.objects.get(id=user_id)
                assignee = f"{user_obj.username} ({user_obj.email})"
            except Exception:
                assignee = None
        elif participant_id.startswith('participant_'):
            try:
                part_id = int(participant_id.split('_')[1])
                from .models import Participant
                participant = Participant.objects.get(id=part_id)
                assignee = f"{participant.name} ({participant.email})"
            except Exception:
                assignee = None
        else:
            # fallback: treat as participant id (legacy)
            try:
                from .models import Participant
                participant = Participant.objects.get(id=participant_id)
                assignee = f"{participant.name} ({participant.email})"
            except Exception:
                assignee = None

    system.status = status
    system.description = description
    system.save()

    # Save status history
    from .models import SystemStatusHistory
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    # Parse status_date if provided
    history_updated_at = None
    if status_date:
        try:
            # Parse the date string (YYYY-MM-DD format from HTML date input)
            history_date = datetime.strptime(status_date, '%Y-%m-%d')
            view_date = history_date.date()
            
            # Check if updating today's date
            today = timezone.now().date()
            
            if view_date == today:
                # If updating today, always use current time for proper ordering
                history_updated_at = timezone.now()
            else:
                # For past/future dates, check for existing records on that date
                existing_records = SystemStatusHistory.objects.filter(
                    system=system,
                    updated_at__date=view_date
                ).order_by('-updated_at')
                
                if existing_records.exists():
                    # Get the latest timestamp on that date and add 1 minute
                    latest_time = existing_records.first().updated_at
                    history_updated_at = latest_time + timedelta(minutes=1)
                else:
                    # No existing records on this date, use noon as the timestamp
                    history_updated_at = timezone.make_aware(
                        datetime.combine(view_date, datetime.min.time().replace(hour=12))
                    )
        except (ValueError, TypeError):
            # If parsing fails, use current time
            history_updated_at = timezone.now()
    else:
        # If no date provided, use current time
        history_updated_at = timezone.now()
    
    # Create history record with the specified date
    history_record = SystemStatusHistory.objects.create(
        system=system,
        status=status,
        description=description,
        assignee=assignee,
        updated_by=request.user.username,
        updated_at=history_updated_at
    )
    
    return JsonResponse({'success': True})

@user_passes_test(is_admin)
@login_required
@csrf_exempt
@require_POST
def reset_system_utilization(request, stream=None):
    """Reset utilization percentage for one or all systems in a stream."""
    try:
        system_id = request.POST.get('system_id')
        reset_all = request.POST.get('reset_all') == 'true'
        
        # Handle empty stream parameter
        if not stream or stream.strip() == '':
            stream = 'HIC'
        
        # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        
        if reset_all:
            # Reset utilization for all systems in the stream
            systems = System.objects.filter(stream=stream_obj)
            count = systems.update(utilization_percentage=0.0)
            
            # Also reset any related metrics
            from .models import SystemMetrics
            for system in systems:
                metrics, created = SystemMetrics.objects.get_or_create(system=system)
                metrics.usage_hours = 0.0
                metrics.total_allocations = 0
                metrics.last_allocation_date = None
                metrics.average_session_duration = None
                metrics.uptime_percentage = 100.0
                metrics.save()
            return JsonResponse({
                'success': True, 
                'message': f'Reset utilization for {count} systems in {stream} stream.'
            })
        
        elif system_id:
            # Reset utilization for a specific system
            system = System.objects.get(id=system_id, stream=stream_obj)
            system.utilization_percentage = 0.0
            system.save(update_fields=['utilization_percentage'])
            
            # Reset related metrics
            from .models import SystemMetrics
            metrics, created = SystemMetrics.objects.get_or_create(system=system)
            metrics.usage_hours = 0.0
            metrics.total_allocations = 0
            metrics.last_allocation_date = None
            metrics.average_session_duration = None
            metrics.uptime_percentage = 100.0
            metrics.save()
            
            return JsonResponse({
                'success': True, 
                'message': f'Reset utilization for system "{system.name}" is successful.'
            })
        
        else:
            return JsonResponse({
                'success': False, 
                'error': 'Either system_id or reset_all parameter is required.'
            }, status=400)
            
    except System.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'System not found.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred: {str(e)}'
        }, status=500)

@login_required
def dashboard(request, stream=None):
    from .models import (Product, Category, Participant, Location, System, Stream, 
                        SystemStatus, UserSession)
    from django.contrib.auth import get_user_model
    from django.utils import timezone
    from datetime import timedelta

    User = get_user_model()
    custom_profile, created = CustomUser.objects.get_or_create(user=request.user)

    # Check if user has any roles (except superusers)
    if not request.user.is_superuser and not custom_profile.user_roles.exists():
        messages.error(request, 'Access denied. You have no assigned roles. Please contact an administrator.')
        context = {
            'total_products': 0,
            'total_categories': 0,
            'total_users': 0,
            'total_locations': 0,
            'total_participants': 0,
            'total_systems': 0,
            'online_users': 0,
            'streams': [],
            'selected_stream': '',
            'system_status': None,
            'current_time': timezone.now(),
            'user_custom_profile': custom_profile,
            'user_permissions': {},
            'access_error_message': 'Access denied. You have no assigned roles. Please contact an administrator.',
        }
        return render(request, 'products/dashboard.html', context)

    # Use Stream model for streams list
    streams = list(Stream.objects.values_list('name', flat=True).order_by('name'))
    if 'PIC' not in streams:
        Stream.objects.get_or_create(name='PIC')
        streams.append('PIC')
    if 'HIC' not in streams:
        Stream.objects.get_or_create(name='HIC')
        streams.append('HIC')
    def stream_sort_key(s):
        if s == 'PIC':
            return (0, s)
        elif s == 'HIC':
            return (1, s)
        return (2, s)
    streams = sorted(set(streams), key=stream_sort_key)

    # Filter streams based on user access (unless superuser)
    if not request.user.is_superuser:
        accessible_streams = custom_profile.get_accessible_streams()
        accessible_stream_names = list(accessible_streams.values_list('name', flat=True))
        streams = [s for s in streams if s in accessible_stream_names]
        # If user has no accessible streams, show message on dashboard only (do not add to Django messages)
        if not streams:
            context = {
                'total_products': 0,
                'total_categories': 0,
                'total_users': 0,
                'total_locations': 0,
                'total_participants': 0,
                'total_systems': 0,
                'online_users': 0,
                'streams': [],
                'selected_stream': '',
                'system_status': None,
                'current_time': timezone.now(),
                'user_custom_profile': custom_profile,
                'user_permissions': {},
                'access_error_message': 'Access denied. You do not have access to any streams.',
            }
            return render(request, 'products/dashboard.html', context)
    
    stream = stream or request.GET.get('stream', streams[0] if streams else 'HIC')
    
    # Check if user can access the requested stream (unless superuser)
    if not request.user.is_superuser and stream not in streams:
        messages.error(request, f'Access denied. You do not have permission to access the {stream} stream.')
        return redirect('dashboard_stream', stream=streams[0])
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Calculate dynamic statistics
    total_products = Product.objects.count()  # Count products across all streams
    total_categories = Category.objects.filter(stream=stream_obj).count()
    total_users = User.objects.filter(is_active=True).count()
    total_locations = Location.objects.filter(stream=stream_obj).count()
    total_participants = Participant.objects.count()    
    total_systems = System.objects.filter(stream=stream_obj).count()
    
    # Calculate online users (active in last 15 minutes)
    cutoff_time = timezone.now() - timedelta(minutes=15)
    
    # First, mark old sessions as inactive
    UserSession.objects.filter(
        last_activity__lt=cutoff_time,
        is_active=True
    ).update(is_active=False)
      # Count truly active users
    online_users = UserSession.objects.filter(
        last_activity__gte=cutoff_time,
        is_active=True
    ).count()
    
    # Get system status
    system_status = SystemStatus.objects.first()
    if not system_status:
        system_status = SystemStatus.objects.create(
            status='online',
            description='System operational'
        )
      # Get current time for weather widget
    current_time = timezone.now()
    
    # Get or create user custom profile
    custom_profile, created = CustomUser.objects.get_or_create(user=request.user)
    
    # Get user permissions for dashboard display
    user_permissions = {
        'can_manage_users': can_manage_users(request.user),
        'can_manage_system_allocation': can_manage_system_allocation(request.user),
        'can_edit_products': can_edit_products(request.user),
        'can_delete_products': can_delete_products(request.user),
        'can_view_analytics': can_view_analytics(request.user),
        'is_admin': is_admin(request.user),
        'is_super_admin': is_super_admin(request.user),
        'is_lab_incharge': is_lab_incharge(request.user),
    }
    
    context = {
        'total_products': total_products,
        'total_categories': total_categories,
        'total_users': total_users,
        'total_locations': total_locations,
        'total_participants': total_participants,
        'total_systems': total_systems,
        'online_users': online_users,
        'streams': streams,
        'selected_stream': stream,
        'system_status': system_status,
        'current_time': current_time,
        'user_custom_profile': custom_profile,
        'user_permissions': user_permissions,
    }

    return render(request, 'products/dashboard.html', context)

@csrf_exempt
@require_POST
def delete_stream(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    stream_name = request.POST.get('stream_name', '').strip()
    password = request.POST.get('password', '').strip()
    if not stream_name or not password:
        return JsonResponse({'success': False, 'error': 'Missing stream name or password.'}, status=400)
    if stream_name in ['HIC', 'PIC']:
        return JsonResponse({'success': False, 'error': 'Cannot delete default streams.'}, status=400)
    user = authenticate(username=request.user.username, password=password)
    if not user:
        return JsonResponse({'success': False, 'error': 'Incorrect password.'}, status=403)
    from .models import Stream, StreamDeletionHistory
    try:
        stream_obj = Stream.objects.get(name=stream_name)
        stream_obj.delete()
        # Log deletion
        StreamDeletionHistory.objects.create(stream_name=stream_name, deleted_by=request.user)
        return JsonResponse({'success': True})
    except Stream.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Stream not found.'}, status=404)

@login_required
def stream_deletion_history(request):
    from .models import StreamDeletionHistory
    history = StreamDeletionHistory.objects.select_related('deleted_by').order_by('-deleted_at')[:100]
    return render(request, 'products/stream_deletion_history.html', {'history': history})

@login_required
def system_status_history(request, stream, system_id):
    # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    system = get_object_or_404(System, id=system_id, stream=stream_obj)
    history = SystemStatusHistory.objects.filter(system=system).order_by('-updated_at')
    return render(request, 'products/system_status_history.html', {
        'system': system,
        'history': history,
        'stream': stream,
        'selected_stream': stream, 
    })

@login_required
def delete_user_backup(request, backup_id):
    if not request.user.is_superuser:
        return HttpResponse('Unauthorized', status=401)
    backup = get_object_or_404(UserDataVersion, id=backup_id)
    backup.delete()
    messages.success(request, 'Backup deleted successfully.')
    return redirect('user_list')

def please_login(request):
    return render(request, 'products/please_login.html')

@login_required
def build_os_info(request):
    # Get all products assigned to a Zenition category
    zenition_category_qs = Category.objects.filter(name__icontains='Zenition')
    zenition_products = ZenitionProduct.objects.all()
    selected_product_id = request.GET.get('selected_product_id')
    warning = None
    delete_result = None
    if request.method == 'POST':
        new_product = request.POST.get('new_zenition_product', '').strip()
        if new_product:
            if ZenitionProduct.objects.filter(name=new_product).exists():
                warning = f"Product '{new_product}' already exists."
            else:
                ZenitionProduct.objects.create(name=new_product)
                zenition_products = ZenitionProduct.objects.all()
        elif 'delete_product_id' in request.POST:
            product_id = request.POST.get('delete_product_id')
            password = request.POST.get('delete_product_password', '').strip()
            if product_id:
                # Password check: use Django's check_password for the current user
                if not password or not request.user.check_password(password):
                    delete_result = 'Incorrect password.'
                else:
                    ZenitionProduct.objects.filter(id=product_id).delete()
                    zenition_products = ZenitionProduct.objects.all()
                    delete_result = 'Product deleted successfully.'
    return render(request, 'products/build_os_info.html', {
        'zenition_products': zenition_products,
        'selected_product_id': selected_product_id,
        'warning': warning,
        'delete_result': delete_result
    })

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import ZenitionProduct, ProductEntry
import json

@csrf_exempt
def product_entries_api(request):
    if request.method == 'GET':
        product_id = request.GET.get('product_id')
        category = request.GET.get('category')  # OS/Binaries
        type_ = request.GET.get('type')         # MVS/Stand PC/Apps PC
        entries = ProductEntry.objects.filter(
            zenition_product_id=product_id,
            entry_type=category,
            category=type_
        )
        data = [
            {
                'id': e.id, 
                'subcategory': e.subcategory,
                'category': e.subcategory,
                'link': e.link,
                'os_system_type': e.os_system_type.id if e.os_system_type else None,
                'os_system_type_name': e.os_system_type.name if e.os_system_type else '',
                'binaries_system_type': e.binaries_system_type.id if e.binaries_system_type else None,
                'binaries_system_type_name': e.binaries_system_type.name if e.binaries_system_type else '',
            }
            for e in entries
        ]
        return JsonResponse({'entries': data})
    elif request.method == 'POST':
        body = json.loads(request.body)
        product_id = body.get('product_id')
        entry_type = body.get('category')  # OS/Binaries
        type_ = body.get('type')           # MVS/Stand PC/Apps PC
        subcategory = body.get('subcategory')
        link = body.get('link')
        os_system_type_id = body.get('os_system_type_id')
        binaries_system_type_id = body.get('binaries_system_type_id')
        
        entry = ProductEntry.objects.create(
            zenition_product_id=product_id,
            entry_type=entry_type,
            category=type_,
            subcategory=subcategory,
            link=link,
            os_system_type_id=os_system_type_id if os_system_type_id else None,
            binaries_system_type_id=binaries_system_type_id if binaries_system_type_id else None,
        )
        return JsonResponse({
            'id': entry.id, 
            'subcategory': entry.subcategory,
            'category': entry.subcategory, 
            'link': entry.link,
            'os_system_type': entry.os_system_type.id if entry.os_system_type else None,
            'os_system_type_name': entry.os_system_type.name if entry.os_system_type else '',
            'binaries_system_type': entry.binaries_system_type.id if entry.binaries_system_type else None,
            'binaries_system_type_name': entry.binaries_system_type.name if entry.binaries_system_type else '',
        })
    elif request.method == 'PUT':
        body = json.loads(request.body)
        entry_id = body.get('id')
        product_id = body.get('product_id')
        entry_type = body.get('category')  # OS/Binaries
        type_ = body.get('type')           # MVS/Stand PC/Apps PC
        subcategory = body.get('subcategory')
        link = body.get('link')
        os_system_type_id = body.get('os_system_type_id')
        binaries_system_type_id = body.get('binaries_system_type_id')
        
        entry = ProductEntry.objects.get(id=entry_id)
        entry.category = type_
        entry.subcategory = subcategory
        entry.link = link
        entry.os_system_type_id = os_system_type_id if os_system_type_id else None
        entry.binaries_system_type_id = binaries_system_type_id if binaries_system_type_id else None
        entry.save()
        return JsonResponse({
            'id': entry.id, 
            'subcategory': entry.subcategory,
            'category': entry.subcategory, 
            'link': entry.link,
            'os_system_type': entry.os_system_type.id if entry.os_system_type else None,
            'os_system_type_name': entry.os_system_type.name if entry.os_system_type else '',
            'binaries_system_type': entry.binaries_system_type.id if entry.binaries_system_type else None,
            'binaries_system_type_name': entry.binaries_system_type.name if entry.binaries_system_type else '',
        })
    elif request.method == 'DELETE':
        body = json.loads(request.body)
        entry_id = body.get('id')
        ProductEntry.objects.filter(id=entry_id).delete()
        return JsonResponse({'deleted': True})

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import ZenitionProduct

@csrf_exempt
def add_zenition_product_api(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Product name required.'})
        # Case-insensitive check for existing product
        if ZenitionProduct.objects.filter(name__iexact=name).exists():
            return JsonResponse({'success': False, 'error': 'Product already exists.'})
        product = ZenitionProduct.objects.create(name=name)
        return JsonResponse({'success': True, 'id': product.id, 'name': product.name})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

@csrf_exempt
def delete_zenition_product_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)
    import json
    data = json.loads(request.body.decode('utf-8'))
    product_id = data.get('id')
    password = data.get('password')
    user = request.user
    if not user.is_authenticated or not (user.is_superuser or user.is_staff):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    if not password or not user.check_password(password):
        return JsonResponse({'success': False, 'error': 'Incorrect password.'}, status=403)
    from .models import ZenitionProduct
    try:
        product = ZenitionProduct.objects.get(id=product_id)
        product.delete()
        return JsonResponse({'success': True, 'message': 'Product deleted successfully.'})
    except ZenitionProduct.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found.'}, status=404)

@csrf_exempt
def os_system_types_api(request):
    """API for managing OS System Types"""
    from .models import OSSystemType
    
    if request.method == 'GET':
        types = OSSystemType.objects.all().order_by('name')
        data = [{'id': t.id, 'name': t.name, 'description': t.description} for t in types]
        return JsonResponse({'types': data})
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required.'}, status=400)
        
        if OSSystemType.objects.filter(name__iexact=name).exists():
            return JsonResponse({'success': False, 'error': 'OS System Type already exists.'}, status=400)
        
        os_type = OSSystemType.objects.create(
            name=name,
            description=description,
            created_by=request.user if request.user.is_authenticated else None
        )
        return JsonResponse({
            'success': True,
            'id': os_type.id,
            'name': os_type.name,
            'description': os_type.description
        })
    
    elif request.method == 'DELETE':
        data = json.loads(request.body)
        type_id = data.get('id')
        
        if not type_id:
            return JsonResponse({'success': False, 'error': 'ID is required.'}, status=400)
        
        try:
            os_type = OSSystemType.objects.get(id=type_id)
            os_type.delete()
            return JsonResponse({'success': True, 'message': 'OS System Type deleted successfully.'})
        except OSSystemType.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'OS System Type not found.'}, status=404)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

@csrf_exempt
def binaries_system_types_api(request):
    """API for managing Binaries System Types"""
    from .models import BinariesSystemType
    
    if request.method == 'GET':
        types = BinariesSystemType.objects.all().order_by('name')
        data = [{'id': t.id, 'name': t.name, 'description': t.description} for t in types]
        return JsonResponse({'types': data})
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required.'}, status=400)
        
        if BinariesSystemType.objects.filter(name__iexact=name).exists():
            return JsonResponse({'success': False, 'error': 'Binaries System Type already exists.'}, status=400)
        
        binaries_type = BinariesSystemType.objects.create(
            name=name,
            description=description,
            created_by=request.user if request.user.is_authenticated else None
        )
        return JsonResponse({
            'success': True,
            'id': binaries_type.id,
            'name': binaries_type.name,
            'description': binaries_type.description
        })
    
    elif request.method == 'DELETE':
        data = json.loads(request.body)
        type_id = data.get('id')
        
        if not type_id:
            return JsonResponse({'success': False, 'error': 'ID is required.'}, status=400)
        
        try:
            binaries_type = BinariesSystemType.objects.get(id=type_id)
            binaries_type.delete()
            return JsonResponse({'success': True, 'message': 'Binaries System Type deleted successfully.'})
        except BinariesSystemType.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Binaries System Type not found.'}, status=404)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

@csrf_exempt
@require_GET
def system_types_category_mapping_api(request):
    """API for fetching system types with their mapped category counts"""
    from .models import OSSystemType, BinariesSystemType, ProductEntry
    from django.db.models import Count
    
    # Get OS System Types with category counts
    os_types_data = []
    os_types = OSSystemType.objects.all().order_by('name')
    for os_type in os_types:
        # Count unique categories mapped to this OS type via ProductEntry
        category_count = ProductEntry.objects.filter(
            os_system_type=os_type
        ).values('category').distinct().count()
        
        os_types_data.append({
            'id': os_type.id,
            'name': os_type.name,
            'type': 'OS',
            'category_count': category_count
        })
    
    # Get Binaries System Types with category counts
    binaries_types_data = []
    binaries_types = BinariesSystemType.objects.all().order_by('name')
    for binaries_type in binaries_types:
        # Count unique categories mapped to this Binaries type via ProductEntry
        category_count = ProductEntry.objects.filter(
            binaries_system_type=binaries_type
        ).values('category').distinct().count()
        
        binaries_types_data.append({
            'id': binaries_type.id,
            'name': binaries_type.name,
            'type': 'Binaries',
            'category_count': category_count
        })
    
    return JsonResponse({
        'os_types': os_types_data,
        'binaries_types': binaries_types_data,
        'total_os_types': len(os_types_data),
        'total_binaries_types': len(binaries_types_data)
    })

@login_required
def download_inventory_with_os_binaries_excel(request):
    if not request.user.is_superuser:
        return HttpResponse('Unauthorized', status=401)
    wb = Workbook()
    # Main Inventory Sheet
    ws_inventory = wb.active
    ws_inventory.title = 'Inventory Data'
    headers = ['S. NO', 'Category name', 'Category Serial Number', 'Product Name', 'Product Serial Number', 'Product Description', 'Product Added date', 'Product added by', 'Stream', 'Location Name', 'Location Address']
    ws_inventory.append(headers)
    for row in get_user_product_data():
        ws_inventory.append(row)
    for col in range(1, len(headers) + 1):
        ws_inventory.column_dimensions[get_column_letter(col)].width = 22

    # OS Sheet
    ws_os = wb.create_sheet(title='OS Data')
    os_headers = ['Zenition Product', 'Type', 'Subcategory', 'Link']
    ws_os.append(os_headers)
    for product in ZenitionProduct.objects.all():
        os_entries = ProductEntry.objects.filter(zenition_product=product, entry_type='OS')
        for entry in os_entries:
            ws_os.append([
                product.name,
                entry.category,
                entry.subcategory or '',
                entry.link
            ])
    for col in range(1, len(os_headers) + 1):
        ws_os.column_dimensions[get_column_letter(col)].width = 22

    # Binaries Sheet
    ws_bin = wb.create_sheet(title='Binaries Data')
    bin_headers = ['Zenition Product', 'Type', 'Subcategory', 'Link']
    ws_bin.append(bin_headers)
    for product in ZenitionProduct.objects.all():
        bin_entries = ProductEntry.objects.filter(zenition_product=product, entry_type='Binaries')
        for entry in bin_entries:
            ws_bin.append([
                product.name,
                entry.category,
                entry.subcategory or '',
                entry.link
            ])
    for col in range(1, len(bin_headers) + 1):
        ws_bin.column_dimensions[get_column_letter(col)].width = 22

    # SubLevel Data Sheet
    ws_sublevel = wb.create_sheet(title='SubLevel Data')
    sublevel_headers = ['Item Name', 'In stock', 'In use', 'Scrapped', 'Stream', 'Last modified by', 'Last modified date']
    ws_sublevel.append(sublevel_headers)
    # Get all sublevels for the export (respect stream if needed)
    sublevels = SubLevel.objects.all()
    for sub in sublevels:
        # Get last history entry for this sublevel (if any)
        last_history = sub.history.order_by('-at').first()
        last_by = last_history.by if last_history else ''
        last_at = last_history.at.strftime('%Y-%m-%d %H:%M:%S') if last_history else ''
        ws_sublevel.append([
            sub.name,
            sub.in_stock,
            sub.in_use,
            sub.scraped,
            sub.stream or '',
            last_by,
            last_at
        ])
    for col in range(1, len(sublevel_headers) + 1):
        ws_sublevel.column_dimensions[get_column_letter(col)].width = 22

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_MoS_Inventory_with_OS_Binaries.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@csrf_exempt
@login_required
def communication_api(request):
    user = request.user
    # Auto-delete comments older than 30 days
    from datetime import datetime, timedelta, timezone
    Communication.objects.filter(page='build_os_info', created_at__lt=datetime.now(timezone.utc) - timedelta(days=30)).delete()
    if request.method == 'GET':
        comms = Communication.objects.filter(page='build_os_info').order_by('-created_at')[:50]
        return JsonResponse({'comments': [
            {
                'id': c.id,
                'user': c.user.username,
                'message': c.message,
                'deleted': c.deleted,
                'created_at': c.created_at.astimezone(timezone.utc).isoformat(),
                'updated_at': (c.updated_at.astimezone(timezone.utc).isoformat() if c.updated_at else c.created_at.astimezone(timezone.utc).isoformat()),
                'attachments': [
                    {
                        'id': att.id,
                        'filename': att.original_filename,
                        'file_size': att.file_size_formatted,
                        'content_type': att.content_type,
                        'is_image': att.is_image,
                        'url': f'/api/communication/attachment/{att.id}/'
                    } for att in c.attachments.all()
                ]
            } for c in comms
        ]})
    elif request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        msg = data.get('message', '').strip()
        attachment_ids = data.get('attachment_ids', [])
        
        if not msg:
            return JsonResponse({'success': False, 'error': 'Message required.'}, status=400)
            
        comm = Communication.objects.create(user=user, message=msg, page='build_os_info')
        
        # Link any pending attachments to this comment
        if attachment_ids:
            attachments_to_update = CommunicationAttachment.objects.filter(
                id__in=attachment_ids, 
                uploaded_by=user, 
                communication__isnull=True
            )
            for attachment in attachments_to_update:
                attachment.communication = comm
                attachment.save()
        
        return JsonResponse({
            'success': True, 
            'id': comm.id, 
            'message': comm.message, 
            'user': user.username, 
            'created_at': comm.created_at.astimezone(timezone.utc).isoformat(),
            'attachments': [
                {
                    'id': att.id,
                    'filename': att.original_filename,
                    'file_size': att.file_size_formatted,
                    'content_type': att.content_type,
                    'is_image': att.is_image,
                    'url': f'/api/communication/attachment/{att.id}/'
                } for att in comm.attachments.all()
            ]
        })
    elif request.method == 'PUT':
        data = json.loads(request.body.decode('utf-8'))
        cid = data.get('id')
        msg = data.get('message', '').strip()
        if not cid or not msg:
            return JsonResponse({'success': False, 'error': 'ID and message required.'}, status=400)
        try:
            comm = Communication.objects.get(id=cid, user=user, page='build_os_info')
        except Communication.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Comment not found.'}, status=404)
        # Only allow edit for 15 minutes after posting
        now = datetime.now(timezone.utc)
        if comm.deleted:
            return JsonResponse({'success': False, 'error': 'Comment deleted.'}, status=403)
        if (now - comm.created_at) > timedelta(minutes=15):
            return JsonResponse({'success': False, 'error': 'Edit window expired.'}, status=403)
        comm.message = msg
        comm.save()
        return JsonResponse({'success': True})
    elif request.method == 'DELETE':
        data = json.loads(request.body.decode('utf-8'))
        cid = data.get('id')
        if not cid:
            return JsonResponse({'success': False, 'error': 'ID required.'}, status=400)
        try:
            comm = Communication.objects.get(id=cid, user=user, page='build_os_info')
            now = datetime.now(timezone.utc)
        except Communication.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Comment not found.'}, status=404)
        now = datetime.now(timezone.utc)
        if comm.deleted:
            return JsonResponse({'success': False, 'error': 'Already deleted.'}, status=403)
        time_diff_hours = (now - comm.created_at).total_seconds() / 3600
        if time_diff_hours > 12:
            return JsonResponse({'success': False, 'error': 'Delete window expired.'}, status=403)
        comm.deleted = True
        comm.save()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid method.'}, status=405)

@csrf_exempt
@login_required
def upload_communication_attachment(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)
    
    uploaded_file = request.FILES['file']
    
    # Validate file size (5MB max)
    if uploaded_file.size > 5 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'File size must be less than 5MB'}, status=400)
    
    # Validate file type
    allowed_types = [
        'image/jpeg', 'image/png', 'image/gif', 'image/bmp', 'image/webp',
        'application/pdf', 'text/plain', 
        'application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ]
    
    if uploaded_file.content_type not in allowed_types:
        return JsonResponse({'success': False, 'error': 'File type not allowed'}, status=400)
    
    try:
        # Create a temporary attachment without a communication (will be linked when comment is posted)
        attachment = CommunicationAttachment.objects.create(
            communication=None,  # Will be set when comment is created
            file=uploaded_file,
            original_filename=uploaded_file.name,
            file_size=uploaded_file.size,
            content_type=uploaded_file.content_type,
            uploaded_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'attachment_id': attachment.id,
            'filename': attachment.original_filename,
            'file_size': attachment.file_size_formatted,
            'content_type': attachment.content_type,
            'is_image': attachment.is_image
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def serve_communication_attachment(request, attachment_id):
    try:
        attachment = CommunicationAttachment.objects.get(id=attachment_id)
        
        # Check if user has permission to view this attachment
        # For now, allow all authenticated users
        
        if not attachment.file:
            return JsonResponse({'error': 'File not found'}, status=404)
        
        response = FileResponse(
            attachment.file.open('rb'),
            content_type=attachment.content_type,
            filename=attachment.original_filename
        )
        response['Content-Length'] = attachment.file_size
        return response
        
    except CommunicationAttachment.DoesNotExist:
        return JsonResponse({'error': 'Attachment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import subprocess
import json

@csrf_exempt
def execute_robocopy(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    try:
        data = json.loads(request.body.decode('utf-8'))
        cmd = data.get('cmd', '')
        if not cmd:
            return JsonResponse({'success': False, 'error': 'No command provided'}, status=400)
        # Ensure the command prompt stays open after execution
        cmd_arg = cmd.replace('"', '""')
        # Add '&& pause' so cmd stays open after robocopy
        powershell_cmd = (
            f"Start-Process -FilePath 'cmd.exe' -ArgumentList '/k', '{cmd_arg} && pause' -Verb RunAs"
        )
        completed = subprocess.run(['powershell', '-Command', powershell_cmd], capture_output=True, text=True)
        if completed.returncode == 0:
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': completed.stderr})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def faq(request, stream=None):
    """Render the FAQ/help page for users."""
    context = {
        'selected_stream': stream or '',
        'stream': stream or '',
    }
    return render(request, 'products/faq.html', context)

@login_required
def user_profile(request, user_id=None):
    if user_id:
        # Only admins can view other users' profiles
        if not can_manage_users(request.user):
            messages.error(request, 'Access denied. You can only view your own profile.')
            return redirect('user_profile')
        
        try:
            profile_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
            return redirect('user_list')
    else:
        profile_user = request.user

    # Get or create custom profile for user
    custom_profile, created = CustomUser.objects.get_or_create(user=profile_user)
    
    # Handle profile image upload
    if request.method == 'POST' and profile_user == request.user:
        profile_image = request.FILES.get('profile_image')
        if profile_image:
            custom_profile.profile_image = profile_image
            custom_profile.save()
            messages.success(request, 'Profile image updated successfully!')
        else:
            messages.error(request, 'No image file selected.')
        return redirect('user_profile')
    
    # Get user's roles and stream access
    user_roles = custom_profile.user_roles.all()
    stream_access = custom_profile.stream_access.select_related('stream').all()
    accessible_streams = custom_profile.get_accessible_streams()
    
    # Get user permissions for display
    permissions = {
        'can_manage_users': can_manage_users(profile_user),
        'can_manage_system_allocation': can_manage_system_allocation(profile_user),
        'can_edit_products': can_edit_products(profile_user),
        'can_delete_products': can_delete_products(profile_user),
        'can_view_analytics': can_view_analytics(profile_user),
        'is_admin': is_admin(profile_user),
        'is_super_admin': is_super_admin(profile_user),
        'is_lab_incharge': is_lab_incharge(profile_user),
    }
    
    return render(request, 'products/user_profile.html', {
        'profile_user': profile_user,
        'custom_profile': custom_profile,
        'user_roles': user_roles,
        'stream_access': stream_access,
        'accessible_streams': accessible_streams,
        'permissions': permissions,
        'is_own_profile': profile_user == request.user,
        'can_edit': request.user == profile_user or can_manage_users(request.user),
        'selected_stream': 'profile'
    })

@login_required
def custom_password_change(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')
        user = request.user
        errors = []
        # Validate old password
        if not user.check_password(old_password):
            errors.append('Old password is incorrect.')
        # Validate new password match
        if new_password1 != new_password2:
            errors.append('New passwords do not match.')
        # Validate password length
        if len(new_password1) < 6:
            errors.append('New password must be at least 6 characters.')
        # Add more password validation as needed
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'products/password_change_form.html', {'selected_stream': 'HIC'})
        user.set_password(new_password1)
        user.save()
        update_session_auth_hash(request, user)
        return redirect('password_change_done')
    return render(request, 'products/password_change_form.html', {'selected_stream': 'HIC'})

@login_required
def sub_level_list(request, stream=None):
    edit_id = request.GET.get('edit_id')
    new_name = request.GET.get('new_name')
    user = request.user if request.user.is_authenticated else None
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    if edit_id and new_name:
        sub = SubLevel.objects.filter(id=edit_id, stream=stream_obj).first()
        if sub:
            old_name = sub.name
            sub.name = new_name
            sub.save()
            SubLevelHistory.objects.create(
                sublevel=sub,
                action='Edited',
                by=user.username if user else 'Unknown',
                details=f"Changed name from '{old_name}' to '{new_name}'"
            )
            return redirect(request.path)
    if request.method == 'POST':        
        if 'delete_id' in request.POST:
            delete_id = request.POST.get('delete_id')
            if delete_id:
                SubLevel.objects.filter(id=delete_id, stream=stream_obj).delete()
                return redirect(request.path)
        if 'sublevel_id' in request.POST:
            # Update counts
            sub = SubLevel.objects.filter(id=request.POST['sublevel_id'], stream=stream_obj).first()
            if sub:
                old_in_stock, old_in_use, old_scraped = sub.in_stock, sub.in_use, sub.scraped
                new_in_stock = int(request.POST.get('in_stock', 0))
                new_in_use = int(request.POST.get('in_use', 0))
                new_scraped = int(request.POST.get('scraped', 0))
                changes = []
                if old_in_stock != new_in_stock:
                    changes.append(f"Changed In Stock from {old_in_stock} to {new_in_stock}")
                if old_in_use != new_in_use:
                    changes.append(f"Changed In Use from {old_in_use} to {new_in_use}")
                if old_scraped != new_scraped:
                    changes.append(f"Changed Scraped from {old_scraped} to {new_scraped}")
                sub.in_stock = new_in_stock
                sub.in_use = new_in_use
                sub.scraped = new_scraped
                sub.save()
                if changes:
                    SubLevelHistory.objects.create(
                        sublevel=sub,
                        action='Edited',
                        by=user.username if user else 'Unknown',
                        details='; '.join(changes)
                    )
        elif 'subitem_name' in request.POST:
            name = request.POST.get('subitem_name')
            if name:                # Prevent duplicate sublevel name (case-insensitive, per stream)
                exists = SubLevel.objects.filter(name__iexact=name.strip(), stream=stream_obj).exists()
                if exists:
                    from django.contrib import messages
                    messages.error(request, f"Sublevel '{name}' already exists.")
                else:
                    sub = SubLevel.objects.create(name=name.strip(), stream=stream_obj)
                    SubLevelHistory.objects.create(
                        sublevel=sub,
                        action='Created',
                        by=user.username if user else 'Admin',
                        details='Initial creation'
                    )
            return redirect(request.path)
    subitems = SubLevel.objects.filter(stream=stream_obj) if stream else SubLevel.objects.all()
    # Attach history for each subitem
    for sub in subitems:
        sub.history_list = list(sub.history.order_by('at').values('action', 'by', 'at', 'details'))
    # Prepare JSON for all sublevel histories
    import json
    sublevel_history_dict = {str(sub.id): sub.history_list for sub in subitems}
    sublevel_history_json = json.dumps(sublevel_history_dict, default=str)
    return render(request, 'products/sub_level_list.html', {
        'subitems': subitems,
        'selected_stream': stream,
        'sublevel_history_json': sublevel_history_json,
    })

@require_POST
def delete_sub_level(request, stream, sublevel_id):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    SubLevel.objects.filter(id=sublevel_id, stream=stream_obj).delete()
    return redirect('sub_level_list_stream', stream=stream)

@login_required
def sub_level_tool_list(request, stream=None):
    edit_id = request.GET.get('edit_id')
    new_name = request.GET.get('new_name')
    user = request.user if request.user.is_authenticated else None
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    if edit_id and new_name:
        tool = SubLevelTool.objects.filter(id=edit_id, stream=stream_obj).first()
        if tool:
            old_name = tool.name
            tool.name = new_name
            tool.save()
            SubLevelToolHistory.objects.create(
                subleveltool=tool,
                action='Edited',
                by=user.username if user else 'Unknown',
                details=f"Changed name from '{old_name}' to '{new_name}'"
            )
            return redirect(request.path)
    if request.method == 'POST':        
        if 'delete_id' in request.POST:
            delete_id = request.POST.get('delete_id')
            if delete_id:
                SubLevelTool.objects.filter(id=delete_id, stream=stream_obj).delete()
                return redirect(request.path)
        if 'subleveltool_id' in request.POST:
            # Update counts
            tool = SubLevelTool.objects.filter(id=request.POST['subleveltool_id'], stream=stream_obj).first()
            if tool:
                old_in_stock, old_in_use, old_scraped = tool.in_stock, tool.in_use, tool.scraped
                new_in_stock = int(request.POST.get('in_stock', 0))
                new_in_use = int(request.POST.get('in_use', 0))
                new_scraped = int(request.POST.get('scraped', 0))
                changes = []
                if old_in_stock != new_in_stock:
                    changes.append(f"Changed In Stock from {old_in_stock} to {new_in_stock}")
                if old_in_use != new_in_use:
                    changes.append(f"Changed In Use from {old_in_use} to {new_in_use}")
                if old_scraped != new_scraped:
                    changes.append(f"Changed Scraped from {old_scraped} to {new_scraped}")
                tool.in_stock = new_in_stock
                tool.in_use = new_in_use
                tool.scraped = new_scraped
                tool.save()
                if changes:
                    SubLevelToolHistory.objects.create(
                        subleveltool=tool,
                        action='Edited',
                        by=user.username if user else 'Unknown',
                        details='; '.join(changes)
                    )
        elif 'subtool_name' in request.POST:
            name = request.POST.get('subtool_name')
            if name:                # Prevent duplicate subleveltool name (case-insensitive, per stream)
                exists = SubLevelTool.objects.filter(name__iexact=name.strip(), stream=stream_obj).exists()
                if exists:
                    from django.contrib import messages
                    messages.error(request, f"Subtool '{name}' already exists.")
                else:
                    tool = SubLevelTool.objects.create(name=name.strip(), stream=stream_obj)
                    SubLevelToolHistory.objects.create(
                        subleveltool=tool,
                        action='Created',
                        by=user.username if user else 'Admin',
                        details='Initial creation'
                    )
            return redirect(request.path)
    subtools = SubLevelTool.objects.filter(stream=stream_obj) if stream else SubLevelTool.objects.all()
    # Attach history for each subtool
    for tool in subtools:
        tool.history_list = list(tool.history.order_by('at').values('action', 'by', 'at', 'details'))
    # Prepare JSON for all subleveltool histories
    import json
    subleveltool_history_dict = {str(tool.id): tool.history_list for tool in subtools}
    subleveltool_history_json = json.dumps(subleveltool_history_dict, default=str)
    return render(request, 'products/sub_level_tool_list.html', {
        'subtools': subtools,
        'selected_stream': stream,
        'subleveltool_history_json': subleveltool_history_json,
    })

@require_POST
def delete_sub_level_tool(request, stream, subleveltool_id):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    SubLevelTool.objects.filter(id=subleveltool_id, stream=stream_obj).delete()
    return redirect('sub_level_tool_list_stream', stream=stream)

@login_required
@csrf_exempt
@require_POST
def bulk_delete_subleveltools(request, stream=None):
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        
        # Handle empty stream parameter
        if not stream or stream.strip() == '':
            stream = 'HIC'
        
        # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        
        deleted_count = SubLevelTool.objects.filter(id__in=ids, stream=stream_obj).delete()[0]
        return JsonResponse({'success': True, 'deleted': deleted_count})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@csrf_exempt
@require_POST
def bulk_update_subleveltools(request, stream=None):
    try:
        data = json.loads(request.body)
        notes = data.get('notes', {})
        user = request.user if request.user.is_authenticated else None
        
        # Handle empty stream parameter
        if not stream or stream.strip() == '':
            stream = 'HIC'
        
        # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        
        for tool_id, note_text in notes.items():
            tool = SubLevelTool.objects.filter(id=tool_id, stream=stream_obj).first()
            if tool:
                old_note = tool.note or ''
                tool.note = note_text
                tool.save()
                if old_note != note_text:
                    SubLevelToolHistory.objects.create(
                        subleveltool=tool,
                        action='Edited',
                        by=user.username if user else 'Unknown',
                        details=f"Updated note from '{old_note}' to '{note_text}'"
                    )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_subleveltools(request, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    subtools = SubLevelTool.objects.filter(stream=stream_obj)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sub-Level Tools'
    
    # Headers
    headers = ['Tool Name', 'In Stock', 'In Use', 'Scraped', 'Stream', 'Note', 'Last Modified By', 'Last Modified Date']
    ws.append(headers)
    
    # Data rows
    for tool in subtools:
        last_history = tool.history.order_by('-at').first()
        last_by = last_history.by if last_history else ''
        last_at = last_history.at.strftime('%Y-%m-%d %H:%M:%S') if last_history else ''
        ws.append([
            tool.name,
            tool.in_stock,
            tool.in_use,
            tool.scraped,
            tool.stream or '',
            tool.note or '',
            last_by,
            last_at
        ])
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    
    # Create response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_{stream}_Sub_Level_Tools.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@csrf_exempt
def get_nc_details(request, stream=None):
    system_id = request.GET.get('system_id')
    try:
        system = System.objects.get(pk=system_id)
        return JsonResponse({'details': system.nc_details or ''})
    except System.DoesNotExist:
        return JsonResponse({'details': ''})

@login_required
@csrf_exempt
def save_nc_details(request, stream=None):
    if request.method == 'POST':
        system_id = request.POST.get('system_id')
        details = request.POST.get('details', '')
        try:
            system = System.objects.get(pk=system_id)
            system.nc_details = details
            system.save()
            return JsonResponse({'success': True})
        except System.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'System not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

from django.views.decorators.http import require_GET

@login_required
@require_GET
def get_all_nc_details(request, stream=None):
    """
    Returns all systems' 12NC details for the given stream as JSON:
    [{id, name, details}]
    """
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    systems = System.objects.filter(stream=stream_obj).order_by('name')
    data = []
    for sys in systems:
        data.append({
            'id': sys.id,
            'name': sys.name,
            'details': sys.nc_details or ''
        })
    return JsonResponse({'systems': data})

@login_required
@csrf_exempt
def mark_notifications_read(request):
    if request.method == "POST":
        request.user.notifications.filter(is_read=False).update(is_read=True)
        return JsonResponse({"status": "ok"})
    return JsonResponse({"status": "error"}, status=400)

@login_required
@require_GET
def get_system_metrics(request, stream=None):
    try:
        # Resolve stream name to Stream object (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')

        # Calculate total systems
        total_systems = System.objects.filter(stream=stream_obj).count()
        active_systems = System.objects.filter(stream=stream_obj, status='Active').count()
        blocked_systems = SystemAllocation.objects.filter(
            stream=stream_obj,
            end_date__gt=timezone.now()
        ).count()

        # Calculate average utilization
        systems = System.objects.filter(stream=stream_obj)
        total_utilization = sum([s.utilization_percentage for s in systems if hasattr(s, 'utilization_percentage')])
        avg_utilization = round(total_utilization / total_systems if total_systems > 0 else 0)

        # Get usage trends for last 7 days
        end_date = timezone.now()
        start_date = end_date - timedelta(days=7)
        daily_usage = SystemAllocation.objects.filter(
            stream=stream_obj,
            start_date__gte=start_date,
            end_date__lte=end_date
        ).values('start_date__date').annotate(count=Count('id'))

        usage_trends = []
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.date().strftime('%Y-%m-%d')
            count = 0
            for usage in daily_usage:
                if usage['start_date__date'].strftime('%Y-%m-%d') == date_str:
                    count = usage['count']
                    break
            usage_trends.append({
                'date': date_str,
                'value': count
            })
            current_date += timedelta(days=1)

        # Get most active systems
        most_active = System.objects.filter(stream=stream_obj).order_by('-utilization_percentage')[:5]
        most_active_data = [{
            'name': system.name,
            'usage_hours': round(system.utilization_percentage * 24 / 100, 1)  # Convert percentage to hours
        } for system in most_active]

        return JsonResponse({
            'success': True,
            'total_systems': total_systems,
            'active_systems': active_systems,
            'blocked_systems': blocked_systems,
            'utilization': avg_utilization,
            'usage_trends': usage_trends,
            'most_active': most_active_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

# @login_required
# @require_GET
# def system_metrics_api(request, stream=None):
#     try:
#         # Calculate total systems
#         total_systems = System.objects.filter(stream=stream).count()
        
#         # Calculate active and blocked systems
#         active_systems = System.objects.filter(stream=stream, status='Active').count()
#         blocked_systems = SystemAllocation.objects.filter(
#             stream=stream, 
#             end_date__gt=timezone.now()
#         ).count()
        
#         # Calculate average utilization
#         current_time = timezone.now()
#         start_time = current_time - timedelta(days=7)
        
#         system_utilization = SystemAllocation.objects.filter(
#             stream=stream,
#             start_date__gte=start_time,
#             end_date__lte=current_time
#         ).aggregate(
#             avg_utilization=Avg(
#                 (F('end_date') - F('start_date')) / timedelta(hours=24)
#             )
#         )['avg_utilization'] or 0
        
#         # Calculate 7-day usage trends
#         usage_trends = []
#         for i in range(7):
#             day = current_time - timedelta(days=i)
#             count = SystemAllocation.objects.filter(
#                 stream=stream,
#                 start_date__date=day.date()
#             ).count()
#             usage_trends.append({
#                 'date': day.strftime('%Y-%m-%d'),
#                 'count': count
#             })
        
#         # Get most active systems
#         most_active = SystemAllocation.objects.filter(
#             stream=stream,
#             start_date__gte=start_time
#         ).values('system__name')\
#          .annotate(usage_time=Count('id'))\
#          .order_by('-usage_time')[:5]
        
#         most_active_data = [{
#             'name': item['system__name'],
#             'usage_time': item['usage_time'],
#             'utilization': round((item['usage_time'] / 7) * 100, 1)
#         } for item in most_active]

#         return JsonResponse({
#             'success': True,
#             'total_systems': total_systems,
#             'active_systems': active_systems,
#             'blocked_systems': blocked_systems,
#             'utilization': round(system_utilization * 100, 1),
#             'usage_trends': usage_trends,
#             'most_active': most_active_data
#         })
#     except Exception as e:
#         return JsonResponse({
#             'success': False,
#             'error': str(e)
#         })

def test_repo_view(request):
    streams = list(Stream.objects.values_list('name', flat=True).order_by('name'))
    if 'PIC' not in streams:
        Stream.objects.get_or_create(name='PIC')
        streams.append('PIC')
    if 'HIC' not in streams:
        Stream.objects.get_or_create(name='HIC')
        streams.append('HIC')
    def stream_sort_key(s):
        if s == 'PIC':
            return (0, s)
        elif s == 'HIC':
            return (1, s)
        return (2, s)
    streams = sorted(set(streams), key=stream_sort_key)
    zenition_products = ZenitionProduct.objects.all().order_by('name')
    # Handle legacy excel upload
    if request.method == 'POST' and request.FILES.get('legacy_excel'):
        excel_file = request.FILES['legacy_excel']
        selected_stream = request.POST.get('selected_stream', streams[0] if streams else '')
        upload = LegacyExcelUpload.objects.create(
            stream=selected_stream,
            file=excel_file,
            uploaded_by=request.user
        )
        # Parse preview data
        try:
            df = pd.read_excel(upload.file.path)
            preview_html = df.head(20).to_html(index=False)
            upload.preview_data = preview_html
            upload.save()
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error parsing Excel: {str(e)}'})
        return JsonResponse({'success': True, 'upload_id': upload.id})
    # Handle ZenitionProduct add
    if request.method == 'POST' and request.POST.get('zenition_name'):
        name = request.POST.get('zenition_name').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Name required'})
        if ZenitionProduct.objects.filter(name__iexact=name).exists():
            return JsonResponse({'success': False, 'error': 'Product already exists'})
        zp = ZenitionProduct.objects.create(name=name)
        zenition_products = ZenitionProduct.objects.all().order_by('name')
        # Return updated list for dropdown
        return JsonResponse({'success': True, 'zenition_products': [{'id': z.id, 'name': z.name} for z in zenition_products]})
    # Handle ZenitionProduct Excel upload
    if request.method == 'POST' and request.FILES.get('zenition_excel'):
        zenition_product_id = request.POST.get('zenition_product_id')
        selected_stream = request.POST.get('selected_stream', streams[0] if streams else '')
        try:
            zp = ZenitionProduct.objects.get(id=zenition_product_id)
        except ZenitionProduct.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Zenition Product not found'})
        excel_file = request.FILES['zenition_excel']
        upload = LegacyExcelUpload.objects.create(
            stream=selected_stream,
            file=excel_file,
            uploaded_by=request.user
        )
        # Parse preview data
        try:
            df = pd.read_excel(upload.file.path)
            preview_html = df.head(20).to_html(index=False)
            upload.preview_data = preview_html
            upload.save()
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error parsing Excel: {str(e)}'})
        return JsonResponse({'success': True, 'upload_id': upload.id})
    return render(request, 'products/test_repo.html', {
        'streams': streams,
        'selected_stream': streams[0] if streams else '',
        'zenition_products': zenition_products
    })

@login_required
@require_GET
def preview_legacy_excel(request, upload_id):
    try:
        excel = LegacyExcelUpload.objects.get(id=upload_id)
        file_path = excel.file.path
        df = pd.read_excel(file_path)
        preview_html = df.head(40).to_html(index=False)
        return JsonResponse({'success': True, 'html': preview_html})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_GET
def preview_zenition_excel(request, upload_id):
    try:
        upload = LegacyExcelUpload.objects.get(id=upload_id)
        html = upload.preview_data or '<div>No preview available.</div>'
        return JsonResponse({'success': True, 'html': html})
    except LegacyExcelUpload.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'File not found'})

import pandas as pd
from django.views.decorators.http import require_GET

def get_merge_pickle_path(legacy_id, zenition_id):
    temp_dir = os.path.join(settings.BASE_DIR, 'products', 'temp')
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    return os.path.join(temp_dir, f'merged_{legacy_id}_{zenition_id}.pkl')

@login_required
@require_GET
def merge_excels(request):
    legacy_id = request.GET.get('legacy_id')
    zenition_id = request.GET.get('zenition_id')
    add_field = request.GET.get('add_field')
    delete_field = request.GET.get('delete_field')
    add_row_testcase = request.GET.get('add_row_testcase')
    add_row_result = request.GET.get('add_row_result')
    add_row_comment = request.GET.get('add_row_comment')
    save_table_data = request.GET.get('save_table_data')
    delete_row_index = request.GET.get('delete_row_index')
    if not legacy_id or not zenition_id:
        return JsonResponse({'success': False, 'error': 'Both file IDs required'})
    try:
        pickle_path = get_merge_pickle_path(legacy_id, zenition_id)
        # If pickle exists, load it; else, create from Excel files
        if os.path.exists(pickle_path):
            with open(pickle_path, 'rb') as f:
                merged_df = pickle.load(f)
        else:
            from .models import LegacyExcelUpload
            legacy_upload = LegacyExcelUpload.objects.get(id=legacy_id)
            zenition_upload = LegacyExcelUpload.objects.get(id=zenition_id)
            df1 = pd.read_excel(legacy_upload.file.path)
            df2 = pd.read_excel(zenition_upload.file.path)
            merged_df = pd.concat([df1, df2], ignore_index=True)
        # Add field(s)
        if add_field:
            for field in add_field.split(','):
                field = field.strip()
                if field:
                    merged_df[field] = ''
        # Add new row if requested
        if add_row_testcase or add_row_result or add_row_comment:
            for col in ['BackEnd Test cases', 'Results', 'Comments']:
                if col not in merged_df.columns:
                    merged_df[col] = ''
            new_row = {
                'BackEnd Test cases': add_row_testcase or '',
                'Results': add_row_result or '',
                'Comments': add_row_comment or ''
            }
            merged_df = pd.concat([merged_df, pd.DataFrame([new_row])], ignore_index=True)
        # Save edited table data
        if save_table_data:
            import json
            try:
                table_data = json.loads(save_table_data)
                for i, row in enumerate(table_data):
                    if i < len(merged_df):
                        merged_df.iloc[i, merged_df.columns.get_loc('BackEnd Test cases')] = row[0]
                        merged_df.iloc[i, merged_df.columns.get_loc('Results')] = row[1]
                        merged_df.iloc[i, merged_df.columns.get_loc('Comments')] = row[2]
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Edit error: {str(e)}'})
        # Delete row by index
        if delete_row_index is not None:
            try:
                idx = int(delete_row_index)
                merged_df = merged_df.drop(idx).reset_index(drop=True)
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Delete row error: {str(e)}'})
        # Delete field(s)
        if delete_field:
            for field in delete_field.split(','):
                field = field.strip()
                if field in merged_df.columns:
                    merged_df = merged_df.drop(columns=[field])
        # Save the updated DataFrame back to pickle
        with open(pickle_path, 'wb') as f:
            pickle.dump(merged_df, f)
        preview_html = merged_df.head(40).to_html(index=False)
        return JsonResponse({'success': True, 'html': preview_html})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Merge error: {str(e)}'})

from django.http import JsonResponse
from .models import LegacyExcelUpload
from django.conf import settings

@login_required
@require_GET
def preview_legacy_excel_by_filename(request, filename):
    folder = os.path.join(settings.BASE_DIR, 'product_images', 'legacy_excels')
    file_path = os.path.join(folder, filename)
    if not os.path.exists(file_path):
        return JsonResponse({'success': False, 'error': 'File not found'})
    try:
        df = pd.read_excel(file_path)
        preview_html = df.head(40).to_html(index=False)
        return JsonResponse({'success': True, 'html': preview_html})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_GET
def fetch_legacy_excels(request):
    # Return all LegacyExcelUpload objects, not files from folder
    excels = LegacyExcelUpload.objects.all().order_by('-uploaded_at')
    files = []
    for excel in excels:
        files.append({
            'file_name': excel.file.name.split('/')[-1],
            'id': excel.id,  # Use database PK as id
        })
    return JsonResponse({'success': True, 'excels': files})

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import TestEnvironment  # Make sure this model exists

@csrf_exempt
@login_required
def save_test_environment(request):
    if request.method == 'POST':
        data = request.POST
        env = TestEnvironment.objects.create(
            mvs_binaries=data.get('mvs_binaries', ''),
            mvs_os=data.get('mvs_os', ''),
            stand_binaries=data.get('stand_binaries', ''),
            stand_os=data.get('stand_os', ''),
            apps_pc_binaries=data.get('apps_pc_binaries', ''),
            apps_pc_os=data.get('apps_pc_os', ''),
            test_environment=data.get('test_environment', ''),
        )
        return JsonResponse({'success': True, 'id': env.id})
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse

@login_required
def export_data(request):
    legacy_id = request.GET.get('legacy_id')
    zenition_id = request.GET.get('zenition_id')
    # Get merged excel data (from pickle or re-merge)
    pickle_path = get_merge_pickle_path(legacy_id, zenition_id)
    if os.path.exists(pickle_path):
        with open(pickle_path, 'rb') as f:
            merged_df = pickle.load(f)
    else:
        merged_df = pd.DataFrame()
    # Get latest Test Environment data
    test_env = TestEnvironment.objects.last()
    # Get selected Zenition Product name
    zenition_product_id = request.GET.get('zenition_id')
    from .models import ZenitionProduct
    try:
        zenition_product = ZenitionProduct.objects.get(id=zenition_product_id)
        zenition_name = zenition_product.name
    except ZenitionProduct.DoesNotExist:
        zenition_name = "TestRepoZenitionProduct"

    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Data"

    # Row 1: TestRepoZenitionProduct name
    ws.append([zenition_name])
    # Row 2: Empty
    ws.append([])
    # Rows 3-8: Test Environment key-value pairs
    env_pairs = [
        ("MVS Binaries", getattr(test_env, "mvs_binaries", "")),
        ("MVS OS", getattr(test_env, "mvs_os", "")),
        ("Stand Binaries", getattr(test_env, "stand_binaries", "")),
        ("Stand OS", getattr(test_env, "stand_os", "")),
        ("APPS PC Binaries", getattr(test_env, "apps_pc_binaries", "")),
        ("APPS PC OS", getattr(test_env, "apps_pc_os", "")),
    ]
    for key, value in env_pairs:
        ws.append([key, value])
    # Row 9: Empty
    ws.append([])
    # Row 10+: Merged Excel Preview
    if not merged_df.empty:
        ws.append(list(merged_df.columns))
        for row in merged_df.itertuples(index=False):
            ws.append(list(row))
    else:
        ws.append(["No merged data available"])

    # Set column widths for readability
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28

    # Prepare response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exported_data.xlsx'
    return response

@login_required
@require_GET
def fetch_zenition_excels(request):
    # Filter by ZenitionProduct if needed
    zenition_excels = LegacyExcelUpload.objects.filter(stream='Zenition').order_by('-uploaded_at')
    files = []
    for excel in zenition_excels:
        files.append({
            'file_name': excel.file.name.split('/')[-1],
            'id': excel.id,
        })
    return JsonResponse({'success': True, 'excels': files})

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import PersonalTask
from django.http import JsonResponse

@login_required
def personal_trackboard(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            title = request.POST.get('title', '').strip()
            if title:
                PersonalTask.objects.create(user=request.user, title=title)
                return JsonResponse({'success': True})
            return JsonResponse({'success': False, 'error': 'Title required'})
        elif action == 'update':
            task_id = request.POST.get('task_id')
            status = request.POST.get('status')
            try:
                task = PersonalTask.objects.get(id=task_id, user=request.user)
                if status in dict(PersonalTask.STATUS_CHOICES):
                    task.status = status
                    task.save()
                    return JsonResponse({'success': True})
            except PersonalTask.DoesNotExist:
                pass
            return JsonResponse({'success': False, 'error': 'Task not found'})
        elif action == 'delete':
            task_id = request.POST.get('task_id')
            try:
                task = PersonalTask.objects.get(id=task_id, user=request.user)
                task.delete()
                return JsonResponse({'success': True})
            except PersonalTask.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Task not found'})
    # GET: Render the page with user's tasks
    tasks = PersonalTask.objects.filter(user=request.user).order_by('created_at')
    
    # Prepare user context data
    user_context = {
        'username': request.user.username,
        'first_name': request.user.first_name,
        'last_name': request.user.last_name,
        'full_name': request.user.get_full_name() or request.user.username,
        'initials': (request.user.first_name[:1] if request.user.first_name else request.user.username[:1]).upper() + 
                   (request.user.last_name[:1] if request.user.last_name else '').upper(),
        'is_admin': request.user.is_superuser
    }
    
    # Try to get custom user profile for role information
    try:
        custom_profile = request.user.custom_profile
        user_context['role'] = custom_profile.get_roles_display() if custom_profile.user_roles.exists() else ''
    except:
        user_context['role'] = ''
    
    # Prepare tasks data for JavaScript
    tasks_data = [
        {
            'id': task.id,
            'title': task.title,
            'status': task.status,
            'created_at': task.created_at.isoformat(),
            'updated_at': task.updated_at.isoformat(),
        }
        for task in tasks
    ]
    
    context = {
        'tasks': tasks,
        'user_context': user_context,
        'tasks_json': json.dumps(tasks_data),
    }
    
    return render(request, 'products/personal_trackboard.html', context)

@login_required
@user_passes_test(is_super_admin)
def usage_tracking(request):
    """
    Display a usage tracking dashboard showing application usage analytics
    """
    # Default to showing last 30 days
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    previous_start_date = start_date - timedelta(days=30)
      # Get usage statistics
    total_users = User.objects.filter(is_active=True).count()
    new_users = User.objects.filter(date_joined__gte=start_date).count()
    previous_new_users = User.objects.filter(date_joined__range=(previous_start_date, start_date)).count()
    
    # Calculate percentage change for new users
    if previous_new_users > 0:
        new_users_percent = round(((new_users - previous_new_users) / previous_new_users) * 100)
    else:
        new_users_percent = 100 if new_users > 0 else 0
        
    # Get active users
    active_users = UsageTracking.objects.filter(
        timestamp__gte=start_date
    ).values('user').distinct().count()
    
    previous_active_users = UsageTracking.objects.filter(
        timestamp__range=(previous_start_date, start_date)
    ).values('user').distinct().count()
    
    # Calculate percentage change for active users
    if previous_active_users > 0:
        active_users_change = round(((active_users - previous_active_users) / previous_active_users) * 100)
    else:
        active_users_change = 100 if active_users > 0 else 0
        
    # Total page views
    total_views = UsageTracking.objects.filter(timestamp__gte=start_date).count()
    previous_total_views = UsageTracking.objects.filter(timestamp__range=(previous_start_date, start_date)).count()
    
    # Calculate percentage change for views
    if previous_total_views > 0:
        views_percent_change = round(((total_views - previous_total_views) / previous_total_views) * 100)
    else:
        views_percent_change = 100 if total_views > 0 else 0
        
    # Get top users
    top_users = []
    user_stats = UsageTracking.objects.filter(
        timestamp__gte=start_date
    ).values('user').annotate(page_views=Count('id')).order_by('-page_views')[:10]
    
    for stat in user_stats:
        user = User.objects.get(pk=stat['user'])
        last_activity = UsageTracking.objects.filter(user=user).order_by('-timestamp').first()
          # Calculate actual average session time for this user
        user_sessions = UsageTracking.objects.filter(user=user).order_by('timestamp')
        total_duration = 0
        session_count = 0
        
        # Simple session calculation: if gap between records > 30 min, consider it a new session
        if user_sessions.count() > 1:
            prev_timestamp = None
            session_durations = []
            
            for activity in user_sessions:
                if prev_timestamp:
                    # Calculate time difference in minutes
                    time_diff = (activity.timestamp - prev_timestamp).total_seconds() / 60
                    
                    if time_diff <= 30:  # Same session (less than 30 min gap)
                        total_duration += time_diff
                    else:  # New session
                        if total_duration > 0:
                            session_durations.append(total_duration)
                            total_duration = 0
                            session_count += 1
                
                prev_timestamp = activity.timestamp
            
            # Add the last session if there's any duration
            if total_duration > 0:
                session_durations.append(total_duration)
                session_count += 1
            
            # Calculate the average
            avg_time = sum(session_durations) / max(len(session_durations), 1)
            avg_session_time_str = f"{int(avg_time)} mins"
        else:
            avg_session_time_str = "N/A"  # Not enough data
        
        top_users.append({
            'username': user.username,
            'page_views': stat['page_views'],
            'last_active': last_activity.timestamp if last_activity else 'N/A',
            'avg_session_time': avg_session_time_str
        })
          # Calculate average session time across all users
    all_sessions = UsageTracking.objects.filter(timestamp__gte=start_date).order_by('user', 'timestamp')
    total_duration = 0
    session_count = 0
    
    current_user = None
    prev_timestamp = None
    session_durations = []
    
    for activity in all_sessions:
        # If user changed, reset the session tracking
        if current_user != activity.user_id:
            if current_user is not None and total_duration > 0:
                session_durations.append(total_duration)
                session_count += 1
            
            current_user = activity.user_id
            prev_timestamp = None
            total_duration = 0
        
        if prev_timestamp:
            # Calculate time difference in minutes
            time_diff = (activity.timestamp - prev_timestamp).total_seconds() / 60
            
            if time_diff <= 30:  # Same session (less than 30 min gap)
                total_duration += time_diff
            else:  # New session
                if total_duration > 0:
                    session_durations.append(total_duration)
                    session_count += 1
                    total_duration = 0
        
        prev_timestamp = activity.timestamp
    
    # Add the last session if there's any duration
    if total_duration > 0:
        session_durations.append(total_duration)
        session_count += 1
    
    # Calculate the overall average
    if session_durations:
        current_avg_time = sum(session_durations) / len(session_durations)
        avg_session_time = f"{int(current_avg_time)} mins"
        
        # Calculate previous period for comparison
        prev_period_start = start_date - (end_date - start_date)
        prev_period_end = start_date
        
        # Get previous period sessions
        prev_sessions = UsageTracking.objects.filter(
            timestamp__gte=prev_period_start,
            timestamp__lt=prev_period_end
        ).order_by('user', 'timestamp')
        
        # Similar calculation for previous period (simplified for brevity)
        prev_session_durations = []
        # Similar logic as above to calculate prev_session_durations
        
        if prev_session_durations:
            prev_avg_time = sum(prev_session_durations) / len(prev_session_durations)
            if prev_avg_time > 0:
                session_time_change = int(((current_avg_time - prev_avg_time) / prev_avg_time) * 100)
            else:
                session_time_change = 100  # If previously 0, it's a 100% increase
        else:
            session_time_change = 100  # No previous data, consider it 100% increase
    else:
        avg_session_time = 'N/A'
        session_time_change = 0
    
    context = {
        'total_users': total_users,
        'new_users_percent': new_users_percent,
        'active_users': active_users,
        'active_users_change': active_users_change,
        'total_views': total_views,
        'views_percent_change': views_percent_change,
        'avg_session_time': avg_session_time,
        'session_time_change': session_time_change,
        'top_users': top_users,
        'selected_stream': 'HIC',  # Default stream
    }
    
    return render(request, 'products/usage_tracking.html', context)

@login_required
@user_passes_test(is_super_admin)
def usage_tracking_data(request):
    """
    API endpoint to get usage tracking data for charts
    """
    date_range = request.GET.get('range', '30')
    
    if date_range == 'custom':        # Handle custom date range
        start_str = request.GET.get('start')
        end_str = request.GET.get('end')
        
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_str, '%Y-%m-%d')
            # Add one day to end date to include the entire day
            end_date = end_date + timedelta(days=1)
        except (ValueError, TypeError):
            # If dates are invalid, fall back to 30 days
            end_date = timezone.now()
            start_date = end_date - timedelta(days=30)
    else:
        # Calculate based on selected range
        days = int(date_range)
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
    
    # Make timezone-aware
    if timezone.is_naive(start_date):
        start_date = timezone.make_aware(start_date)
    if timezone.is_naive(end_date):
        end_date = timezone.make_aware(end_date)
    
    # Calculate previous period for comparison
    period_length = end_date - start_date
    previous_start_date = start_date - period_length
    previous_end_date = start_date
    
    # Get statistics for current period
    total_views = UsageTracking.objects.filter(timestamp__gte=start_date, timestamp__lt=end_date).count()
    previous_total_views = UsageTracking.objects.filter(timestamp__gte=previous_start_date, timestamp__lt=previous_end_date).count()
    
    # Calculate percentage change for views
    if previous_total_views > 0:
        views_percent_change = round(((total_views - previous_total_views) / previous_total_views) * 100)
    else:
        views_percent_change = 100 if total_views > 0 else 0
    
    # Get active users
    active_users = UsageTracking.objects.filter(
        timestamp__gte=start_date, timestamp__lt=end_date
    ).values('user').distinct().count()
    
    previous_active_users = UsageTracking.objects.filter(
        timestamp__gte=previous_start_date, timestamp__lt=previous_end_date
    ).values('user').distinct().count()
    
    # Calculate percentage change for active users
    if previous_active_users > 0:
        active_users_change = round(((active_users - previous_active_users) / previous_active_users) * 100)
    else:
        active_users_change = 100 if active_users > 0 else 0
    
    # Get total users (all time)
    total_users = User.objects.filter(is_active=True).count()
    new_users = User.objects.filter(date_joined__gte=start_date, date_joined__lt=end_date).count()
    previous_new_users = User.objects.filter(date_joined__gte=previous_start_date, date_joined__lt=previous_end_date).count()
    
    # Calculate percentage change for new users
    if previous_new_users > 0:
        new_users_percent = round(((new_users - previous_new_users) / previous_new_users) * 100)
    else:
        new_users_percent = 100 if new_users > 0 else 0
    
    # Calculate average session time
    all_sessions = UsageTracking.objects.filter(
        timestamp__gte=start_date, timestamp__lt=end_date
    ).order_by('user', 'timestamp')
    
    session_durations = []
    current_user = None
    prev_timestamp = None
    total_duration = 0
    
    for activity in all_sessions:
        if current_user != activity.user_id:
            if current_user is not None and total_duration > 0:
                session_durations.append(total_duration)
            current_user = activity.user_id
            prev_timestamp = None
            total_duration = 0
        
        if prev_timestamp:
            time_diff = (activity.timestamp - prev_timestamp).total_seconds() / 60
            if time_diff <= 30:  # Same session
                total_duration += time_diff
            else:  # New session
                if total_duration > 0:
                    session_durations.append(total_duration)
                    total_duration = 0
        
        prev_timestamp = activity.timestamp
    
    if total_duration > 0:
        session_durations.append(total_duration)
    
    # Calculate averages
    if session_durations:
        current_avg_time = sum(session_durations) / len(session_durations)
        avg_session_time = f"{int(current_avg_time)} mins"
        
        # Previous period sessions
        prev_sessions = UsageTracking.objects.filter(
            timestamp__gte=previous_start_date, timestamp__lt=previous_end_date
        ).order_by('user', 'timestamp')
        
        prev_session_durations = []
        current_user = None
        prev_timestamp = None
        total_duration = 0
        
        for activity in prev_sessions:
            if current_user != activity.user_id:
                if current_user is not None and total_duration > 0:
                    prev_session_durations.append(total_duration)
                current_user = activity.user_id
                prev_timestamp = None
                total_duration = 0
            
            if prev_timestamp:
                time_diff = (activity.timestamp - prev_timestamp).total_seconds() / 60
                if time_diff <= 30:
                    total_duration += time_diff
                else:
                    if total_duration > 0:
                        prev_session_durations.append(total_duration)
                        total_duration = 0
            
            prev_timestamp = activity.timestamp
        
        if total_duration > 0:
            prev_session_durations.append(total_duration)
        
        if prev_session_durations:
            prev_avg_time = sum(prev_session_durations) / len(prev_session_durations)
            if prev_avg_time > 0:
                session_time_change = int(((current_avg_time - prev_avg_time) / prev_avg_time) * 100)
            else:
                session_time_change = 100
        else:
            session_time_change = 100
    else:
        avg_session_time = 'N/A'
        session_time_change = 0
    
    # Get daily usage data
    daily_data = generate_daily_usage_data(start_date, end_date)
    
    # Get most visited pages
    page_views_data = generate_page_views_data(start_date, end_date)
    
    # Get hourly activity data
    hourly_data = generate_hourly_activity_data(start_date, end_date)
    
    return JsonResponse({
        'daily_usage': daily_data,
        'page_views': page_views_data,
        'hourly_activity': hourly_data,
        'stats': {
            'total_views': total_views,
            'views_percent_change': views_percent_change,
            'active_users': active_users,
            'active_users_change': active_users_change,
            'total_users': total_users,
            'new_users_percent': new_users_percent,
            'avg_session_time': avg_session_time,
            'session_time_change': session_time_change,
        }
    })

# Helper functions for usage tracking data
def generate_daily_usage_data(start_date, end_date):
    """
    Generate daily usage data for the specified time range
    """
    from django.utils import timezone as dj_timezone
    # Ensure start/end are timezone-aware and work in current (local) timezone for display
    local_tz = dj_timezone.get_current_timezone()
    # Normalize to local midnight for start and end (end is exclusive)
    local_start = dj_timezone.localtime(start_date, local_tz).replace(hour=0, minute=0, second=0, microsecond=0)
    local_end = dj_timezone.localtime(end_date, local_tz).replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)

    date_range = (local_end - local_start).days
    dates = []
    views_counts = []
    users_counts = []

    # For each date in the range (using local dates for labels)
    for i in range(date_range):
        current_local = local_start + timedelta(days=i)
        next_local = current_local + timedelta(days=1)

        # Format the date for display using local date (this fixes off-by-one day labels)
        display_date = current_local.strftime('%Y-%m-%d')
        dates.append(display_date)

        # Convert local day boundaries back to UTC for querying the DB (timestamps are stored in UTC)
        try:
            current_utc = current_local.astimezone(dj_timezone.utc)
            next_utc = next_local.astimezone(dj_timezone.utc)
        except Exception:
            # Fallback: if astimezone fails, use the naive datetimes as-is (best-effort)
            current_utc = current_local
            next_utc = next_local

        # Count page views for this day
        day_views = UsageTracking.objects.filter(
            timestamp__gte=current_utc,
            timestamp__lt=next_utc
        ).count()
        views_counts.append(day_views)

        # Count unique users for this day
        day_users = UsageTracking.objects.filter(
            timestamp__gte=current_utc,
            timestamp__lt=next_utc
        ).values('user').distinct().count()
        users_counts.append(day_users)

    return {
        'dates': dates,
        'views': views_counts,
        'users': users_counts
    }

def generate_page_views_data(start_date, end_date):
    """
    Generate data for most visited pages, excluding the usage tracking page 
    and user-related pages (login, register, profile, etc.)
    """
    from django.db.models import Q
    
    # Create a combined Q object for all excluded terms (more precise, avoid over-filtering)
    exclude_terms = (
        Q(page_name__iexact='/usage-tracking/') |
        Q(page_name__iexact='/usage_tracking/') |
        Q(page_name__iexact='/favicon.ico') |
        Q(page_name__iexact='/accounts/login/') |
        Q(page_name__iexact='/accounts/logout/') |
        Q(page_name__iexact='/accounts/register/') |
        Q(page_name__iexact='/accounts/profile/') |
        Q(page_name__icontains='.well-known') |
        Q(page_name__icontains='appspecific')
    )
    
    page_counts = UsageTracking.objects.filter(
        timestamp__gte=start_date,
        timestamp__lte=end_date
    ).exclude(exclude_terms).values('page_name').annotate(count=Count('id')).order_by('-count')[:10]
    
    pages = []
    counts = []
    
    for item in page_counts:
        page_name = item['page_name']
        pages.append(page_name[:25] + '...' if len(page_name) > 25 else page_name)
        counts.append(item['count'])
    
    return {
        'pages': pages,
        'counts': counts
    }

def generate_hourly_activity_data(start_date, end_date):
    """
    Generate data for hourly user activity
    """
    hours = list(range(24))
    counts = [0] * 24  # Initialize with zeros
    
    # Get all records in the date range
    records = UsageTracking.objects.filter(
        timestamp__gte=start_date,
        timestamp__lte=end_date
    )
    
    # Count by hour
    for record in records:
        hour = record.timestamp.hour
        counts[hour] += 1
    
    # Format hours for display (e.g., "12 AM", "1 PM", etc.)
    hour_labels = []
    for h in hours:
        if h == 0:
            hour_labels.append("12 AM")
        elif h < 12:
            hour_labels.append(f"{h} AM")
        elif h == 12:
            hour_labels.append("12 PM")
        else:
            hour_labels.append(f"{h-12} PM")
    
    return {
        'hours': hour_labels,
        'counts': counts
    }


def get_client_ip(request):
    """Get client IP address from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

# API endpoints for dashboard real-time data
@csrf_exempt
@require_GET
def dashboard_api_data(request):
    """API endpoint for real-time dashboard data."""
    from .models import (Product, Category, System, Stream, 
                        SystemStatus, UserSession)
    from django.utils import timezone
    from datetime import timedelta
    
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    stream = request.GET.get('stream', 'HIC')
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'PIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'PIC' when missing.
    stream_obj = get_stream_or_404(stream, default='PIC')
    
    # Get current statistics
    total_products = Product.objects.count()  # Count products across all streams
    total_categories = Category.objects.filter(stream=stream_obj).count()
    
    # Calculate online users (active in last 15 minutes)
    cutoff_time = timezone.now() - timedelta(minutes=15)
    
    # First, mark old sessions as inactive
    UserSession.objects.filter(
        last_activity__lt=cutoff_time,
        is_active=True
    ).update(is_active=False)
    
    # Count truly active users
    online_users = UserSession.objects.filter(
        last_activity__gte=cutoff_time,
        is_active=True
    ).count()
    
    # Get system status
    system_status = SystemStatus.objects.first()
    if not system_status:
        system_status = SystemStatus.objects.create(
            status='online',
            description='System operational'
        )
    data = {
        'total_products': total_products,
        'total_categories': total_categories,
        'online_users': online_users,
        'system_status': {
            'status': system_status.get_status_display(),
            'description': system_status.description,
            'uptime': system_status.uptime_percentage
        },
        'timestamp': timezone.now().isoformat()
    }
    
    return JsonResponse(data)

@csrf_exempt
@require_POST
def update_user_activity(request):
    """Update user activity for session tracking."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    session_key = request.session.session_key
    if session_key:
        from .models import UserSession
        from django.utils import timezone
        session_obj, created = UserSession.objects.get_or_create(
            session_key=session_key,
            defaults={
                'user': request.user,
                'ip_address': get_client_ip(request),
                'user_agent': request.META.get('HTTP_USER_AGENT', '')
            }
        )
        if not created:
            session_obj.last_activity = timezone.now()
            session_obj.save()
    
    return JsonResponse({'status': 'success'})

@login_required
@require_POST
def category_bulk_delete(request, stream=None):
    """Delete multiple categories sent as JSON list of ids in request body.
    Only superusers may perform this action.
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
        ids = payload.get('ids', [])
    except Exception:
        ids = request.POST.get('ids', '')
        ids = [int(x) for x in ids.split(',') if x]

    if not ids:
        return JsonResponse({'success': False, 'error': 'No IDs provided.'}, status=400)

    # Resolve stream param to a Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    # Query using the Stream instance rather than a raw string
    categories = Category.objects.filter(id__in=ids, stream=stream_obj).annotate(product_count=Count('products'))

    deleted_count = categories.count()
    categories.delete()
    return JsonResponse({'success': True, 'deleted': deleted_count})

@login_required
@require_POST
def category_export_csv(request, stream=None):
    """Export selected categories (ids) to a CSV. Accepts JSON body with ids list or form-encoded 'ids'."""
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
        ids = payload.get('ids', [])
    except Exception:
        ids = request.POST.get('ids', '')
        ids = [int(x) for x in ids.split(',') if x]

    if not ids:
        return JsonResponse({'success': False, 'error': 'No IDs provided.'}, status=400)

    # Resolve stream param to a Stream object (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')

    # Query using the Stream instance rather than a raw string
    categories = Category.objects.filter(id__in=ids, stream=stream_obj).annotate(product_count=Count('products'))

    # Build CSV in memory
    import csv
    from io import StringIO

    csvfile = StringIO()
    writer = csv.writer(csvfile)
    writer.writerow(['ID', 'Name', 'Serial Number', 'Product Count', 'Created At'])
    for c in categories:
        writer.writerow([c.id, c.name, c.serial_number, getattr(c, 'product_count', 0), c.created_at.strftime('%Y-%m-%d %H:%M:%S') if getattr(c, 'created_at', None) else ''])

    resp = HttpResponse(csvfile.getvalue(), content_type='text/csv')
    resp['Content-Disposition'] = f'attachment; filename="categories_{stream or "all"}.csv"'
    return resp

@login_required
@csrf_exempt
def bulk_delete_sublevels(request, stream=None):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        ids = data.get('ids', [])
        deleted = 0
        
        # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        
        for sub_id in ids:
            try:
                sub = SubLevel.objects.get(id=sub_id, stream=stream_obj)
                sub.delete()
                deleted += 1
            except SubLevel.DoesNotExist:
                continue
        return JsonResponse({'success': True, 'deleted': deleted})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
@csrf_exempt
def bulk_update_sublevels(request, stream=None):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        notes = data.get('notes', {})
        updated = 0
        
        # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        
        for sub_id, note in notes.items():
            try:
                sub = SubLevel.objects.get(id=sub_id, stream=stream_obj)
                sub.note = note
                sub.save()
                updated += 1
            except SubLevel.DoesNotExist:
                continue
        return JsonResponse({'success': True, 'updated': updated})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
@require_GET
def export_sublevels(request, stream=None):
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    sublevels = SubLevel.objects.filter(stream=stream_obj)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sublevels_export.csv"'
    writer = csv.writer(response)
    writer.writerow(['Name', 'In Stock', 'In Use', 'Scrapped', 'Note'])
    for sub in sublevels:
        writer.writerow([sub.name, sub.in_stock, sub.in_use, sub.scraped, sub.note or ''])
    return response

# API endpoints for system allocation real-time features

@login_required
@require_GET
def system_status_api(request, stream=None):
    """API endpoint to get current system status for real-time updates"""
    from django.utils import timezone
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    systems = System.objects.filter(stream=stream_obj).order_by('name')
    system_data = []
    
    for system in systems:
        system_data.append({
            'id': system.id,
            'name': system.name,
            'status': system.status,
            'health': system.health,
            'utilization': system.utilization_percentage,
            'last_updated': system.last_updated.isoformat() if system.last_updated else None
        })
    
    return JsonResponse({
        'success': True,
        'systems': system_data,
        'timestamp': timezone.now().isoformat()
    })

@login_required
@require_GET
def system_metrics_api(request, stream=None):
    """API endpoint to get system metrics data"""
    from django.utils import timezone
    from django.db.models import Count, Avg, Sum
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    systems = System.objects.filter(stream=stream_obj)
    
    # Calculate basic metrics
    total_systems = systems.count()
    active_systems = systems.filter(status='Active').count()
    blocked_systems = systems.exclude(status='Active').count()
    
    # Calculate average utilization
    avg_utilization = systems.aggregate(
        avg_util=Avg('utilization_percentage')
    )['avg_util'] or 0
    
    # Get most active systems (by utilization)
    most_active = systems.filter(
        utilization_percentage__gt=0
    ).order_by('-utilization_percentage')[:5]
    
    most_active_data = [{
        'name': system.name,
        'usage_hours': round(system.utilization_percentage * 24 * 30 / 100, 1)  # Convert percentage to hours
    } for system in most_active]
    
    # Generate usage trends (last 7 days)
    from datetime import datetime, timedelta
    now = timezone.now()
    usage_trends = []
    
    for i in range(7):
        date = now - timedelta(days=6-i)
        day_name = date.strftime('%a')
        
        # Calculate allocations for this day
        day_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        
        allocations_count = SystemAllocation.objects.filter(
            start_date__lt=day_end,
            end_date__gt=day_start
        ).count()
        
        usage_trends.append({
            'label': day_name,
            'value': allocations_count
        })
    
    return JsonResponse({
        'success': True,
        'total_systems': total_systems,
        'active_systems': active_systems,
        'blocked_systems': blocked_systems,
        'utilization': round(avg_utilization, 1),
        'most_active': most_active_data,
        'usage_trends': usage_trends
    })

@login_required
@require_GET
def get_historical_system_status(request, stream=None):
    """API endpoint to get historical system statuses for a specific date"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Get view_date parameter (format: YYYY-MM-DD)
    view_date_str = request.GET.get('view_date')
    if not view_date_str:
        return JsonResponse({
            'success': False,
            'error': 'view_date parameter required (format: YYYY-MM-DD)'
        }, status=400)
    
    try:
        view_date = datetime.strptime(view_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    
    # Handle empty stream parameter
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Resolve stream to Stream object (404 handled below)
    from django.http import Http404
    try:
        stream_obj = get_stream_or_404(stream, default='HIC')
    except Http404:
        return JsonResponse({
            'success': False,
            'error': f'Stream "{stream}" not found'
        }, status=404)
    
    # Get all systems in the stream
    systems = System.objects.filter(stream=stream_obj).order_by('name')
    
    # Fetch historical status for each system on the given date
    systems_data = []
    for system in systems:
        # Get historical status for the view_date
        view_date_obj = view_date  # This is already a date object
        
        # Create date range for the entire day (24 hours)
        # Start: 00:00:00 of view_date
        # End: 00:00:00 of next day (exclusive)
        day_start = timezone.make_aware(datetime.combine(view_date_obj, datetime.min.time()))
        day_end = day_start + timedelta(days=1)
        
        # First, try to find the most recent status on the view_date itself
        historical_status = SystemStatusHistory.objects.filter(
            system=system,
            updated_at__gte=day_start,
            updated_at__lt=day_end
        ).order_by('-updated_at').first()
        
        # If no status found on that exact day, look for the most recent status before that day
        if not historical_status:
            historical_status = SystemStatusHistory.objects.filter(
                system=system,
                updated_at__lt=day_start
            ).order_by('-updated_at').first()
        
        if historical_status:
            status_value = historical_status.status
            status_display = historical_status.get_status_display()
            description = historical_status.description
            assignee = str(historical_status.assignee) if historical_status.assignee else ''
        else:
            status_value = system.status
            status_display = system.get_status_display()
            description = system.description
            assignee = getattr(system, 'assignee', '')
        
        systems_data.append({
            'id': system.id,
            'name': system.name,
            'status': status_value,
            'status_display': status_display,
            'description': description,
            'assignee': assignee,
            'health': system.health,
            'utilization': system.utilization_percentage
        })
    
    return JsonResponse({
        'success': True,
        'view_date': view_date_str,
        'systems': systems_data
    })

@login_required
@require_GET
def system_details_api(request, stream=None, system_id=None):
    """API endpoint to get detailed information about a specific system"""    
    try:
        # Resolve stream to Stream object (404 if not found)
        from django.http import Http404
        try:
            stream_obj = get_stream_or_404(stream, default='HIC')
        except Http404:
            return JsonResponse({'success': False, 'error': f'Stream "{stream}" not found'}, status=404)

        system = System.objects.get(id=system_id, stream=stream_obj)
        
        # Get recent allocations
        recent_allocations = SystemAllocation.objects.filter(
            system_type=system.name
        ).order_by('-created_at')[:10]
        
        allocations_data = [{
            'user': allocation.user.username,
            'start_date': allocation.start_date.isoformat(),
            'end_date': allocation.end_date.isoformat(),
            'participant': allocation.blocked_for_participant.name if allocation.blocked_for_participant else None
        } for allocation in recent_allocations]
        
        # Get status history
        status_history = SystemStatusHistory.objects.filter(
            system=system
        ).order_by('-updated_at')[:10]
        
        history_data = [{
            'status': history.status,
            'description': history.description,
            'updated_by': history.updated_by,
            'updated_at': history.updated_at.isoformat()
        } for history in status_history]
        
        return JsonResponse({
            'success': True,
            'system': {
                'id': system.id,
                'name': system.name,
                'status': system.status,
                'description': system.description,
                'health': system.health,
                'utilization': system.utilization_percentage,
                'nc_details': system.nc_details,
                'last_updated': system.last_updated.isoformat(),
                'recent_allocations': allocations_data,
                'status_history': history_data
            }
        })
        
    except System.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'System not found'
        }, status=404)

@login_required
@require_GET
def system_metrics(request, system_id):
    """Get metrics for a specific system including downtime data"""
    system = get_object_or_404(System, id=system_id)
    try:
        metrics = SystemMetrics.objects.get(system=system)
    except SystemMetrics.DoesNotExist:
        metrics = SystemMetrics.objects.create(system=system)
    
    # Get downtime metrics (this calculates real availability)
    downtime_metrics = system.get_downtime_metrics(30)
    current_downtime = system.get_current_downtime()
    
    # Use calculated availability as uptime percentage instead of static value
    actual_uptime_percentage = downtime_metrics.availability_percentage if hasattr(downtime_metrics, 'availability_percentage') else metrics.uptime_percentage
    
    # Update the metrics with calculated uptime
    if hasattr(downtime_metrics, 'availability_percentage'):
        metrics.uptime_percentage = downtime_metrics.availability_percentage
        metrics.save()
    
    data = {
        'usage_hours': metrics.usage_hours,
        'total_allocations': metrics.total_allocations,
        'last_allocation_date': metrics.last_allocation_date.isoformat() if metrics.last_allocation_date else None,
        'average_session_duration': str(metrics.average_session_duration) if metrics.average_session_duration else None,
        'uptime_percentage': actual_uptime_percentage,  # Use calculated value
        'utilization_percentage': system.utilization_percentage,
        # Downtime metrics
        'downtime_metrics': {
            'availability_percentage': round(downtime_metrics.availability_percentage, 2),
            'total_downtime_hours': round(downtime_metrics.total_downtime_hours, 2),
            'total_incidents': downtime_metrics.total_incidents,
            'planned_downtime_hours': round(downtime_metrics.planned_downtime_hours, 2),
            'unplanned_downtime_hours': round(downtime_metrics.unplanned_downtime_hours, 2),
            'mean_time_to_repair_hours': round(downtime_metrics.mean_time_to_repair_hours, 2) if downtime_metrics.mean_time_to_repair_hours else None,
            'mean_time_between_failures_hours': round(downtime_metrics.mean_time_between_failures_hours, 2) if downtime_metrics.mean_time_between_failures_hours else None,
        },
        'current_downtime': {
            'is_down': system.is_currently_down(),
            'downtime_id': current_downtime.id if current_downtime else None,
            'title': current_downtime.title if current_downtime else None,
            'start_time': current_downtime.start_time.isoformat() if current_downtime else None,
            'duration_hours': round(current_downtime.duration_hours, 2) if current_downtime else 0,
            'impact_level': current_downtime.get_impact_level_display() if current_downtime else None
        }
    }
    return JsonResponse(data)

@login_required
@csrf_exempt
@require_POST
def update_system_metrics(request, system_id):
    """Update metrics for a specific system"""
    system = get_object_or_404(System, id=system_id)
    try:
        metrics = SystemMetrics.objects.get(system=system)
    except SystemMetrics.DoesNotExist:
        metrics = SystemMetrics.objects.create(system=system)
    
    data = json.loads(request.body)
    
    # Update the metrics
    if 'usage_hours' in data:
        metrics.usage_hours = data['usage_hours']
    if 'total_allocations' in data:
        metrics.total_allocations = data['total_allocations']
    if 'uptime_percentage' in data:
        metrics.uptime_percentage = data['uptime_percentage']
    
    metrics.save()
    
    # Update system utilization
    if 'utilization_percentage' in data:
        system.utilization_percentage = data['utilization_percentage']
        system.save()
    
    return JsonResponse({'status': 'success'})

@login_required
def export_systems_log(request, stream=None):
    """
    Export a detailed log of all systems including booking, utilization, status, and all possible details with timestamps.
    """
    if stream:
        # Handle empty stream parameter
        if not stream or stream.strip() == '':
            stream = 'HIC'
        
        # Get the Stream object for database queries (404 if not found). Default to 'HIC' when missing.
        stream_obj = get_stream_or_404(stream, default='HIC')
        systems = System.objects.filter(stream=stream_obj)
    else:
        systems = System.objects.all()
    response = HttpResponse(content_type='text/csv')
    filename = f"systems_detailed_log_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow([
        'System Name', 'Status', 'Utilization (%)', 'Health','Description',
        'Booking User', 'Booking Start', 'Booking End',
        'Last Status Change', 'Last Utilization Update', 'History Details'
    ])
    for system in systems:
        allocation = SystemAllocation.objects.filter(system_type=system.name, stream=system.stream).order_by('-end_date').first()
        booking_user = allocation.user.username if allocation and allocation.user else ''
        booking_start = allocation.start_date.strftime('%Y-%m-%d %H:%M:%S') if allocation and allocation.start_date else ''
        booking_end = allocation.end_date.strftime('%Y-%m-%d %H:%M:%S') if allocation and allocation.end_date else ''
        booking_type = getattr(allocation, 'blocked_for_participant', None)
        booking_type = booking_type.name if booking_type else ''
        last_status = SystemStatusHistory.objects.filter(system=system).order_by('-updated_at').first()
        last_status_change = last_status.updated_at.strftime('%Y-%m-%d %H:%M:%S') if last_status and last_status.updated_at else ''
        last_util_update = last_status.updated_at.strftime('%Y-%m-%d %H:%M:%S') if last_status and last_status.updated_at else ''
        history_details = []
        for h in SystemStatusHistory.objects.filter(system=system).order_by('-updated_at'):
            history_details.append(f"[{h.updated_at.strftime('%Y-%m-%d %H:%M:%S')}] {h.status}")
        row = [
            system.name or '',
            system.get_status_display() or '',
            getattr(system, 'utilization_percentage', '') or '',
            getattr(system, 'health', '') or '',
            getattr(system, 'description', ''),
            booking_user,
            booking_start,
            booking_end,
            last_status_change,
            last_util_update,
            '\n'.join(history_details)
        ]
        # Ensure row has exactly 10 columns
        while len(row) < 10:
            row.append('')
        writer.writerow(row)
    return response

from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import Note

@login_required
def notes_list(request):
    """List all notes accessible to the current user."""
    # Get public notes and user's private notes
    notes = Note.objects.filter(
        Q(is_public=True) | Q(created_by=request.user)
    ).select_related('created_by').prefetch_related('attachments').order_by('-created_at')
    
    # Add search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        notes = notes.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(created_by__username__icontains=search_query) |
            Q(created_by__first_name__icontains=search_query) |
            Q(created_by__last_name__icontains=search_query)
        )
    
    # Get shared notes count for the current user
    shared_notes_count = SharedNote.objects.filter(shared_with=request.user).count()
    
    return render(request, 'products/notes_list.html', {
        'notes': notes,
        'shared_notes_count': shared_notes_count,
        'selected_stream': 'HIC'  # Default stream
    })

@login_required
def note_detail(request, pk):
    """View a single note."""
    note = get_object_or_404(Note, pk=pk)
    # Check if user has permission to view the note
    if not note.is_public and note.created_by != request.user:
        messages.error(request, "You don't have permission to view this note.")
        return redirect('notes_list')
    
    # Get all attachments for this note
    attachments = note.attachments.all()
    
    return render(request, 'products/note_detail.html', {
        'note': note,
        'attachments': attachments,
        'selected_stream': 'HIC'
    })

@login_required
def note_create(request):
    """Create a new note."""
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        is_public = request.POST.get('is_public') == 'on'
        
        if not title or not content:
            messages.error(request, 'Title and content are required.')
            return render(request, 'products/note_form.html', {
                'note': {'title': title, 'content': content, 'is_public': is_public},
                'selected_stream': 'HIC'
            })
        
        note = Note.objects.create(
            title=title,
            content=content,
            is_public=is_public,
            created_by=request.user,
            updated_by=request.user
        )
        
        # Handle file attachments
        files = request.FILES.getlist('attachments')
        for file in files:
            # Validate file size (5MB limit)
            if file.size > 5 * 1024 * 1024:
                messages.warning(request, f'File "{file.name}" is larger than 5MB and was not uploaded.')
                continue
            
            # Validate file type
            allowed_types = [
                'image/jpeg', 'image/png', 'image/gif', 'image/bmp', 'image/webp',
                'application/pdf',
                'application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ]
            
            if file.content_type not in allowed_types:
                messages.warning(request, f'File "{file.name}" has an unsupported format and was not uploaded.')
                continue
            
            NoteAttachment.objects.create(
                note=note,
                file=file,
                original_filename=file.name,
                content_type=file.content_type,
                uploaded_by=request.user
            )
        
        messages.success(request, 'Note created successfully.')
        return redirect('note_detail', pk=note.pk)
    
    return render(request, 'products/note_form.html', {
        'note': None,
        'selected_stream': 'HIC'
    })

@login_required
def note_edit(request, pk):
    """Edit an existing note."""
    note = get_object_or_404(Note, pk=pk)
    
    # Check if user has permission to edit the note
    if note.created_by != request.user:
        messages.error(request, "You don't have permission to edit this note.")
        return redirect('notes_list')
    
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        is_public = request.POST.get('is_public') == 'on'
        
        if not title or not content:
            messages.error(request, 'Title and content are required.')
            return render(request, 'products/note_form.html', {
                'note': note,
                'selected_stream': 'HIC'
            })
        
        note.title = title
        note.content = content
        note.is_public = is_public
        note.updated_by = request.user
        note.updated_at = timezone.now()
        note.save()
        
        # Handle new file attachments
        files = request.FILES.getlist('attachments')
        for file in files:
            # Validate file size (5MB limit)
            if file.size > 5 * 1024 * 1024:
                messages.warning(request, f'File "{file.name}" is larger than 5MB and was not uploaded.')
                continue
            
            # Validate file type
            allowed_types = [
                'image/jpeg', 'image/png', 'image/gif', 'image/bmp', 'image/webp',
                'application/pdf',
                'application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ]
            
            if file.content_type not in allowed_types:
                messages.warning(request, f'File "{file.name}" has an unsupported format and was not uploaded.')
                continue
            
            NoteAttachment.objects.create(
                note=note,
                file=file,
                original_filename=file.name,
                content_type=file.content_type,
                uploaded_by=request.user
            )        # Handle attachment deletions
        attachment_ids_to_delete = request.POST.getlist('delete_attachments')
        if attachment_ids_to_delete:
            attachments_to_delete = NoteAttachment.objects.filter(
                id__in=attachment_ids_to_delete,
                note=note
            )
            for attachment in attachments_to_delete:
                # Delete each attachment individually to ensure file deletion
                attachment.delete()
        
        messages.success(request, 'Note updated successfully.')
        return redirect('note_detail', pk=note.pk)
    
    return render(request, 'products/note_form.html', {
        'note': note,
        'selected_stream': 'HIC'
    })

@login_required
def note_delete(request, pk):
    """Delete a note."""
    note = get_object_or_404(Note, pk=pk)
    
    # Check if user has permission to delete the note
    if note.created_by != request.user:
        messages.error(request, "You don't have permission to delete this note.")
        return redirect('notes_list')    
    if request.method == 'POST':
        # Delete the note (this will also delete attachments and files due to CASCADE and overridden delete method)
        note.delete()
        messages.success(request, 'Note and all attachments deleted successfully.')
        return redirect('notes_list')
    
    return render(request, 'products/note_confirm_delete.html', {
        'note': note,
        'selected_stream': 'HIC'
    })

@login_required
def share_note(request, pk):
    """Share a note with other users."""
    note = get_object_or_404(Note, pk=pk)
    
    # Check if user has permission to share the note (must be public or owner)
    if not note.is_public and note.created_by != request.user:
        messages.error(request, "You don't have permission to share this note.")
        return redirect('notes_list')
    
    if request.method == 'POST':
        user_ids = request.POST.getlist('users')
        message = request.POST.get('message', '')
        
        if not user_ids:
            messages.error(request, 'Please select at least one user to share with.')
            users = User.objects.exclude(id=request.user.id).order_by('username')
            return render(request, 'products/share_note.html', {
                'note': note,
                'users': users,
                'selected_stream': 'HIC'
            })
        
        shared_count = 0
        for user_id in user_ids:
            try:
                user = User.objects.get(id=user_id)
                shared_note, created = SharedNote.objects.get_or_create(
                    note=note,
                    shared_by=request.user,
                    shared_with=user,
                    defaults={'message': message}
                )
                if created:
                    shared_count += 1
            except User.DoesNotExist:
                continue
        
        if shared_count > 0:
            messages.success(request, f'Note shared with {shared_count} user(s).')
        else:
            messages.info(request, 'Note was already shared with selected users.')
        
        return redirect('note_detail', pk=pk)
    
    # GET request - show share form
    users = User.objects.exclude(id=request.user.id).order_by('username')
    return render(request, 'products/share_note.html', {
        'note': note,
        'users': users,
        'selected_stream': 'HIC'
    })

@login_required
def shared_notes(request):
    """View notes shared with the current user."""
    shared_notes = SharedNote.objects.filter(shared_with=request.user).select_related(
        'note', 'shared_by', 'note__created_by'
    ).order_by('-shared_at')
    
    # Mark as read when viewed
    SharedNote.objects.filter(shared_with=request.user, is_read=False).update(is_read=True)
    
    return render(request, 'products/shared_notes.html', {
        'shared_notes': shared_notes,
        'selected_stream': 'HIC'
    })

@login_required
def remove_shared_note(request, pk):
    """Remove a shared note from user's shared list."""
    shared_note = get_object_or_404(SharedNote, pk=pk, shared_with=request.user)
    
    if request.method == 'POST':
        shared_note.delete()
        messages.success(request, 'Shared note removed from your list.')
        return redirect('shared_notes')
    
    return render(request, 'products/remove_shared_note.html', {
        'shared_note': shared_note,
        'selected_stream': 'HIC'
    })

@login_required
@require_POST  
def assign_role(request, user_id):
    """Assign a role to a user"""
    if not is_super_admin(request.user):
        messages.error(request, 'Only Super Admins can assign roles.')
        return redirect('user_list')
    
    user = get_object_or_404(User, id=user_id)
    role = request.POST.get('role')
    
    if role not in ['user', 'lab_incharge', 'admin', 'super_admin']:
        messages.error(request, 'Invalid role selected.')
        return redirect('user_list')
        
    custom_user, created = CustomUser.objects.get_or_create(user=user)
    
    # Check if user already has this role
    if custom_user.user_roles.filter(role=role).exists():
        messages.info(request, f'User {user.username} already has the {role} role.')
        return redirect('user_list')
    
    # Create the role assignment
    UserRole.objects.create(
        custom_user=custom_user,
        role=role,
        assigned_by=request.user
    )
    
    messages.success(request, f'Role {role} assigned to {user.username}.')
    return redirect('user_list')

@login_required
@require_POST
def remove_role(request, user_id):
    """Remove a role from a user"""
    if not is_super_admin(request.user):
        messages.error(request, 'Only Super Admins can remove roles.')
        return redirect('user_list')
    
    user = get_object_or_404(User, id=user_id)
    role = request.POST.get('role')
    
    try:
        custom_user = user.custom_profile
        role_obj = custom_user.user_roles.get(role=role)
        role_obj.delete()
        
        # Check if user has any roles left
        remaining_roles = custom_user.user_roles.count()
        if remaining_roles == 0:
            # Remove all stream access when user has no roles
            custom_user.stream_access.all().delete()
            messages.success(request, f'Role {role} removed from {user.username}. All stream access has been revoked as user has no remaining roles.')
        else:
            messages.success(request, f'Role {role} removed from {user.username}.')
    
    except (CustomUser.DoesNotExist, UserRole.DoesNotExist):
        messages.error(request, 'Role assignment not found.')
    
    return redirect('user_list')

@login_required
@require_POST
def grant_stream_access(request, user_id):
    """Grant stream access to a user"""
    if not is_super_admin(request.user):
        messages.error(request, 'Only Super Admins can grant stream access.')
        return redirect('user_list')
    
    user = get_object_or_404(User, id=user_id)
    stream_id = request.POST.get('stream_id')
    
    try:
        stream = Stream.objects.get(id=stream_id)        
        custom_user, created = CustomUser.objects.get_or_create(user=user)
        
        UserStreamAccess.objects.get_or_create(
            custom_user=custom_user,
            stream=stream,
            defaults={'granted_by': request.user}
        )
        
        if created:
            messages.success(request, f'Stream access granted: {user.username} can now access {stream.name}.')
        else:
            messages.info(request, f'User {user.username} already has access to {stream.name}.')
            
    except Stream.DoesNotExist:
        messages.error(request, 'Invalid stream selected.')
    
    return redirect('user_list')

@login_required
@require_POST
def revoke_stream_access(request, user_id):
    """Revoke stream access from a user"""
    if not is_super_admin(request.user):
        messages.error(request, 'Only Super Admins can revoke stream access.')
        return redirect('user_list')
    
    user = get_object_or_404(User, id=user_id)
    stream_id = request.POST.get('stream_id')
    
    try:
        stream = Stream.objects.get(id=stream_id)
        custom_user = user.custom_profile
        access = custom_user.stream_access.get(stream=stream)
        access.delete()
        messages.success(request, f'Stream access revoked: {user.username} can no longer access {stream.name}.')    
    
    except (Stream.DoesNotExist, CustomUser.DoesNotExist, UserStreamAccess.DoesNotExist):
        messages.error(request, 'Stream access not found.')
    
    return redirect('user_list')

@login_required
def manage_streams(request):
    """Manage dynamic streams"""
    if not is_super_admin(request.user):
        messages.error(request, 'Only Super Admins can manage streams.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            name = request.POST.get('name')
            description = request.POST.get('description', '')
            allow_public_registration = request.POST.get('allow_public_registration') == 'on'
            requires_approval = request.POST.get('requires_approval') == 'on'
            
            if Stream.objects.filter(name=name).exists():
                messages.error(request, f'Stream "{name}" already exists.')
            else:
                Stream.objects.create(
                    name=name,
                    description=description,
                    allow_public_registration=allow_public_registration,
                    requires_approval=requires_approval,
                    created_by=request.user
                )
                messages.success(request, f'Stream "{name}" created successfully.')
        
        elif action == 'update':
            stream_id = request.POST.get('stream_id')
            try:
                stream = Stream.objects.get(id=stream_id)
                stream.name = request.POST.get('name')
                stream.description = request.POST.get('description', '')
                stream.allow_public_registration = request.POST.get('allow_public_registration') == 'on'
                stream.requires_approval = request.POST.get('requires_approval') == 'on'
                stream.is_active = request.POST.get('is_active') == 'on'
                stream.save()
                messages.success(request, f'Stream "{stream.name}" updated successfully.')
            except Stream.DoesNotExist:
                messages.error(request, 'Stream not found.')
        
        elif action == 'delete':
            stream_id = request.POST.get('stream_id')
            try:
                stream = Stream.objects.get(id=stream_id)
                stream_name = stream.name
                
                # Record deletion in history
                from .models import StreamDeletionHistory
                StreamDeletionHistory.objects.create(
                    stream_name=stream_name,
                    deleted_by=request.user
                )
                
                stream.delete()
                messages.success(request, f'Stream "{stream_name}" deleted successfully.')
            except Stream.DoesNotExist:
                messages.error(request, 'Stream not found.')
    
    streams = Stream.objects.all().order_by('name')
    return render(request, 'products/manage_streams.html', {
        'streams': streams,
        'selected_stream': 'management'
    })

# ========== System Tags and Allocation Tree Views ==========

@login_required
def allocation_tree(request, stream=None):
    """Display tree structure of systems with their tagged products and components"""
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Get all systems for this stream with their allocations
    from .models import SystemTag, Project
    systems = System.objects.filter(stream=stream_obj).prefetch_related(
        'tags__products__category',
        'tags__sublevels',
        'tags__sublevel_tools',
        'tags__projects'
    ).order_by('name')
    
    # Get recent allocations for each system
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    today = timezone.now()
    fifteen_days_ago = today - timedelta(days=15)
    
    systems_data = []
    for system in systems:
        # Get tags for this system
        tags = system.tags.all()
        
        # Get recent allocations
        recent_allocations = SystemAllocation.objects.filter(
            system_type=system.name,
            stream=stream_obj,
            end_date__gte=fifteen_days_ago
        ).select_related('user', 'blocked_for_participant').order_by('-start_date')[:5]
        
        systems_data.append({
            'system': system,
            'tags': tags,
            'recent_allocations': recent_allocations
        })
    
    context = {
        'systems_data': systems_data,
        'stream': stream,
        'selected_stream': stream,
    }
    
    return render(request, 'products/allocation_tree.html', context)

@login_required
@require_POST
def create_system_tag(request, stream=None):
    """Create or update a system tag (only one tag allowed per system)"""
    if not can_edit_products(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    system_id = request.POST.get('system_id')
    tag_id = request.POST.get('tag_id', '').strip()  # For updates
    tag_name = request.POST.get('tag_name', '').strip()
    description = request.POST.get('description', '').strip()
    
    if not system_id or not tag_name:
        return JsonResponse({'success': False, 'error': 'System and tag name are required'}, status=400)
    
    try:
        from .models import SystemTag
        system = System.objects.get(id=system_id, stream=stream_obj)
        
        # Check if we're updating an existing tag
        if tag_id:
            tag = SystemTag.objects.get(id=tag_id, system=system, stream=stream_obj)
            tag.tag_name = tag_name
            tag.description = description
            tag.save()
            message = f'Tag "{tag_name}" updated successfully'
        else:
            # Check if system already has a tag (only one allowed)
            existing_tag = SystemTag.objects.filter(system=system).first()
            if existing_tag:
                return JsonResponse({
                    'success': False, 
                    'error': 'This system already has a tag. Please edit the existing tag instead.'
                }, status=400)
            
            tag = SystemTag.objects.create(
                system=system,
                tag_name=tag_name,
                stream=stream_obj,
                description=description,
                created_by=request.user
            )
            message = f'Tag "{tag_name}" created successfully'
        
        return JsonResponse({
            'success': True,
            'tag_id': tag.id,
            'message': message
        })
    except System.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'System not found'}, status=404)
    except SystemTag.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tag not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def delete_system_tag(request, stream=None, tag_id=None):
    """Delete a system tag"""
    if not can_delete_products(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    try:
        from .models import SystemTag
        tag = SystemTag.objects.get(id=tag_id, stream=stream_obj)
        tag_name = tag.tag_name
        tag.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Tag "{tag_name}" deleted successfully'
        })
    except SystemTag.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tag not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def manage_tag_items(request, stream=None, tag_id=None):
    """Add or remove items (products, sublevels, sublevel_tools) from a tag"""
    if not can_edit_products(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    try:
        from .models import SystemTag, Project
        tag = SystemTag.objects.get(id=tag_id, stream=stream_obj)
        
        action = request.POST.get('action')  # 'add' or 'remove'
        item_type = request.POST.get('item_type')  # 'product', 'sublevel', 'sublevel_tool', or 'project'
        item_id = request.POST.get('item_id')
        
        if action not in ['add', 'remove'] or item_type not in ['product', 'sublevel', 'sublevel_tool', 'project']:
            return JsonResponse({'success': False, 'error': 'Invalid action or item type'}, status=400)
        
        if not item_id:
            return JsonResponse({'success': False, 'error': 'Item ID is required'}, status=400)
        
        # Get the appropriate model and relation
        if item_type == 'product':
            item = Product.objects.get(id=item_id, stream=stream_obj)
            relation = tag.products
        elif item_type == 'sublevel':
            item = SubLevel.objects.get(id=item_id)
            relation = tag.sublevels
        elif item_type == 'sublevel_tool':
            item = SubLevelTool.objects.get(id=item_id)
            relation = tag.sublevel_tools
        else:  # project
            # Get project regardless of stream since we now show all projects
            item = Project.objects.get(id=item_id)
            relation = tag.projects
        
        # Perform action
        if action == 'add':
            relation.add(item)
            message = f'{item_type.replace("_", " ").title()} added to tag'
        else:
            relation.remove(item)
            message = f'{item_type.replace("_", " ").title()} removed from tag'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'item_count': tag.get_all_components_count()
        })
        
    except (SystemTag.DoesNotExist, Product.DoesNotExist, SubLevel.DoesNotExist, SubLevelTool.DoesNotExist, Project.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Tag or item not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def get_available_items(request, stream=None):
    """Get available products, sublevels, and sublevel_tools for tagging"""
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    tag_id = request.GET.get('tag_id')
    item_type = request.GET.get('item_type', 'product')
    
    try:
        from .models import SystemTag
        
        # Get already tagged items if tag_id is provided
        tagged_ids = []
        if tag_id:
            tag = SystemTag.objects.get(id=tag_id, stream=stream_obj)
            if item_type == 'product':
                tagged_ids = list(tag.products.values_list('id', flat=True))
            elif item_type == 'sublevel':
                tagged_ids = list(tag.sublevels.values_list('id', flat=True))
            elif item_type == 'sublevel_tool':
                tagged_ids = list(tag.sublevel_tools.values_list('id', flat=True))
            elif item_type == 'project':
                tagged_ids = list(tag.projects.values_list('id', flat=True))
        
        # Get available items
        if item_type == 'product':
            items = Product.objects.filter(stream=stream_obj).select_related('category').values(
                'id', 'name', 'serial_number', 'category__name'
            )
            items_list = [{
                'id': item['id'],
                'name': item['name'],
                'serial_number': item['serial_number'],
                'category': item['category__name'],
                'is_tagged': item['id'] in tagged_ids
            } for item in items]
        elif item_type == 'sublevel':
            items = SubLevel.objects.all().values('id', 'name', 'stream', 'in_stock', 'in_use')
            items_list = [{
                'id': item['id'],
                'name': item['name'],
                'stream': item['stream'],
                'in_stock': item['in_stock'],
                'in_use': item['in_use'],
                'is_tagged': item['id'] in tagged_ids
            } for item in items]
        elif item_type == 'sublevel_tool':
            items = SubLevelTool.objects.all().values('id', 'name', 'stream', 'in_stock', 'in_use')
            items_list = [{
                'id': item['id'],
                'name': item['name'],
                'stream': item['stream'],
                'in_stock': item['in_stock'],
                'in_use': item['in_use'],
                'is_tagged': item['id'] in tagged_ids
            } for item in items]
        else:  # project
            from .models import Project
            items = Project.objects.all().values(
                'id', 'name', 'status', 'priority', 'progress_percentage'
            )
            items_list = [{
                'id': item['id'],
                'name': item['name'],
                'status': item['status'],
                'priority': item['priority'],
                'progress_percentage': item['progress_percentage'],
                'is_tagged': item['id'] in tagged_ids
            } for item in items]
        
        return JsonResponse({
            'success': True,
            'items': items_list
        })
        
    except SystemTag.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tag not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# Project Status Management Views
from django.views.decorators.cache import never_cache

@login_required
@never_cache
def project_status(request):
    """View to manage projects - list, create, edit, delete"""
    # Get user's accessible streams - use first one or default
    if hasattr(request.user, 'custom_profile'):
        accessible_streams = request.user.custom_profile.get_accessible_streams()
        if accessible_streams.exists():
            stream_obj = accessible_streams.first()
        else:
            # Fallback to any active stream
            stream_obj = Stream.objects.filter(is_active=True).first()
    else:
        stream_obj = Stream.objects.filter(is_active=True).first()
    
    if not stream_obj:
        messages.error(request, 'No active streams available.')
        return redirect('dashboard')
    
    # Import here to avoid circular imports
    from .models import Project
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            # Create new project
            try:
                # Get selected stream from form or use default
                selected_stream_id = request.POST.get('stream')
                if selected_stream_id:
                    try:
                        project_stream = Stream.objects.get(id=selected_stream_id, is_active=True)
                    except Stream.DoesNotExist:
                        project_stream = stream_obj  # fallback to default
                else:
                    project_stream = stream_obj  # use default if not specified
                
                project = Project.objects.create(
                    name=request.POST.get('name'),
                    description=request.POST.get('description', ''),
                    duration=request.POST.get('duration', ''),
                    start_date=request.POST.get('start_date'),
                    initial_release_date=request.POST.get('initial_release_date'),
                    final_release_date=request.POST.get('final_release_date'),
                    status=request.POST.get('status', 'running'),
                    priority=request.POST.get('priority', 'medium'),
                    progress_percentage=int(request.POST.get('progress_percentage', 0)),
                    stream=project_stream,
                    created_by=request.user
                )
                
                # Calculate expected progress if running, otherwise set to 0
                if project.status == 'running':
                    project.expected_progress = project.calculate_expected_progress()
                else:
                    project.expected_progress = 0
                project.save()
                
                # Handle team members if provided
                team_member_ids = request.POST.getlist('team_members')
                if team_member_ids:
                    project.team_members.set(team_member_ids)
                
            except Exception as e:
                messages.error(request, f'Error creating project: {str(e)}')
        
        elif action == 'update':
            # Update existing project
            try:
                project_id = request.POST.get('project_id')
                project = Project.objects.get(id=project_id, stream=stream_obj)
                from datetime import datetime
                project.name = request.POST.get('name')
                project.description = request.POST.get('description', '')
                project.duration = request.POST.get('duration', '')
                # Parse and assign start_date
                start_date_str = request.POST.get('start_date')
                if start_date_str:
                    try:
                        project.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                    except Exception as e:
                        print(f'ERROR: Could not parse start_date: {start_date_str} ({e})')
                        project.start_date = None
                else:
                    project.start_date = None
                # Parse and assign initial_release_date
                initial_release_date_str = request.POST.get('initial_release_date')
                if initial_release_date_str:
                    try:
                        project.initial_release_date = datetime.strptime(initial_release_date_str, '%Y-%m-%d').date()
                    except Exception as e:
                        print(f'ERROR: Could not parse initial_release_date: {initial_release_date_str} ({e})')
                        project.initial_release_date = None
                else:
                    project.initial_release_date = None
                # Parse and assign final_release_date
                final_release_date_str = request.POST.get('final_release_date')
                if final_release_date_str:
                    try:
                        project.final_release_date = datetime.strptime(final_release_date_str, '%Y-%m-%d').date()
                    except Exception as e:
                        print(f'ERROR: Could not parse final_release_date: {final_release_date_str} ({e})')
                        project.final_release_date = None
                else:
                    project.final_release_date = None
                project.status = request.POST.get('status', 'running')
                project.priority = request.POST.get('priority', 'medium')
                project.progress_percentage = int(request.POST.get('progress_percentage', 0))
                # Calculate expected progress if running, otherwise set to 0
                if project.status == 'running':
                    project.expected_progress = project.calculate_expected_progress()
                else:
                    project.expected_progress = 0
                project.save()
                # Handle team members
                team_member_ids = request.POST.getlist('team_members')
                if team_member_ids:
                    project.team_members.set(team_member_ids)
                else:
                    project.team_members.clear()
                messages.success(request, f'Project "{project.name}" updated successfully!')
            except Project.DoesNotExist:
                messages.error(request, 'Project not found.')
            except Exception as e:
                messages.error(request, f'Error updating project: {str(e)}')
        
        return redirect('project_status')
    
    # GET request - display project list
    # Show all projects regardless of stream when accessed via /projects/ URL
    projects = Project.objects.all().select_related('created_by', 'stream').prefetch_related('team_members')
    
    # Get all users for team member selection
    users = User.objects.filter(is_active=True).order_by('username')
    
    # Get all active streams for project assignment
    all_streams = Stream.objects.filter(is_active=True).order_by('name')

    context = {
        'selected_stream': stream_obj.name,
        'stream': stream_obj,
        'projects': projects,
        'users': users,
        'all_streams': all_streams,
    }
    
    return render(request, 'products/project_status.html', context)

@login_required
@require_POST
def delete_project(request):
    """Delete a project with password confirmation"""
    # Get user's accessible streams
    if hasattr(request.user, 'custom_profile'):
        accessible_streams = request.user.custom_profile.get_accessible_streams()
        if accessible_streams.exists():
            stream_obj = accessible_streams.first()
        else:
            stream_obj = Stream.objects.filter(is_active=True).first()
    else:
        stream_obj = Stream.objects.filter(is_active=True).first()
    
    if not stream_obj:
        return JsonResponse({'success': False, 'error': 'No active streams available'}, status=400)
    
    from .models import Project
    
    try:
        data = json.loads(request.body)
        project_id = data.get('project_id')
        password = data.get('password')
        
        # Verify password
        if not check_password(password, request.user.password):
            return JsonResponse({'success': False, 'error': 'Invalid password'}, status=400)
        
        # Delete project
        project = Project.objects.get(id=project_id, stream=stream_obj)
        project_name = project.name
        project.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Project "{project_name}" deleted successfully'
        })
        
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# =============================================
# HOLISTIC DASHBOARD VIEWS
# =============================================

@login_required
def holistic_dashboard(request, stream=None):
    """Main view for the advanced holistic dashboard"""
    from .models import HolisticSystem, HolisticWeeklyData
    
    # Handle stream
    stream = stream or request.GET.get('stream', 'HIC')
    if not stream or stream.strip() == '':
        stream = 'HIC'
    
    # Check user access
    has_access, error_message, custom_profile = check_user_access(request, stream)
    if not has_access:
        logout(request)
        messages.error(request, error_message)
        return redirect('please_login')
    
    # Get stream object
    stream_obj = get_stream_or_404(stream)
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '').strip()
    week_filter = request.GET.get('week', '')
    
    # Base queryset
    systems = HolisticSystem.objects.filter(stream=stream_obj).select_related('created_by', 'updated_by')
    
    # Apply filters
    if status_filter:
        systems = systems.filter(system_availability=status_filter)
    
    if search_query:
        systems = systems.filter(
            Q(sr_no__icontains=search_query) |
            Q(system_owner__icontains=search_query) |
            Q(stmi_number__icontains=search_query) |
            Q(test_engineer__icontains=search_query) |
            Q(location_info__icontains=search_query)
        )
    
    # Get current week
    from datetime import date
    current_week = date.today().isocalendar()[1]
    current_year = date.today().year
    
    # Get all unique weeks for dropdown
    all_weeks = HolisticWeeklyData.objects.values('week_number', 'year').distinct().order_by('-year', '-week_number')
    
    # Prepare week data and project IDs for each system
    import json as _json
    for system in systems:
        if week_filter:
            try:
                # Expecting format W47-2025
                if '-' in week_filter:
                    week_part, year_part = week_filter.split('-')
                    week_num = int(week_part.replace('W', ''))
                    year_num = int(year_part)
                else:
                    week_num = int(week_filter.replace('W', ''))
                    year_num = current_year
                system.current_week_data = system.weekly_data.filter(
                    week_number=week_num,
                    year=year_num
                ).first()
            except Exception:
                system.current_week_data = system.get_current_week_data()
        else:
            system.current_week_data = system.get_current_week_data()

        # If the week is in the past, set utilization_percentage to 100 for display
        if system.current_week_data:
            week_is_past = (
                (system.current_week_data.year < current_year) or
                (system.current_week_data.year == current_year and system.current_week_data.week_number < current_week)
            )
            if week_is_past:
                system.current_week_data.utilization_percentage = 100

        # Precompute current project ID for template (if any)
        current_project_id = None
        if system.current_week_data and system.current_week_data.project:
            current_project_id = system.current_week_data.project.id
        system.current_project_id_json = _json.dumps(current_project_id)
        
        # Get project timeline for this system
        system.project_timeline = system.get_project_timeline()
    
    # Calculate statistics
    total_systems = systems.count()
    available_count = systems.filter(system_availability='available').count()
    allocated_count = systems.filter(system_availability='allocated').count()
    maintenance_count = systems.filter(system_availability='maintenance').count()
    offline_count = systems.filter(system_availability='offline').count()
    reserved_count = systems.filter(system_availability='reserved').count()
    
    # Get recent weeks for display (last 8 weeks)
    recent_weeks = []
    for i in range(8):
        week_date = date.today() - timedelta(weeks=i)
        week_num = week_date.isocalendar()[1]
        week_year = week_date.year
        recent_weeks.append({
            'week': week_num,
            'year': week_year,
            'label': f'W{week_num}',
            'value': f'W{week_num}-{week_year}'
        })
    
    from .models import Project
    projects = Project.objects.all().order_by('name')
    context = {
        'systems': systems,
        'stream': stream,
        'selected_stream': stream,
        'status_filter': status_filter,
        'search_query': search_query,
        'week_filter': week_filter,
        'all_weeks': all_weeks,
        'recent_weeks': recent_weeks,
        'current_week': current_week,
        'current_year': current_year,
        'total_systems': total_systems,
        'available_count': available_count,
        'allocated_count': allocated_count,
        'maintenance_count': maintenance_count,
        'offline_count': offline_count,
        'reserved_count': reserved_count,
        'status_choices': HolisticSystem.STATUS_CHOICES,
        'projects': projects,
    }
    
    return render(request, 'products/holistic_dashboard.html', context)

@login_required
def holistic_system_create(request, stream=None):
    """Create a new holistic system"""
    from .models import HolisticSystem, HolisticSystemHistory
    
    stream = stream or 'HIC'
    stream_obj = get_stream_or_404(stream)
    
    if request.method == 'POST':
        try:
            # Create system
            system = HolisticSystem.objects.create(
                sr_no=request.POST.get('sr_no'),
                system_availability=request.POST.get('system_availability', 'available'),
                allocation_to_sl_no=request.POST.get('allocation_to_sl_no'),
                location_info=request.POST.get('location_info'),
                stmi_number=request.POST.get('stmi_number'),
                system_owner=request.POST.get('system_owner'),
                ecr_number=request.POST.get('ecr_number'),
                test_engineer=request.POST.get('test_engineer'),
                description=request.POST.get('description'),
                notes=request.POST.get('notes'),
                priority=request.POST.get('priority', 'medium'),
                stream=stream_obj,
                created_by=request.user,
                updated_by=request.user
            )
            
            # Create history entry
            HolisticSystemHistory.objects.create(
                holistic_system=system,
                action='created',
                user=request.user,
                details=f'System {system.sr_no} created'
            )
            
            return redirect('holistic_dashboard')
            
        except IntegrityError:
            messages.error(request, 'A system with this Serial Number already exists.')
        except Exception as e:
            messages.error(request, f'Error creating system: {str(e)}')
    
    context = {
        'stream': stream,
        'selected_stream': stream,
        'status_choices': HolisticSystem.STATUS_CHOICES,
        'edit': False,
    }
    
    return render(request, 'products/holistic_system_form.html', context)

@login_required
def holistic_system_edit(request, pk, stream=None):
    """Edit an existing holistic system"""
    from .models import HolisticSystem, HolisticSystemHistory
    
    stream = stream or 'HIC'
    stream_obj = get_stream_or_404(stream)
    
    system = get_object_or_404(HolisticSystem, pk=pk, stream=stream_obj)
    
    if request.method == 'POST':
        try:
            # Track changes
            changes = []
            old_values = {
                'sr_no': system.sr_no,
                'system_availability': system.system_availability,
                'allocation_to_sl_no': system.allocation_to_sl_no,
                'location_info': system.location_info,
                'stmi_number': system.stmi_number,
                'system_owner': system.system_owner,
                'ecr_number': system.ecr_number,
                'test_engineer': system.test_engineer,
                'description': system.description,
                'notes': system.notes,
                'priority': system.priority,
            }
            
            # Update fields
            system.sr_no = request.POST.get('sr_no')
            system.system_availability = request.POST.get('system_availability')
            system.allocation_to_sl_no = request.POST.get('allocation_to_sl_no')
            system.location_info = request.POST.get('location_info')
            system.stmi_number = request.POST.get('stmi_number')
            system.system_owner = request.POST.get('system_owner')
            system.ecr_number = request.POST.get('ecr_number')
            system.test_engineer = request.POST.get('test_engineer')
            system.description = request.POST.get('description')
            system.notes = request.POST.get('notes')
            system.priority = request.POST.get('priority', 'medium')
            system.updated_by = request.user
            
            # Track changes
            for field, old_value in old_values.items():
                new_value = getattr(system, field)
                if old_value != new_value:
                    changes.append(f"{field}: '{old_value}' → '{new_value}'")
            
            system.save()
            
            # Create history entry
            if changes:
                HolisticSystemHistory.objects.create(
                    holistic_system=system,
                    action='edited',
                    user=request.user,
                    details='; '.join(changes)
                )
            
            messages.success(request, f'System {system.sr_no} updated successfully!')
            return redirect('holistic_dashboard')
            
        except IntegrityError:
            messages.error(request, 'A system with this Serial Number already exists.')
        except Exception as e:
            messages.error(request, f'Error updating system: {str(e)}')
    
    context = {
        'system': system,
        'stream': stream,
        'selected_stream': stream,
        'status_choices': HolisticSystem.STATUS_CHOICES,
        'edit': True,
    }
    
    return render(request, 'products/holistic_system_form.html', context)

@login_required
def holistic_system_delete(request, pk, stream=None):
    """Delete a holistic system"""
    from .models import HolisticSystem
    
    stream = stream or 'HIC'
    stream_obj = get_stream_or_404(stream)
    
    system = get_object_or_404(HolisticSystem, pk=pk, stream=stream_obj)
    
    if request.method == 'POST':
        sr_no = system.sr_no
        system.delete()
        messages.success(request, f'System {sr_no} deleted successfully!')
        return redirect('holistic_dashboard')
    
    context = {
        'system': system,
        'stream': stream,
        'selected_stream': stream,
    }
    
    return render(request, 'products/holistic_system_confirm_delete.html', context)

@login_required
def holistic_system_detail(request, pk, stream=None):
    """View detailed information about a holistic system"""
    from .models import HolisticSystem
    
    stream = stream or 'HIC'
    stream_obj = get_stream_or_404(stream)
    
    system = get_object_or_404(HolisticSystem, pk=pk, stream=stream_obj)
    
    # Get all weekly data
    weekly_data = system.weekly_data.all().order_by('-year', '-week_number')[:20]
    
    # Get history
    history = system.history.all().order_by('-timestamp')[:50]
    
    context = {
        'system': system,
        'weekly_data': weekly_data,
        'history': history,
        'stream': stream,
        'selected_stream': stream,
    }
    
    return render(request, 'products/holistic_system_detail.html', context)

from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
@require_POST
def holistic_weekly_data_update(request, stream=None):
    """Update or create weekly data for a system"""
    from .models import HolisticSystem, HolisticWeeklyData
    
    try:
        system_id = request.POST.get('system_id')
        week_number = int(request.POST.get('week_number'))
        year = int(request.POST.get('year'))
        
        system = HolisticSystem.objects.get(id=system_id)
        
        # Get or create weekly data
        weekly_data, created = HolisticWeeklyData.objects.get_or_create(
            holistic_system=system,
            week_number=week_number,
            year=year,
            defaults={'updated_by': request.user}
        )
        
        # Update fields
        weekly_data.allocation_status = request.POST.get('allocation_status', '')
        weekly_data.utilization_percentage = request.POST.get('utilization_percentage', 0)
        weekly_data.assigned_to = request.POST.get('assigned_to', '')
        weekly_data.task_description = request.POST.get('task_description', '')
        weekly_data.hours_used = request.POST.get('hours_used', 0)
        weekly_data.availability_hours = request.POST.get('availability_hours', 40)
        weekly_data.notes = request.POST.get('notes', '')
        weekly_data.updated_by = request.user
        weekly_data.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Week {week_number} data updated successfully',
            'week_label': f'W{week_number}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
def holistic_export_excel(request, stream=None):
    """Export holistic systems to Excel"""
    from .models import HolisticSystem, HolisticWeeklyData
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    stream = stream or 'HIC'
    stream_obj = get_stream_or_404(stream)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Holistic Systems'
    
    # Define headers
    headers = [
        'Sr No', 'System Availability', 'Allocation to Sl No', 'Location Info',
        'STMi Number', 'System Owner', 'ECR#', 'Test Engineer',
        'Description', 'Priority', 'Notes'
    ]
    
    # Add week columns (last 12 weeks)
    current_week = date.today().isocalendar()[1]
    for i in range(12):
        week_num = current_week - i
        if week_num < 1:
            week_num += 52
        headers.append(f'W{week_num}')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='005fa3', end_color='005fa3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    systems = HolisticSystem.objects.filter(stream=stream_obj).order_by('sr_no')
    
    for row_num, system in enumerate(systems, 2):
        ws.cell(row=row_num, column=1, value=system.sr_no)
        ws.cell(row=row_num, column=2, value=system.get_system_availability_display())
        ws.cell(row=row_num, column=3, value=system.allocation_to_sl_no or '')
        ws.cell(row=row_num, column=4, value=system.location_info or '')
        ws.cell(row=row_num, column=5, value=system.stmi_number or '')
        ws.cell(row=row_num, column=6, value=system.system_owner or '')
        ws.cell(row=row_num, column=7, value=system.ecr_number or '')
        ws.cell(row=row_num, column=8, value=system.test_engineer or '')
        ws.cell(row=row_num, column=9, value=system.description or '')
        ws.cell(row=row_num, column=10, value=system.priority or '')
        ws.cell(row=row_num, column=11, value=system.notes or '')
        
        # Add weekly data
        col_offset = 12
        for i in range(12):
            week_num = current_week - i
            if week_num < 1:
                week_num += 52
            
            week_data = system.weekly_data.filter(
                week_number=week_num,
                year=date.today().year
            ).first()
            
            if week_data:
                value = f"{week_data.allocation_status or ''} ({week_data.utilization_percentage}%)"
            else:
                value = '-'
            
            ws.cell(row=row_num, column=col_offset + i, value=value)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Create response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_Holistic_Systems_{stream}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def holistic_export_pdf(request, stream=None):
    """Export holistic systems to PDF"""
    from .models import HolisticSystem
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from django.contrib.staticfiles import finders
    
    stream = stream or 'HIC'
    stream_obj = get_stream_or_404(stream)
    
    buffer = BytesIO()
    page_width, page_height = landscape(A3)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(page_width, page_height),
        leftMargin=30,
        rightMargin=30,
        topMargin=60,
        bottomMargin=30,
    )
    
    # Enhanced helper functions to truncate text for table cells
    def truncate_text(text, max_length=30):
        """Aggressively truncate text to fit in table cells and handle None values"""
        if text is None:
            return '-'
        text = str(text).strip()
        if len(text) <= max_length:
            return text
        return text[:max_length-3] + '...'
    
    def wrap_text(text, max_length=25, max_lines=2):
        """Wrap long text into multiple lines with strict limits"""
        if text is None or text == '-':
            return '-'
        text = str(text).strip()
        if len(text) <= max_length:
            return text
        
        # Break text into lines
        words = text.split(' ')
        lines = []
        current_line = ''
        
        for word in words:
            # If adding this word would exceed max_length, start a new line
            if len(current_line + ' ' + word) <= max_length:
                current_line += (' ' + word if current_line else word)
            else:
                if current_line:
                    lines.append(current_line)
                    if len(lines) >= max_lines:  # Stop if we've reached max lines
                        break
                current_line = word[:max_length]  # Truncate very long words
        
        if current_line and len(lines) < max_lines:
            lines.append(current_line)
        
        # If we had to truncate due to max_lines, add ellipsis
        if len(lines) == max_lines and len(words) > len(' '.join(lines).split()):
            if len(lines) > 0:
                lines[-1] = lines[-1][:max_length-3] + '...'
        
        return '\n'.join(lines[:max_lines])

    def smart_truncate(text, max_length=20):
        """Smart truncation that preserves important parts"""
        if text is None:
            return '-'
        text = str(text).strip()
        if len(text) <= max_length:
            return text
        
        # For certain patterns, try to keep meaningful parts
        if '@' in text:  # Email addresses
            parts = text.split('@')
            if len(parts) == 2:
                username_len = max(3, max_length - len(parts[1]) - 4)
                return parts[0][:username_len] + '@' + parts[1]
        
        return text[:max_length-3] + '...'

    # Prepare data
    headers = [
        'Sr No', 'Availability', 'Allocation', 'Location', 'STMi#',
        'Owner', 'ECR#', 'Test Eng', 'Priority', 'Description', 'Notes'
    ]
    
    data = [headers]
    
    systems = HolisticSystem.objects.filter(stream=stream_obj).order_by('sr_no')
    
    for system in systems:
        data.append([
            truncate_text(system.sr_no, 10),
            truncate_text(system.get_system_availability_display(), 8),
            truncate_text(system.allocation_to_sl_no, 12),
            wrap_text(system.location_info, 15, 2),
            truncate_text(system.stmi_number, 10),
            smart_truncate(system.system_owner, 12),
            truncate_text(system.ecr_number, 10),
            smart_truncate(system.test_engineer, 12),
            truncate_text(system.priority, 8),
            wrap_text(system.description, 20, 2),
            wrap_text(system.notes, 20, 2),
        ])
    
    # Set more conservative column widths for better fitting
    col_widths = [50, 60, 65, 80, 60, 75, 60, 75, 50, 100, 100]
    
    table = Table(data, repeatRows=1, colWidths=col_widths)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005fa3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align text to top for wrapped content
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),  # Further reduced font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('WORDWRAP', (0, 0), (-1, -1), 'LTR'),  # Enable word wrapping
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),  # Alternating row colors
    ])
    table.setStyle(style)

    # Styles for titles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.fontName = 'Helvetica-Bold'
    title_style.fontSize = 22
    title_style.alignment = 1  # Center

    subtitle_style = styles['Heading2']
    subtitle_style.fontName = 'Helvetica-Bold'
    subtitle_style.fontSize = 18
    subtitle_style.alignment = 1

    subsubtitle_style = styles['Heading3']
    subsubtitle_style.fontName = 'Helvetica-Bold'
    subsubtitle_style.fontSize = 16
    subsubtitle_style.alignment = 1

    note_style = styles['Italic']
    note_style.fontSize = 12
    note_style.alignment = 1

    # Logo path
    logo_path = finders.find('products/philips.png')

    def draw_header(canvas, doc):
        # Draw logo
        if logo_path:
            logo_width = 50
            logo_height = 50
            x_logo = page_width - logo_width - 40
            y_logo = page_height - logo_height - 20
            canvas.drawImage(logo_path, x_logo, y_logo, width=logo_width, height=logo_height, mask='auto')
        # Draw titles
        canvas.setFont("Helvetica-Bold", 22)
        canvas.drawCentredString(page_width / 2, page_height - 60, "Image Guided Therapy (IGT)")
        canvas.setFont("Helvetica-Bold", 18)
        canvas.drawCentredString(page_width / 2, page_height - 90, "Mobile Surgery (MoS)")
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(page_width / 2, page_height - 120, "Holistic Systems Dashboard")
        canvas.setFont("Helvetica-Oblique", 12)
        canvas.drawCentredString(page_width / 2, page_height - 145, "(Note: Automated data output. Verification recommended to ensure reliability and compliance with organizational protocols.)")

    elements = [Spacer(1, 120), table]  # Reduced gap above the table

    # --- Weekly Data Table ---
    from .models import HolisticWeeklyData
    from datetime import date, timedelta
    
    # Get last 4 weeks data
    current_week = date.today().isocalendar()[1]
    current_year = date.today().year
    
    weekly_headers = ['System Sr No', 'Week', 'Project', 'Utilization %', 'Assigned To', 'Task Description', 'Hours Used', 'Notes']
    weekly_data = [weekly_headers]
    
    for i in range(4):
        week_num = current_week - i
        if week_num < 1:
            week_num += 52
            year = current_year - 1
        else:
            year = current_year
            
        week_data = HolisticWeeklyData.objects.filter(
            week_number=week_num,
            year=year,
            holistic_system__stream=stream_obj
        ).select_related('holistic_system', 'project')
        
        for week in week_data:
            weekly_data.append([
                truncate_text(week.holistic_system.sr_no, 10),
                f'W{week.week_number}',
                truncate_text(week.project.name if week.project else '-', 15),
                f'{week.utilization_percentage}%' if week.utilization_percentage else '0%',
                smart_truncate(week.assigned_to, 12),
                wrap_text(week.task_description, 18, 2),
                f'{week.hours_used}h' if week.hours_used else '0h',
                wrap_text(week.notes, 15, 2)
            ])
    
    if len(weekly_data) > 1:
        weekly_table = Table(weekly_data, repeatRows=1)
        weekly_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005fa3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align text to top for wrapped content
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),  # Further reduced font size
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('WORDWRAP', (0, 0), (-1, -1), 'LTR'),  # Enable word wrapping
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),  # Alternating row colors
        ])
        weekly_table.setStyle(weekly_style)
        # Set more conservative column widths for Weekly Data table
        col_widths = [65, 40, 85, 50, 75, 120, 50, 90]
        weekly_table._argW = col_widths
        elements.append(Spacer(1, 40))
        elements.append(Paragraph('Weekly Assignment Data (Last 4 Weeks)', title_style))
        elements.append(Spacer(1, 10))
        elements.append(weekly_table)

    doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header)
    
    pdf = buffer.getvalue()
    buffer.close()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_Holistic_Systems_{stream}.pdf"
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    
    return response

@login_required
def holistic_bulk_update(request, stream=None):
    """Bulk update holistic systems"""
    from .models import HolisticSystem, HolisticSystemHistory
    
    if request.method == 'POST':
        try:
            system_ids = request.POST.getlist('system_ids[]')
            action = request.POST.get('action')
            value = request.POST.get('value')
            
            systems = HolisticSystem.objects.filter(id__in=system_ids)
            
            updated_count = 0
            for system in systems:
                old_value = getattr(system, action)
                
                if action == 'system_availability':
                    system.system_availability = value
                elif action == 'priority':
                    system.priority = value
                
                system.updated_by = request.user
                system.save()
                
                # Create history
                HolisticSystemHistory.objects.create(
                    holistic_system=system,
                    action='bulk_updated',
                    user=request.user,
                    details=f'{action}: {old_value} → {value}'
                )
                
                updated_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'{updated_count} systems updated successfully'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

from django.views.decorators.csrf import csrf_exempt
# AJAX endpoint to assign project to a specific week for a HolisticSystem
@login_required
@csrf_exempt
def holistic_assign_project_to_week(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})
    try:
        data = json.loads(request.body.decode('utf-8'))
        system_id = data.get('system_id')
        project_id = data.get('project_id')
        week_number = data.get('week_number')
        year = data.get('year', date.today().year)
        
        from .models import HolisticSystem, Project, HolisticWeeklyData
        
        system = HolisticSystem.objects.get(id=system_id)
        project = Project.objects.get(id=project_id) if project_id else None
        
        # Get or create weekly data for this week
        weekly_data, created = HolisticWeeklyData.objects.get_or_create(
            holistic_system=system,
            week_number=week_number,
            year=year,
            defaults={'updated_by': request.user}
        )
        
        # Assign project to this week
        weekly_data.project = project
        weekly_data.updated_by = request.user
        weekly_data.save()
        
        # Create history record
        from .models import HolisticSystemHistory
        action = f"Assigned project '{project.name}' to week W{week_number} {year}" if project else f"Removed project from week W{week_number} {year}"
        HolisticSystemHistory.objects.create(
            holistic_system=system,
            action='project_assigned',
            user=request.user,
            details=action
        )
        
        return JsonResponse({'success': True, 'message': f'Project assignment updated for week W{week_number}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# AJAX endpoint to get weekly data for a system
@login_required
@csrf_exempt
def holistic_get_week_data(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})
    try:
        data = json.loads(request.body.decode('utf-8'))
        system_id = data.get('system_id')
        week_number = data.get('week_number')
        year = data.get('year')
        
        from .models import HolisticSystem, HolisticWeeklyData
        
        system = HolisticSystem.objects.get(id=system_id)
        
        # Try to get existing weekly data
        try:
            weekly_data = HolisticWeeklyData.objects.get(
                holistic_system=system,
                week_number=week_number,
                year=year
            )
            
            week_data = {
                'week_number': weekly_data.week_number,
                'year': weekly_data.year,
                'allocation_status': weekly_data.allocation_status or '',
                'utilization_percentage': float(weekly_data.utilization_percentage) if weekly_data.utilization_percentage else 0,
                'assigned_to': weekly_data.assigned_to or '',
                'hours_used': float(weekly_data.hours_used) if weekly_data.hours_used else 0,
                'availability_hours': float(weekly_data.availability_hours) if weekly_data.availability_hours else 40,
                'task_description': weekly_data.task_description or '',
                'notes': weekly_data.notes or '',
            }
            
            return JsonResponse({'success': True, 'week_data': week_data})
            
        except HolisticWeeklyData.DoesNotExist:
            # Return empty data for new week
            week_data = {
                'week_number': week_number,
                'year': year,
                'allocation_status': '',
                'utilization_percentage': 0,
                'assigned_to': '',
                'hours_used': 0,
                'availability_hours': 40,
                'task_description': '',
                'notes': '',
            }
            
            return JsonResponse({'success': False, 'week_data': week_data})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# AJAX endpoint to get project assignments for a system across multiple weeks
@login_required
@csrf_exempt
def holistic_get_project_assignments(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})
    try:
        data = json.loads(request.body.decode('utf-8'))
        system_id = data.get('system_id')
        
        from .models import HolisticSystem, HolisticWeeklyData
        
        system = HolisticSystem.objects.get(id=system_id)
        weekly_data = system.weekly_data.select_related('project').order_by('year', 'week_number')
        
        assignments = []
        for week in weekly_data:
            assignments.append({
                'week_number': week.week_number,
                'year': week.year,
                'project_id': week.project.id if week.project else None,
                'project_name': week.project.name if week.project else None,
            })
        
        return JsonResponse({'success': True, 'assignments': assignments})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def holistic_all_systems_list(request):
    """Fetch list of all systems for selection"""
    from .models import HolisticSystem
    
    try:
        systems = HolisticSystem.objects.all().select_related('created_by').order_by('sr_no')
        
        systems_data = []
        for system in systems:
            # Calculate average utilization for each system
            weekly_data = system.weekly_data.all()
            utilizations = [float(w.utilization_percentage) for w in weekly_data if w.utilization_percentage is not None]
            avg_utilization = sum(utilizations) / len(utilizations) if utilizations else 0
            systems_data.append({
                'id': system.id,
                'sr_no': system.sr_no,
                'status': system.get_system_availability_display(),
                'owner': system.system_owner or '-',
                'location': system.location_info or '-',
                'stmi': system.stmi_number or '-',
                'priority': system.get_priority_display() if hasattr(system, 'get_priority_display') else system.priority or '-',
                'avg_utilization': round(avg_utilization, 2)
            })
        
        return JsonResponse({'success': True, 'systems': systems_data})
    
    except Exception as e:
        logger.error(f"Error fetching systems list: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def holistic_graph_data(request, system_id):
    """Fetch allocation data for graph visualization"""
    from .models import HolisticSystem, HolisticWeeklyData, Project
    from django.db.models import Count, Q
    
    try:
        system = HolisticSystem.objects.get(id=system_id)
        
        # Get all weekly data for this system
        weekly_data = HolisticWeeklyData.objects.filter(
            holistic_system=system
        ).order_by('year', 'week_number')
        
        # Prepare weekly data response
        weekly_data_list = []
        utilization_values = []
        
        for week in weekly_data:
            # Get the project assigned directly in this week
            project_name = week.project.name if week.project else 'Unassigned'
            
            utilization = float(week.utilization_percentage) if week.utilization_percentage else 0
            utilization_values.append(utilization)
            
            weekly_data_list.append({
                'week': week.week_number,
                'year': week.year,
                'status': week.allocation_status or system.get_system_availability_display(),
                'project_name': project_name,
                'utilization': utilization,
                'hours_used': float(week.hours_used) if week.hours_used else 0,
                'assigned_to': week.assigned_to or '-'
            })
        
        # Calculate project distribution from weekly assignments
        project_distribution = []
        project_weeks = {}
        
        # Count weeks per project from weekly data
        for week in weekly_data_list:
            project_name = week['project_name']
            if project_name in project_weeks:
                project_weeks[project_name] += 1
            else:
                project_weeks[project_name] = 1
        
        # Create distribution list
        for project_name, weeks_count in project_weeks.items():
            project_distribution.append({
                'project_name': project_name,
                'weeks_count': weeks_count
            })
        
        # Calculate statistics
        avg_utilization = sum(utilization_values) / len(utilization_values) if utilization_values else 0
        max_utilization = max(utilization_values) if utilization_values else 0
        
        response_data = {
            'success': True,
            'system': {
                'id': system.id,
                'name': system.sr_no,
                'status': system.get_system_availability_display(),
                'owner': system.system_owner,
                'location': system.location_info
            },
            'weekly_data': weekly_data_list,
            'project_distribution': project_distribution,
            'statistics': {
                'total_weeks': len(weekly_data_list),
                'avg_utilization': avg_utilization,
                'max_utilization': max_utilization,
                'unique_projects': len(project_weeks)
            }
        }
        
        return JsonResponse(response_data)
    
    except HolisticSystem.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'System not found'}, status=404)
    except Exception as e:
        logger.error(f"Error fetching graph data: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def holistic_export_graph_data(request, system_id):
    """Export system allocation data as CSV"""
    from .models import HolisticSystem, HolisticWeeklyData
    import csv
    from io import StringIO
    
    try:
        system = HolisticSystem.objects.get(id=system_id)
        weekly_data = HolisticWeeklyData.objects.filter(
            holistic_system=system
        ).order_by('year', 'week_number')
        
        # Create CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'System Name', 'Week', 'Year', 'Status', 'Allocation Status',
            'Utilization %', 'Hours Used', 'Available Hours', 'Assigned To',
            'Task Description', 'Notes'
        ])
        
        # Write data rows
        for week in weekly_data:
            writer.writerow([
                system.sr_no,
                f'W{week.week_number}',
                week.year,
                system.get_system_availability_display(),
                week.allocation_status or '-',
                float(week.utilization_percentage) or 0,
                float(week.hours_used) or 0,
                float(week.availability_hours) or 40,
                week.assigned_to or '-',
                week.task_description or '-',
                week.notes or '-'
            ])
        
        # Create response
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="system_{system.sr_no}_allocation.csv"'
        return response
    
    except HolisticSystem.DoesNotExist:
        return JsonResponse({'error': 'System not found'}, status=404)
    except Exception as e:
        logger.error(f"Error exporting graph data: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# ================================
# DOWNTIME TRACKING VIEWS
# ================================

@login_required
@csrf_exempt
@require_http_methods(["GET", "POST"])
def system_downtime_events(request, stream=None, system_id=None):
    """
    Handle downtime events - GET to retrieve, POST to create
    """
    from .models import SystemDowntime
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check permissions
    if not can_manage_system_allocation(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'GET':
        # Retrieve downtime events
        if system_id:
            system = get_object_or_404(System, id=system_id, stream=stream_obj)
            downtimes = SystemDowntime.objects.filter(system=system)
        else:
            downtimes = SystemDowntime.objects.filter(stream=stream_obj)
        
        # Apply filters
        status_filter = request.GET.get('status')
        if status_filter:
            downtimes = downtimes.filter(status=status_filter)
        
        limit = int(request.GET.get('limit', 50))
        downtimes = downtimes.order_by('-start_time')[:limit]
        
        downtime_data = []
        for downtime in downtimes:
            downtime_data.append({
                'id': downtime.id,
                'system_id': downtime.system.id,
                'system_name': downtime.system.name,
                'title': downtime.title,
                'description': downtime.description,
                'downtime_type': downtime.downtime_type,
                'downtime_type_display': downtime.get_downtime_type_display(),
                'impact_level': downtime.impact_level,
                'impact_level_display': downtime.get_impact_level_display(),
                'status': downtime.status,
                'status_display': downtime.get_status_display(),
                'start_time': downtime.start_time.isoformat(),
                'end_time': downtime.end_time.isoformat() if downtime.end_time else None,
                'duration_hours': round(downtime.duration_hours, 2),
                'is_ongoing': downtime.is_ongoing,
                'reported_by': downtime.reported_by.username if downtime.reported_by else None,
                'assigned_to': downtime.assigned_to.username if downtime.assigned_to else None,
                'resolved_by': downtime.resolved_by.username if downtime.resolved_by else None,
                'root_cause': downtime.root_cause,
                'resolution_steps': downtime.resolution_steps,
                'external_ticket_id': downtime.external_ticket_id,
                'created_at': downtime.created_at.isoformat(),
                'updated_at': downtime.updated_at.isoformat(),
            })
        
        return JsonResponse({'downtimes': downtime_data})
    
    elif request.method == 'POST':
        # Create new downtime event
        try:
            data = json.loads(request.body)
            
            system_id = data.get('system_id')
            if not system_id:
                return JsonResponse({'error': 'system_id is required'}, status=400)
            
            system = get_object_or_404(System, id=system_id, stream=stream_obj)
            
            # Parse datetime
            start_time_str = data.get('start_time')
            if not start_time_str:
                start_time = timezone.now()
            else:
                try:
                    start_time = datetime.fromisoformat(start_time_str.replace('Z', '+00:00'))
                    if timezone.is_naive(start_time):
                        start_time = timezone.make_aware(start_time)
                except ValueError:
                    return JsonResponse({'error': 'Invalid start_time format'}, status=400)
            
            end_time = None
            if data.get('end_time'):
                try:
                    end_time = datetime.fromisoformat(data['end_time'].replace('Z', '+00:00'))
                    if timezone.is_naive(end_time):
                        end_time = timezone.make_aware(end_time)
                except ValueError:
                    return JsonResponse({'error': 'Invalid end_time format'}, status=400)
            
            # Create downtime event
            downtime = SystemDowntime.objects.create(
                system=system,
                stream=stream_obj,
                title=data.get('title', 'System Downtime'),
                description=data.get('description', ''),
                downtime_type=data.get('downtime_type', 'other'),
                impact_level=data.get('impact_level', 'medium'),
                status=data.get('status', 'ongoing'),
                start_time=start_time,
                end_time=end_time,
                root_cause=data.get('root_cause', ''),
                resolution_steps=data.get('resolution_steps', ''),
                external_ticket_id=data.get('external_ticket_id', ''),
                reported_by=request.user
            )
            
            # Handle assigned_to
            assigned_to_id = data.get('assigned_to_id')
            if assigned_to_id:
                try:
                    assigned_user = User.objects.get(id=assigned_to_id)
                    downtime.assigned_to = assigned_user
                    downtime.save()
                except User.DoesNotExist:
                    pass
            
            return JsonResponse({
                'success': True,
                'downtime_id': downtime.id,
                'message': 'Downtime event created successfully'
            })
        
        except Exception as e:
            logger.error(f"Error creating downtime event: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)

@login_required
@csrf_exempt
@require_http_methods(["PUT", "DELETE"])
def system_downtime_event_detail(request, stream=None, downtime_id=None):
    """
    Handle individual downtime events - PUT to update, DELETE to remove
    """
    from .models import SystemDowntime
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check permissions
    if not can_manage_system_allocation(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    downtime = get_object_or_404(SystemDowntime, id=downtime_id, stream=stream_obj)
    
    if request.method == 'PUT':
        # Update downtime event
        try:
            data = json.loads(request.body)
            
            # Update fields if provided
            if 'title' in data:
                downtime.title = data['title']
            if 'description' in data:
                downtime.description = data['description']
            if 'downtime_type' in data:
                downtime.downtime_type = data['downtime_type']
            if 'impact_level' in data:
                downtime.impact_level = data['impact_level']
            if 'status' in data:
                downtime.status = data['status']
            if 'root_cause' in data:
                downtime.root_cause = data['root_cause']
            if 'resolution_steps' in data:
                downtime.resolution_steps = data['resolution_steps']
            if 'external_ticket_id' in data:
                downtime.external_ticket_id = data['external_ticket_id']
            
            # Handle end_time
            if 'end_time' in data:
                if data['end_time']:
                    try:
                        end_time = datetime.fromisoformat(data['end_time'].replace('Z', '+00:00'))
                        if timezone.is_naive(end_time):
                            end_time = timezone.make_aware(end_time)
                        downtime.end_time = end_time
                    except ValueError:
                        return JsonResponse({'error': 'Invalid end_time format'}, status=400)
                else:
                    downtime.end_time = None
            
            # Handle resolution
            if data.get('resolve') and not downtime.end_time:
                downtime.resolve(resolved_by_user=request.user, resolution_notes=data.get('resolution_notes'))
            
            # Handle assigned_to
            if 'assigned_to_id' in data:
                if data['assigned_to_id']:
                    try:
                        assigned_user = User.objects.get(id=data['assigned_to_id'])
                        downtime.assigned_to = assigned_user
                    except User.DoesNotExist:
                        return JsonResponse({'error': 'Assigned user not found'}, status=400)
                else:
                    downtime.assigned_to = None
            
            downtime.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Downtime event updated successfully'
            })
        
        except Exception as e:
            logger.error(f"Error updating downtime event: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'DELETE':
        # Delete downtime event
        try:
            downtime.delete()
            return JsonResponse({
                'success': True,
                'message': 'Downtime event deleted successfully'
            })
        
        except Exception as e:
            logger.error(f"Error deleting downtime event: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_GET
def system_downtime_metrics(request, stream=None, system_id=None):
    """
    Get downtime metrics for a system or all systems in a stream
    """
    from .models import SystemDowntime, SystemDowntimeMetrics
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check permissions
    if not can_view_analytics(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    # Get time period (default to last 30 days)
    days = int(request.GET.get('days', 30))
    end_date = timezone.now()
    start_date = end_date - timedelta(days=days)
    
    if system_id:
        # Get metrics for specific system
        system = get_object_or_404(System, id=system_id, stream=stream_obj)
        metrics = system.get_downtime_metrics(days)
        
        # Get recent downtime events
        recent_events = SystemDowntime.objects.filter(
            system=system,
            start_time__gte=start_date
        ).order_by('-start_time')[:10]
        
        events_data = []
        for event in recent_events:
            events_data.append({
                'id': event.id,
                'title': event.title,
                'downtime_type': event.get_downtime_type_display(),
                'impact_level': event.get_impact_level_display(),
                'status': event.get_status_display(),
                'start_time': event.start_time.isoformat(),
                'duration_hours': round(event.duration_hours, 2),
                'is_ongoing': event.is_ongoing
            })
        
        return JsonResponse({
            'system_id': system.id,
            'system_name': system.name,
            'period_days': days,
            'metrics': {
                'availability_percentage': round(metrics.availability_percentage, 2),
                'total_downtime_hours': round(metrics.total_downtime_hours, 2),
                'total_incidents': metrics.total_incidents,
                'planned_downtime_hours': round(metrics.planned_downtime_hours, 2),
                'unplanned_downtime_hours': round(metrics.unplanned_downtime_hours, 2),
                'mean_time_to_repair_hours': round(metrics.mean_time_to_repair_hours, 2) if metrics.mean_time_to_repair_hours else None,
                'mean_time_between_failures_hours': round(metrics.mean_time_between_failures_hours, 2) if metrics.mean_time_between_failures_hours else None,
                'most_common_downtime_type': metrics.most_common_downtime_type,
                'most_common_impact_level': metrics.most_common_impact_level,
            },
            'recent_events': events_data
        })
    
    else:
        # Get metrics for all systems in stream
        systems = System.objects.filter(stream=stream_obj)
        systems_metrics = []
        
        for system in systems:
            metrics = system.get_downtime_metrics(days)
            systems_metrics.append({
                'system_id': system.id,
                'system_name': system.name,
                'availability_percentage': round(metrics.availability_percentage, 2),
                'total_downtime_hours': round(metrics.total_downtime_hours, 2),
                'total_incidents': metrics.total_incidents,
                'is_currently_down': system.is_currently_down(),
                'current_status': system.status
            })
        
        # Calculate stream-wide metrics
        total_systems = len(systems_metrics)
        total_incidents = sum(s['total_incidents'] for s in systems_metrics)
        avg_availability = sum(s['availability_percentage'] for s in systems_metrics) / total_systems if total_systems > 0 else 100
        systems_down = sum(1 for s in systems_metrics if s['is_currently_down'])
        
        return JsonResponse({
            'stream': stream,
            'period_days': days,
            'summary': {
                'total_systems': total_systems,
                'systems_currently_down': systems_down,
                'average_availability': round(avg_availability, 2),
                'total_incidents': total_incidents
            },
            'systems': systems_metrics
        })

@login_required
@csrf_exempt
@require_POST
def resolve_downtime(request, stream=None, downtime_id=None):
    """
    Resolve a specific downtime event
    """
    from .models import SystemDowntime
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check permissions
    if not can_manage_system_allocation(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    downtime = get_object_or_404(SystemDowntime, id=downtime_id, stream=stream_obj)
    
    try:
        data = json.loads(request.body)
        resolution_notes = data.get('resolution_notes', '')
        
        downtime.resolve(resolved_by_user=request.user, resolution_notes=resolution_notes)
        
        return JsonResponse({
            'success': True,
            'message': 'Downtime resolved successfully',
            'end_time': downtime.end_time.isoformat(),
            'duration_hours': round(downtime.duration_hours, 2)
        })
    
    except Exception as e:
        logger.error(f"Error resolving downtime: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_GET
def downtime_dashboard(request, stream=None):
    """
    Render the downtime dashboard with current status and metrics
    """
    from .models import SystemDowntime
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check permissions
    if not can_view_analytics(request.user):
        messages.error(request, 'Access denied. You need appropriate privileges to view downtime dashboard.')
        return redirect('dashboard')
    
    # Get systems and current downtime status
    systems = System.objects.filter(stream=stream_obj).order_by('name')
    
    # Get ongoing downtimes
    ongoing_downtimes = SystemDowntime.objects.filter(
        stream=stream_obj,
        status='ongoing'
    ).order_by('-start_time')
    
    # Get recent resolved downtimes
    recent_downtimes = SystemDowntime.objects.filter(
        stream=stream_obj,
        status='resolved'
    ).order_by('-end_time')[:20]
    
    # Get users for assignment
    users = User.objects.filter(is_active=True).order_by('username')
    
    context = {
        'stream': stream or 'HIC',
        'selected_stream': stream or 'HIC',
        'systems': systems,
        'ongoing_downtimes': ongoing_downtimes,
        'recent_downtimes': recent_downtimes,
        'users': users,
        'downtime_types': SystemDowntime.DOWNTIME_TYPES,
        'impact_levels': SystemDowntime.IMPACT_LEVELS,
        'status_choices': SystemDowntime.STATUS_CHOICES,
        'can_manage': can_manage_system_allocation(request.user)
    }
    
    return render(request, 'products/downtime_dashboard.html', context)


# ================================
# BUILD SERVERS VIEWS
# ================================

@login_required
def build_servers_dashboard(request):
    """
    Main Build Servers dashboard with stream selection
    """
    # Get all streams for selection
    streams = Stream.objects.filter(is_active=True).order_by('name')
    
    # If user is not super admin, limit to their accessible streams
    if hasattr(request.user, 'custom_profile') and not request.user.custom_profile.is_super_admin():
        user_streams = request.user.custom_profile.get_accessible_streams()
        streams = user_streams.filter(is_active=True).order_by('name')
    
    # Set a default stream for the base template
    selected_stream = streams.first() if streams.exists() else None
    
    context = {
        'streams': streams,
        'selected_stream': selected_stream,
        'stream': selected_stream,  # For base template compatibility
    }
    
    return render(request, 'products/build_servers_dashboard.html', context)

@login_required
def build_servers_list(request, stream=None):
    """
    List all build servers for a specific stream
    """
    from .models import BuildServer
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check if user can access this stream
    if hasattr(request.user, 'custom_profile') and not request.user.custom_profile.can_access_stream(stream_obj.name):
        messages.error(request, f'Access denied. You do not have permission to view {stream_obj.name} build servers.')
        return redirect('build_servers_dashboard')
    
    # Get servers for this stream, ordered by hostname
    servers = BuildServer.objects.filter(stream=stream_obj).order_by('hostname')
    
    # Calculate status counts for all servers in this stream (before filtering)
    from django.db.models import Count, Case, When, IntegerField
    all_servers = BuildServer.objects.filter(stream=stream_obj)
    status_counts = {}
    for status, _ in BuildServer.STATUS_CHOICES:
        status_counts[status] = all_servers.filter(status=status).count()
    
    # Apply filters
    status_filter = request.GET.get('status')
    if status_filter:
        servers = servers.filter(status=status_filter)
    
    stream_type_filter = request.GET.get('stream_type')
    if stream_type_filter:
        servers = servers.filter(stream_type=stream_type_filter)
    
    floor_filter = request.GET.get('floor')
    if floor_filter:
        servers = servers.filter(floor=floor_filter)
    
    search = request.GET.get('search')
    if search:
        servers = servers.filter(
            Q(hostname__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(location__icontains=search) |
            Q(owner__icontains=search) |
            Q(purpose__icontains=search)
        )
    
    context = {
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,  # Add for base template compatibility
        'servers': servers,
        'status_counts': status_counts,
        'status_choices': BuildServer.STATUS_CHOICES,
        'stream_type_choices': BuildServer.SERVER_TYPES,
        'floor_choices': [(floor.id, floor.name) for floor in Floor.objects.filter(stream=stream_obj, is_active=True)],
        'current_filters': {
            'status': status_filter,
            'stream_type': stream_type_filter,
            'floor': floor_filter,
            'search': search,
        }
    }
    
    return render(request, 'products/build_servers_list.html', context)

@login_required
def build_server_create(request, stream=None):
    """
    Create a new build server
    """
    from .models import BuildServer, BuildServerHistory
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check permissions
    if not can_edit_products(request.user):
        messages.error(request, 'Access denied. You do not have permission to create build servers.')
        return redirect('build_servers_list', stream=stream)
    
    if request.method == 'POST':
        try:
            # Create the build server
            floor_id = request.POST.get('floor')
            floor_instance = None
            if floor_id:
                try:
                    floor_instance = Floor.objects.get(id=floor_id)
                except Floor.DoesNotExist:
                    floor_instance = None
            server = BuildServer(
                hostname=request.POST.get('hostname'),
                ip_address=request.POST.get('ip_address'),
                location=request.POST.get('location'),
                floor=floor_instance,
                owner=request.POST.get('owner'),
                stream_type=(stream_obj.name if stream_obj.name in ['PIC', 'HIC'] else 'Other'),
                stream=stream_obj,
                status=request.POST.get('status', 'Active'),
                operating_system=request.POST.get('operating_system', ''),
                cpu_cores=request.POST.get('cpu_cores') or None,
                ram_gb=request.POST.get('ram_gb') or None,
                storage_gb=request.POST.get('storage_gb') or None,
                mac_address=request.POST.get('mac_address', ''),
                domain=request.POST.get('domain', ''),
                ssh_port=request.POST.get('ssh_port', 22),
                purpose=request.POST.get('purpose', ''),
                project_allocation=request.POST.get('project_allocation', ''),
                cost_center=request.POST.get('cost_center', ''),
                primary_contact=request.POST.get('primary_contact', ''),
                secondary_contact=request.POST.get('secondary_contact', ''),
                contact_email=request.POST.get('contact_email', ''),
                notes=request.POST.get('notes', ''),
                tags=request.POST.get('tags', ''),
                created_by=request.user,
                updated_by=request.user
            )
            
            # Handle date fields
            procurement_date = request.POST.get('procurement_date')
            if procurement_date:
                server.procurement_date = datetime.strptime(procurement_date, '%Y-%m-%d').date()
            
            warranty_expiry = request.POST.get('warranty_expiry')
            if warranty_expiry:
                server.warranty_expiry = datetime.strptime(warranty_expiry, '%Y-%m-%d').date()
            
            last_maintenance = request.POST.get('last_maintenance')
            if last_maintenance:
                server.last_maintenance = datetime.strptime(last_maintenance, '%Y-%m-%d').date()
            
            next_maintenance = request.POST.get('next_maintenance')
            if next_maintenance:
                server.next_maintenance = datetime.strptime(next_maintenance, '%Y-%m-%d').date()
            
            server.save()
            
            # Create history entry
            BuildServerHistory.objects.create(
                build_server=server,
                action='created',
                user=request.user,
                details=f'Build server {server.hostname} created'
            )
            
            messages.success(request, f'Build server {server.hostname} created successfully!')
            return redirect('build_servers_list', stream=stream)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            messages.error(request, f'Error creating build server: {str(e)}')
            messages.error(request, f'Details: {error_details}')
    
    context = {
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,
        'status_choices': BuildServer.STATUS_CHOICES,
        'stream_type_choices': BuildServer.SERVER_TYPES,
        'floor_choices': [(floor.id, floor.name) for floor in Floor.objects.filter(stream=stream_obj, is_active=True)],
    }
    
    return render(request, 'products/build_server_form.html', context)

@login_required
def build_server_detail(request, stream, server_id):
    """
    View detailed information about a build server
    """
    from .models import BuildServer, BuildServerHistory, BuildServerMaintenanceLog
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Get the server
    server = get_object_or_404(BuildServer, id=server_id, stream=stream_obj)
    
    # Get history
    history = BuildServerHistory.objects.filter(build_server=server).order_by('-timestamp')[:20]
    
    # Get maintenance logs
    maintenance_logs = BuildServerMaintenanceLog.objects.filter(build_server=server).order_by('-scheduled_date')[:10]
    
    context = {
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,
        'server': server,
        'history': history,
        'maintenance_logs': maintenance_logs,
    }
    
    return render(request, 'products/build_server_detail.html', context)

@login_required
def build_server_edit(request, stream, server_id):
    """
    Edit a build server
    """
    from .models import BuildServer, BuildServerHistory
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Get the server
    server = get_object_or_404(BuildServer, id=server_id, stream=stream_obj)
    
    # Check permissions
    if not can_edit_products(request.user):
        messages.error(request, 'Access denied. You do not have permission to edit build servers.')
        return redirect('build_server_detail', stream=stream, server_id=server_id)
    
    if request.method == 'POST':
        try:
            # Store old values for history
            old_values = {
                'hostname': server.hostname,
                'ip_address': server.ip_address,
                'location': server.location,
                'floor': server.floor,
                'owner': server.owner,
                'status': server.status
            }
            
            # Update the build server
            server.hostname = request.POST.get('hostname')
            server.ip_address = request.POST.get('ip_address')
            server.location = request.POST.get('location')
            server.floor = request.POST.get('floor')
            server.owner = request.POST.get('owner')
            # Stream type is now static, not editable from form
            # server.stream_type remains unchanged
            
            # Handle status change with explicit validation
            new_status = request.POST.get('status')
            if new_status and new_status in dict(BuildServer.STATUS_CHOICES):
                server.status = new_status
            else:
                # Keep existing status if no valid status provided
                pass
            server.operating_system = request.POST.get('operating_system', '')
            server.cpu_cores = request.POST.get('cpu_cores') or None
            server.ram_gb = request.POST.get('ram_gb') or None
            server.storage_gb = request.POST.get('storage_gb') or None
            server.mac_address = request.POST.get('mac_address', '')
            server.domain = request.POST.get('domain', '')
            server.ssh_port = request.POST.get('ssh_port', 22)
            server.purpose = request.POST.get('purpose', '')
            server.project_allocation = request.POST.get('project_allocation', '')
            server.cost_center = request.POST.get('cost_center', '')
            server.primary_contact = request.POST.get('primary_contact', '')
            server.secondary_contact = request.POST.get('secondary_contact', '')
            server.contact_email = request.POST.get('contact_email', '')
            server.notes = request.POST.get('notes', '')
            server.tags = request.POST.get('tags', '')
            server.updated_by = request.user
            
            # Handle date fields
            procurement_date = request.POST.get('procurement_date')
            if procurement_date:
                server.procurement_date = datetime.strptime(procurement_date, '%Y-%m-%d').date()
            else:
                server.procurement_date = None
            
            warranty_expiry = request.POST.get('warranty_expiry')
            if warranty_expiry:
                server.warranty_expiry = datetime.strptime(warranty_expiry, '%Y-%m-%d').date()
            else:
                server.warranty_expiry = None
            
            last_maintenance = request.POST.get('last_maintenance')
            if last_maintenance:
                server.last_maintenance = datetime.strptime(last_maintenance, '%Y-%m-%d').date()
            else:
                server.last_maintenance = None
            
            next_maintenance = request.POST.get('next_maintenance')
            if next_maintenance:
                server.next_maintenance = datetime.strptime(next_maintenance, '%Y-%m-%d').date()
            else:
                server.next_maintenance = None
            
            server.save()
            
            # Create history entry
            new_values = {
                'hostname': server.hostname,
                'ip_address': server.ip_address,
                'location': server.location,
                'floor': server.floor,
                'owner': server.owner,
                'status': server.status
            }
            
            BuildServerHistory.objects.create(
                build_server=server,
                action='updated',
                user=request.user,
                details=f'Build server {server.hostname} updated',
                old_values=old_values,
                new_values=new_values
            )
            
            messages.success(request, f'Build server {server.hostname} updated successfully!')
            return redirect('build_server_detail', stream=stream, server_id=server_id)
            
        except Exception as e:
            messages.error(request, f'Error updating build server: {str(e)}')
    
    context = {
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,
        'server': server,
        'status_choices': BuildServer.STATUS_CHOICES,
        'stream_type_choices': BuildServer.SERVER_TYPES,
        'floor_choices': [(floor.id, floor.name) for floor in Floor.objects.filter(stream=stream_obj, is_active=True)],
        'is_edit': True,
    }
    
    return render(request, 'products/build_server_form.html', context)

@login_required
def build_server_delete(request, stream, server_id):
    """
    Delete a build server
    """
    from .models import BuildServer, BuildServerHistory
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Get the server
    server = get_object_or_404(BuildServer, id=server_id, stream=stream_obj)
    
    # Check permissions
    if not can_delete_products(request.user):
        messages.error(request, 'Access denied. You do not have permission to delete build servers.')
        return redirect('build_server_detail', stream=stream, server_id=server_id)
    
    if request.method == 'POST':
        hostname = server.hostname
        server.delete()
        messages.success(request, f'Build server {hostname} deleted successfully!')
        return redirect('build_servers_list', stream=stream)
    
    context = {
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,
        'server': server,
    }
    
    return render(request, 'products/build_server_confirm_delete.html', context)

@login_required
@require_http_methods(["GET"])
def build_servers_api(request, stream=None):
    """
    API endpoint to get build servers data for AJAX requests
    """
    from .models import BuildServer
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check if user can access this stream
    if hasattr(request.user, 'custom_profile') and not request.user.custom_profile.can_access_stream(stream_obj.name):
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        servers = BuildServer.objects.filter(stream=stream_obj)
        
        # Apply filters
        status_filter = request.GET.get('status')
        if status_filter:
            servers = servers.filter(status=status_filter)
        
        stream_type_filter = request.GET.get('stream_type')
        if stream_type_filter:
            servers = servers.filter(stream_type=stream_type_filter)
        
        # Serialize data
        servers_data = []
        for server in servers:
            servers_data.append({
                'id': server.id,
                'hostname': server.hostname,
                'ip_address': server.ip_address,
                'location': server.location,
                'floor': server.floor,
                'owner': server.owner,
                'stream_type': server.stream_type,
                'status': server.status,
                'status_display': server.get_status_display(),
                'operating_system': server.operating_system,
                'cpu_cores': server.cpu_cores,
                'ram_gb': server.ram_gb,
                'storage_gb': server.storage_gb,
                'uptime_percentage': float(server.uptime_percentage),
                'warranty_expiring_soon': server.is_warranty_expiring_soon(),
                'days_until_warranty_expiry': server.days_until_warranty_expiry(),
                'url': reverse('build_server_detail', kwargs={'stream': stream, 'server_id': server.id})
            })
        
        return JsonResponse({
            'servers': servers_data,
            'total_count': len(servers_data)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def build_servers_export(request, stream=None):
    """
    Export build servers data to Excel
    """
    from .models import BuildServer
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    import io
    
    # Get the stream object
    stream_obj = get_stream_or_404(stream, default='HIC')
    
    # Check if user can access this stream
    if hasattr(request.user, 'custom_profile') and not request.user.custom_profile.can_access_stream(stream_obj.name):
        messages.error(request, f'Access denied. You do not have permission to export {stream_obj.name} build servers.')
        return redirect('build_servers_list', stream=stream)
    
    # Get servers
    servers = BuildServer.objects.filter(stream=stream_obj).order_by('hostname')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{stream_obj.name} Build Servers'
    
    # Headers
    headers = [
        'Hostname', 'IP Address', 'Location', 'Floor', 'Owner', 'Stream Type', 'Status',
        'Operating System', 'CPU Cores', 'RAM (GB)', 'Storage (GB)', 'MAC Address',
        'Domain', 'SSH Port', 'Purpose', 'Project Allocation', 'Primary Contact',
        'Secondary Contact', 'Contact Email', 'Procurement Date', 'Warranty Expiry',
        'Last Maintenance', 'Next Maintenance', 'Uptime %', 'Notes', 'Tags'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, server in enumerate(servers, 2):
        data = [
            server.hostname,
            server.ip_address,
            server.location,
            server.floor,
            server.owner,
            server.stream_type,
            server.status,
            server.operating_system or '',
            server.cpu_cores or '',
            server.ram_gb or '',
            server.storage_gb or '',
            server.mac_address or '',
            server.domain or '',
            server.ssh_port,
            server.purpose or '',
            server.project_allocation or '',
            server.primary_contact or '',
            server.secondary_contact or '',
            server.contact_email or '',
            server.procurement_date.strftime('%Y-%m-%d') if server.procurement_date else '',
            server.warranty_expiry.strftime('%Y-%m-%d') if server.warranty_expiry else '',
            server.last_maintenance.strftime('%Y-%m-%d') if server.last_maintenance else '',
            server.next_maintenance.strftime('%Y-%m-%d') if server.next_maintenance else '',
            float(server.uptime_percentage),
            server.notes or '',
            server.tags or ''
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{stream_obj.name}_build_servers_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


# Floor Management Views
@login_required
def floor_list(request, stream=None):
    """List all floors with management options for specific stream"""
    stream_obj = get_stream_or_404(stream)
    floors = Floor.objects.filter(stream=stream_obj).order_by('name')
    
    context = {
        'floors': floors,
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,
    }
    return render(request, 'products/floor_list.html', context)

@login_required
def floor_create(request, stream=None):
    """Create a new floor for specific stream"""
    stream_obj = get_stream_or_404(stream)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        if name:
            try:
                floor = Floor.objects.create(
                    name=name,
                    description=description,
                    stream=stream_obj,
                    is_active=is_active
                )
                messages.success(request, f'Floor "{name}" created successfully for {stream_obj.name} stream!')
                return redirect('floor_list', stream=stream)
            except IntegrityError:
                messages.error(request, f'Floor "{name}" already exists in {stream_obj.name} stream!')
        else:
            messages.error(request, 'Floor name is required!')
    
    context = {
        'is_edit': False,
        'title': 'Add New Floor',
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,
    }
    return render(request, 'products/floor_form.html', context)

@login_required
def floor_edit(request, stream=None, floor_id=None):
    """Edit an existing floor within specific stream"""
    stream_obj = get_stream_or_404(stream)
    floor = get_object_or_404(Floor, id=floor_id, stream=stream_obj)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        if name:
            try:
                floor.name = name
                floor.description = description
                floor.is_active = is_active
                floor.save()
                messages.success(request, f'Floor "{name}" updated successfully!')
                return redirect('floor_list', stream=stream)
            except IntegrityError:
                messages.error(request, f'Floor "{name}" already exists in {stream_obj.name} stream!')
        else:
            messages.error(request, 'Floor name is required!')
    
    context = {
        'floor': floor,
        'is_edit': True,
        'title': 'Edit Floor',
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,
    }
    return render(request, 'products/floor_form.html', context)

@login_required
def floor_delete(request, stream=None, floor_id=None):
    """Delete a floor within specific stream"""
    stream_obj = get_stream_or_404(stream)
    floor = get_object_or_404(Floor, id=floor_id, stream=stream_obj)
    
    if request.method == 'POST':
        # Check if any build servers are using this floor
        servers_using_floor = BuildServer.objects.filter(floor=floor).count()
        
        if servers_using_floor > 0:
            messages.error(request, f'Cannot delete floor "{floor.name}" as it is being used by {servers_using_floor} build server(s)!')
        else:
            floor_name = floor.name
            floor.delete()
            messages.success(request, f'Floor "{floor_name}" deleted successfully!')
        
        return redirect('floor_list', stream=stream)
    
    context = {
        'floor': floor,
        'servers_count': BuildServer.objects.filter(floor=floor).count(),
        'stream': stream,
        'stream_obj': stream_obj,
        'selected_stream': stream_obj,
    }
    return render(request, 'products/floor_confirm_delete.html', context)

