"""Products app - Holistic Weekly Data, Excel/PDF Export, and Bulk Update views."""

# pylint: disable=broad-exception-caught,protected-access

from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib.staticfiles import finders

from ..models import BusinessUnit as _BU  # noqa: N814  # pylint: disable=relative-beyond-top-level
from ._helpers import (
    A3,
    AuditLog,
    BytesIO,
    HolisticSystem,
    HolisticSystemHistory,
    HolisticWeeklyData,
    HttpResponse,
    JsonResponse,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    colors,
    date,
    datetime,
    get_column_letter,
    get_default_stream_name,
    get_stream_or_404,
    getSampleStyleSheet,
    landscape,
    logger,
    login_required,
    openpyxl,
    require_POST,
)

__all__ = [
    "holistic_weekly_data_update",
    "holistic_export_excel",
    "holistic_export_pdf",
    "holistic_bulk_update",
]


@login_required
@require_POST
def holistic_weekly_data_update(request, stream=None):  # pylint: disable=unused-argument
    """Update or create weekly data for a system."""
    try:
        system_id = request.POST.get("system_id")
        week_number = int(request.POST.get("week_number"))
        year = int(request.POST.get("year"))

        system = HolisticSystem.objects.get(id=system_id)

        weekly_data, _created = HolisticWeeklyData.objects.get_or_create(
            holistic_system=system, week_number=week_number, year=year, defaults={"updated_by": request.user}
        )

        weekly_data.allocation_status = request.POST.get("allocation_status", "")
        weekly_data.utilization_percentage = request.POST.get("utilization_percentage", 0)
        weekly_data.assigned_to = request.POST.get("assigned_to", "")
        weekly_data.task_description = request.POST.get("task_description", "")
        weekly_data.hours_used = request.POST.get("hours_used", 0)
        weekly_data.availability_hours = request.POST.get("availability_hours", 40)
        weekly_data.notes = request.POST.get("notes", "")
        weekly_data.updated_by = request.user
        weekly_data.save()

        AuditLog.log(
            action="update",
            title=f"Updated weekly data for system {system.sr_no} (W{week_number}/{year})",
            user=request.user,
            request=request,
            module="holistic",
            severity="info",
            stream=system.stream,
        )
        return JsonResponse(
            {
                "success": True,
                "message": f"Week {week_number} data updated successfully",
                "week_label": f"W{week_number}",
            }
        )

    except Exception:
        logger.exception("Operation failed")
        return JsonResponse({"success": False, "error": "An unexpected error occurred"}, status=400)


@login_required
def holistic_export_excel(request, stream=None):  # noqa: CCR001
    # pylint: disable=too-many-locals
    """Export holistic systems to Excel."""
    stream = stream or get_default_stream_name(request)
    stream_obj = get_stream_or_404(stream)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Holistic Systems"

    headers = [
        "Sr No",
        "System Availability",
        "Allocation to Sl No",
        "Location Info",
        "STMi Number",
        "System Owner",
        "ECR#",
        "Test Engineer",
        "Description",
        "Priority",
        "Notes",
    ]

    current_week = date.today().isocalendar()[1]
    for i in range(12):
        week_num = current_week - i
        if week_num < 1:
            week_num += 52
        headers.append(f"W{week_num}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="005fa3", end_color="005fa3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    systems = HolisticSystem.objects.filter(stream=stream_obj).order_by("sr_no")

    for row_num, system in enumerate(systems, 2):
        ws.cell(row=row_num, column=1, value=system.sr_no)
        ws.cell(row=row_num, column=2, value=system.get_system_availability_display())
        ws.cell(row=row_num, column=3, value=system.allocation_to_sl_no or "")
        ws.cell(row=row_num, column=4, value=system.location_info or "")
        ws.cell(row=row_num, column=5, value=system.stmi_number or "")
        ws.cell(row=row_num, column=6, value=system.system_owner or "")
        ws.cell(row=row_num, column=7, value=system.ecr_number or "")
        ws.cell(row=row_num, column=8, value=system.test_engineer or "")
        ws.cell(row=row_num, column=9, value=system.description or "")
        ws.cell(row=row_num, column=10, value=system.priority or "")
        ws.cell(row=row_num, column=11, value=system.notes or "")

        col_offset = 12
        for i in range(12):
            week_num = current_week - i
            if week_num < 1:
                week_num += 52

            week_data = system.weekly_data.filter(week_number=week_num, year=date.today().year).first()

            if week_data:
                value = f"{week_data.allocation_status or ''} ({week_data.utilization_percentage}%)"
            else:
                value = "-"

            ws.cell(row=row_num, column=col_offset + i, value=value)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_Holistic_Systems_{stream}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def holistic_export_pdf(request, stream=None):  # noqa: C901, CCR001
    # pylint: disable=too-many-locals,too-many-statements,too-complex
    """Export holistic systems to PDF."""
    stream = stream or get_default_stream_name(request)
    stream_obj = get_stream_or_404(stream)

    buffer = BytesIO()
    page_width, page_height = landscape(A3)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(page_width, page_height),
        leftMargin=30,
        rightMargin=30,
        topMargin=60,
        bottomMargin=30,
    )

    def truncate_text(text, max_length=30):
        """Aggressively truncate text to fit in table cells and handle None values."""
        if text is None:
            return "-"
        text = str(text).strip()
        if len(text) <= max_length:
            return text
        return text[: max_length - 3] + "..."

    def wrap_text(text, max_length=25, max_lines=2):  # noqa: CCR001
        """Wrap long text into multiple lines with strict limits."""
        if text is None or text == "-":
            return "-"
        text = str(text).strip()
        if len(text) <= max_length:
            return text

        words = text.split(" ")
        lines = []
        current_line = ""

        for word in words:
            if len(current_line + " " + word) <= max_length:
                current_line += " " + word if current_line else word
            else:
                if current_line:
                    lines.append(current_line)
                    if len(lines) >= max_lines:
                        break
                current_line = word[:max_length]

        if current_line and len(lines) < max_lines:
            lines.append(current_line)

        if len(lines) == max_lines and len(words) > len(" ".join(lines).split()):
            if len(lines) > 0:
                lines[-1] = lines[-1][: max_length - 3] + "..."

        return "\n".join(lines[:max_lines])

    def smart_truncate(text, max_length=20):
        """Smart truncation that preserves important parts."""
        if text is None:
            return "-"
        text = str(text).strip()
        if len(text) <= max_length:
            return text

        if "@" in text:
            parts = text.split("@")
            if len(parts) == 2:
                username_len = max(3, max_length - len(parts[1]) - 4)
                return parts[0][:username_len] + "@" + parts[1]

        return text[: max_length - 3] + "..."

    headers = [
        "Sr No",
        "Availability",
        "Allocation",
        "Location",
        "STMi#",
        "Owner",
        "ECR#",
        "Test Eng",
        "Priority",
        "Description",
        "Notes",
    ]

    data = [headers]

    systems = HolisticSystem.objects.filter(stream=stream_obj).order_by("sr_no")

    for system in systems:
        data.append(
            [
                truncate_text(system.sr_no, 10),
                truncate_text(system.get_system_availability_display(), 8),
                truncate_text(system.allocation_to_sl_no, 12),
                wrap_text(system.location_info, 15, 2),
                truncate_text(system.stmi_number, 10),
                smart_truncate(system.system_owner, 12),
                truncate_text(system.ecr_number, 10),
                smart_truncate(system.test_engineer, 12),
                truncate_text(system.priority, 8),
                wrap_text(system.description, 20, 2),
                wrap_text(system.notes, 20, 2),
            ]
        )

    col_widths = [50, 60, 65, 80, 60, 75, 60, 75, 50, 100, 100]

    table = Table(data, repeatRows=1, colWidths=col_widths)
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#005fa3")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 1), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("WORDWRAP", (0, 0), (-1, -1), "LTR"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ]
    )
    table.setStyle(style)

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = "Helvetica-Bold"
    title_style.fontSize = 22
    title_style.alignment = 1

    subtitle_style = styles["Heading2"]
    subtitle_style.fontName = "Helvetica-Bold"
    subtitle_style.fontSize = 18
    subtitle_style.alignment = 1

    subsubtitle_style = styles["Heading3"]
    subsubtitle_style.fontName = "Helvetica-Bold"
    subsubtitle_style.fontSize = 16
    subsubtitle_style.alignment = 1

    note_style = styles["Italic"]
    note_style.fontSize = 12
    note_style.alignment = 1

    logo_path = finders.find("products/Philips_NoBG.png")

    # Dynamic BU name for PDF header
    _bu_obj = getattr(request, "current_bu", None)
    if not _bu_obj:
        _bu_id = request.session.get("selected_bu_id")
        if _bu_id:
            try:
                _bu_obj = _BU.objects.get(id=_bu_id)
            except _BU.DoesNotExist:
                pass
    _pdf_bu_line1 = _bu_obj.name if _bu_obj else "NexusOps"
    _pdf_bu_line2 = f"{_bu_obj.bu_name} {_bu_obj.division}" if _bu_obj else ""

    def draw_header(canvas, doc):  # pylint: disable=unused-argument
        if logo_path:
            logo_width = 50
            logo_height = 50
            x_logo = page_width - logo_width - 40
            y_logo = page_height - logo_height - 20
            canvas.drawImage(logo_path, x_logo, y_logo, width=logo_width, height=logo_height, mask="auto")
        canvas.setFont("Helvetica-Bold", 22)
        canvas.drawCentredString(page_width / 2, page_height - 60, _pdf_bu_line1)
        canvas.setFont("Helvetica-Bold", 18)
        canvas.drawCentredString(page_width / 2, page_height - 90, _pdf_bu_line2)
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(page_width / 2, page_height - 120, "Holistic Systems Dashboard")
        canvas.setFont("Helvetica-Oblique", 12)
        canvas.drawCentredString(
            page_width / 2,
            page_height - 145,
            "(Note: Automated data output. Verification recommended to ensure reliability"
            " and compliance with organizational protocols.)",
        )

    elements = [Spacer(1, 120), table]

    # Weekly Data Table

    # Get last 4 weeks data
    current_week = date.today().isocalendar()[1]
    current_year = date.today().year

    weekly_headers = [
        "System Sr No",
        "Week",
        "Project",
        "Utilization %",
        "Assigned To",
        "Task Description",
        "Hours Used",
        "Notes",
    ]
    weekly_data = [weekly_headers]

    for i in range(4):
        week_num = current_week - i
        if week_num < 1:
            week_num += 52
            year = current_year - 1
        else:
            year = current_year

        week_data = HolisticWeeklyData.objects.filter(
            week_number=week_num, year=year, holistic_system__stream=stream_obj
        ).select_related("holistic_system", "project")

        for week in week_data:
            weekly_data.append(
                [
                    truncate_text(week.holistic_system.sr_no, 10),
                    f"W{week.week_number}",
                    truncate_text(week.project.name if week.project else "-", 15),
                    f"{week.utilization_percentage}%" if week.utilization_percentage else "0%",
                    smart_truncate(week.assigned_to, 12),
                    wrap_text(week.task_description, 18, 2),
                    f"{week.hours_used}h" if week.hours_used else "0h",
                    wrap_text(week.notes, 15, 2),
                ]
            )

    if len(weekly_data) > 1:
        weekly_table = Table(weekly_data, repeatRows=1)
        weekly_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#005fa3")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 1), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("WORDWRAP", (0, 0), (-1, -1), "LTR"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
            ]
        )
        weekly_table.setStyle(weekly_style)
        col_widths = [65, 40, 85, 50, 75, 120, 50, 90]
        weekly_table._argW = col_widths
        elements.append(Spacer(1, 40))
        elements.append(Paragraph("Weekly Assignment Data (Last 4 Weeks)", title_style))
        elements.append(Spacer(1, 10))
        elements.append(weekly_table)

    doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header)

    pdf = buffer.getvalue()
    buffer.close()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_Holistic_Systems_{stream}.pdf"

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(pdf)

    return response


@login_required
def holistic_bulk_update(request, stream=None):  # pylint: disable=unused-argument
    """Bulk update holistic systems."""
    if request.method == "POST":
        try:
            system_ids = request.POST.getlist("system_ids[]")
            action = request.POST.get("action")
            value = request.POST.get("value")

            systems = HolisticSystem.objects.filter(id__in=system_ids)

            updated_count = 0
            for system in systems:
                old_value = getattr(system, action)

                if action == "system_availability":
                    system.system_availability = value
                elif action == "priority":
                    system.priority = value

                system.updated_by = request.user
                system.save()

                HolisticSystemHistory.objects.create(
                    holistic_system=system,
                    action="bulk_updated",
                    user=request.user,
                    details=f"{action}: {old_value} → {value}",
                )

                updated_count += 1

            AuditLog.log(
                action="update",
                title=f"Bulk updated {updated_count} holistic systems ({action}: {value})",
                user=request.user,
                request=request,
                module="holistic",
                severity="info",
            )
            return JsonResponse({"success": True, "message": f"{updated_count} systems updated successfully"})

        except Exception:
            logger.exception("Operation failed")
            return JsonResponse({"success": False, "error": "An unexpected error occurred"}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)
