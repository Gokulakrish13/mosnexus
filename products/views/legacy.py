"""Products app — Legacy views."""

# pylint: disable=broad-exception-caught,import-error,invalid-name,reimported,relative-beyond-top-level,too-many-lines
# pylint: disable=undefined-variable,ungrouped-imports,wrong-import-position


from ._helpers import (
    ALLOWED_EXCEL_EXTENSIONS,
    ALLOWED_EXCEL_TYPES,
    MAX_EXCEL_SIZE,
    BytesIO,
    HttpResponse,
    JsonResponse,
    LegacyExcelUpload,
    Q,
    TestEnvironment,
    Workbook,
    ZenitionProduct,
    get_bu_streams,
    get_column_letter,
    get_current_bu,
    is_super_admin,
    json,
    logger,
    login_required,
    os,
    pd,
    re,
    render,
    require_GET,
    require_http_methods,
    require_POST,
    settings,
    timedelta,
    validate_uploaded_file,
)

__all__ = [
    "test_repo_view",
    "preview_legacy_excel",
    "preview_zenition_excel",
    "get_merge_pickle_path",
    "merge_excels",
    "preview_legacy_excel_by_filename",
    "fetch_legacy_excels",
    "delete_legacy_excel",
    "save_test_environment",
    "export_data",
    "fetch_zenition_excels",
]


@login_required
def test_repo_view(request):  # noqa: C901, CCR001
    """Test repo view."""
    # pylint: disable=too-many-locals,too-many-return-statements,too-many-branches
    # pylint: disable=too-many-statements,too-complex,invalid-name
    streams = list(get_bu_streams(request).values_list("name", flat=True).order_by("name"))
    bu = get_current_bu(request)

    def stream_sort_key(stream_name):
        if stream_name == "PIC":
            return (0, stream_name)
        if stream_name == "HIC":
            return (1, stream_name)
        return (2, stream_name)

    streams = sorted(set(streams), key=stream_sort_key)
    if bu:
        zenition_products = ZenitionProduct.objects.filter(
            Q(business_unit=bu) | Q(business_unit__isnull=True)
        ).order_by("name")
    else:
        zenition_products = ZenitionProduct.objects.all().order_by("name")
    # Handle legacy excel upload
    if request.method == "POST" and request.FILES.get("legacy_excel"):
        excel_file = request.FILES["legacy_excel"]
        is_valid, error_msg = validate_uploaded_file(
            excel_file, ALLOWED_EXCEL_TYPES, ALLOWED_EXCEL_EXTENSIONS, MAX_EXCEL_SIZE
        )
        if not is_valid:
            return JsonResponse({"success": False, "error": error_msg})
        selected_stream = request.POST.get("selected_stream", streams[0] if streams else "")
        upload = LegacyExcelUpload.objects.create(
            stream=selected_stream, business_unit=bu, file=excel_file, uploaded_by=request.user
        )
        try:
            df = pd.read_excel(upload.file.path)
            preview_html = df.head(20).to_html(index=False, escape=True)
            upload.preview_data = preview_html
            upload.save()
        except Exception:
            logger.exception("Operation failed")
            return JsonResponse({"success": False, "error": "An unexpected error occurred"})
        return JsonResponse({"success": True, "upload_id": upload.id})
    # Handle ZenitionProduct delete
    if request.method == "POST" and request.POST.get("zenition_delete_id"):
        delete_id = request.POST.get("zenition_delete_id")
        try:
            zp = ZenitionProduct.objects.get(id=delete_id)
            zp.delete()
        except ZenitionProduct.DoesNotExist:
            return JsonResponse({"success": False, "error": "Product not found"})
        if bu:
            zenition_products = ZenitionProduct.objects.filter(
                Q(business_unit=bu) | Q(business_unit__isnull=True)
            ).order_by("name")
        else:
            zenition_products = ZenitionProduct.objects.all().order_by("name")
        return JsonResponse(
            {"success": True, "zenition_products": [{"id": z.id, "name": z.name} for z in zenition_products]}
        )
    # Handle ZenitionProduct add
    if request.method == "POST" and request.POST.get("zenition_name"):
        name = request.POST.get("zenition_name").strip()
        if not name:
            return JsonResponse({"success": False, "error": "Name required"})
        if ZenitionProduct.objects.filter(name__iexact=name, business_unit=bu).exists():
            return JsonResponse({"success": False, "error": "Product already exists"})
        zp = ZenitionProduct.objects.create(name=name, business_unit=bu)
        if bu:
            zenition_products = ZenitionProduct.objects.filter(
                Q(business_unit=bu) | Q(business_unit__isnull=True)
            ).order_by("name")
        else:
            zenition_products = ZenitionProduct.objects.all().order_by("name")
        # Return updated list for dropdown
        return JsonResponse(
            {"success": True, "zenition_products": [{"id": z.id, "name": z.name} for z in zenition_products]}
        )
    # Handle ZenitionProduct Excel upload
    if request.method == "POST" and request.FILES.get("zenition_excel"):
        zenition_product_id = request.POST.get("zenition_product_id")
        selected_stream = request.POST.get("selected_stream", streams[0] if streams else "")
        try:
            zp = ZenitionProduct.objects.get(id=zenition_product_id)
        except ZenitionProduct.DoesNotExist:
            return JsonResponse({"success": False, "error": "Zenition Product not found"})
        excel_file = request.FILES["zenition_excel"]
        is_valid, error_msg = validate_uploaded_file(
            excel_file, ALLOWED_EXCEL_TYPES, ALLOWED_EXCEL_EXTENSIONS, MAX_EXCEL_SIZE
        )
        if not is_valid:
            return JsonResponse({"success": False, "error": error_msg})
        upload = LegacyExcelUpload(
            stream=selected_stream, business_unit=bu, zenition_product=zp, uploaded_by=request.user
        )
        upload.file = excel_file
        upload.save()
        try:
            df = pd.read_excel(upload.file.path)
            preview_html = df.head(20).to_html(index=False, escape=True)
            upload.preview_data = preview_html
            upload.save()
        except Exception:
            logger.exception("Operation failed")
            return JsonResponse({"success": False, "error": "An unexpected error occurred"})
        return JsonResponse({"success": True, "upload_id": upload.id})
    # Check if user can delete excel files (super_admin only via Feature Access Control)
    can_delete_excel = is_super_admin(request.user)
    return render(
        request,
        "products/test_repo.html",
        {
            "streams": streams,
            "selected_stream": streams[0] if streams else "",
            "zenition_products": zenition_products,
            "can_delete_excel": can_delete_excel,
        },
    )


@login_required
@require_GET
def preview_legacy_excel(request, upload_id):
    """Preview legacy excel."""
    try:
        excel = LegacyExcelUpload.objects.get(id=upload_id)
        file_path = excel.file.path
        df = pd.read_excel(file_path)
        preview_html = df.head(40).to_html(index=False, escape=True)
        return JsonResponse({"success": True, "html": preview_html})
    except LegacyExcelUpload.DoesNotExist:
        return JsonResponse({"success": False, "error": "File not found"})
    except Exception:
        logger.exception("Failed to preview legacy excel")
        return JsonResponse({"success": False, "error": "Failed to preview file"})


@login_required
@require_GET
def preview_zenition_excel(request, upload_id):
    """Preview zenition excel."""
    try:
        upload = LegacyExcelUpload.objects.get(id=upload_id)
        html = upload.preview_data or "<div>No preview available.</div>"
        return JsonResponse({"success": True, "html": html})
    except LegacyExcelUpload.DoesNotExist:
        return JsonResponse({"success": False, "error": "File not found"})


import pandas as pd  # noqa: E402, F811

from django.views.decorators.http import require_GET  # noqa: E402


def get_merge_pickle_path(legacy_id, zenition_id):
    """Return path for cached merged Excel data (JSON format, not pickle)."""
    # Validate IDs are integers to prevent path traversal
    try:
        legacy_id = int(legacy_id)
        zenition_id = int(zenition_id)
    except (ValueError, TypeError) as exc:
        raise ValueError("Invalid file IDs: must be integers") from exc
    temp_dir = os.path.join(settings.BASE_DIR, "products", "temp")
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
    return os.path.join(temp_dir, f"merged_{legacy_id}_{zenition_id}.json")


@login_required
@require_http_methods(["GET", "POST"])
def merge_excels(request):  # noqa: C901, CCR001
    """Merge excels."""
    # pylint: disable=too-many-locals,too-many-branches,too-many-statements,too-complex
    # Accept params from POST (state-changing) or GET (read-only preview)
    params = request.POST if request.method == "POST" else request.GET
    legacy_id = params.get("legacy_id")
    zenition_id = params.get("zenition_id")
    add_field = params.get("add_field")
    delete_field = params.get("delete_field")
    add_row_testcase = params.get("add_row_testcase")
    add_row_result = params.get("add_row_result")
    add_row_comment = params.get("add_row_comment")
    save_table_data = params.get("save_table_data")
    delete_row_index = params.get("delete_row_index")
    # Require POST for any state-changing operation
    is_mutating = any([add_field, delete_field, add_row_testcase, save_table_data, delete_row_index])
    if is_mutating and request.method != "POST":
        return JsonResponse({"success": False, "error": "State-changing operations require POST"}, status=405)
    if not legacy_id or not zenition_id:
        return JsonResponse({"success": False, "error": "Both file IDs required"})
    try:
        pickle_path = get_merge_pickle_path(legacy_id, zenition_id)
        # If cached JSON exists, load it; else, create from Excel files
        if os.path.exists(pickle_path):
            merged_df = pd.read_json(pickle_path, orient="records")
        else:
            legacy_upload = LegacyExcelUpload.objects.get(id=int(legacy_id))
            zenition_upload = LegacyExcelUpload.objects.get(id=int(zenition_id))
            df1 = pd.read_excel(legacy_upload.file.path)
            df2 = pd.read_excel(zenition_upload.file.path)
            merged_df = pd.concat([df1, df2], ignore_index=True)
        # Add field(s)
        if add_field:
            for field in add_field.split(","):
                field = field.strip()
                if field:
                    merged_df[field] = ""
        # Add new row if requested
        if add_row_testcase or add_row_result or add_row_comment:
            for col in ["BackEnd Test cases", "Results", "Comments"]:
                if col not in merged_df.columns:
                    merged_df[col] = ""
            new_row = {
                "BackEnd Test cases": add_row_testcase or "",
                "Results": add_row_result or "",
                "Comments": add_row_comment or "",
            }
            merged_df = pd.concat([merged_df, pd.DataFrame([new_row])], ignore_index=True)
        # Save edited table data
        if save_table_data:
            try:
                table_data = json.loads(save_table_data)
                for i, row in enumerate(table_data):
                    if i < len(merged_df):
                        merged_df.iloc[i, merged_df.columns.get_loc("BackEnd Test cases")] = row[0]
                        merged_df.iloc[i, merged_df.columns.get_loc("Results")] = row[1]
                        merged_df.iloc[i, merged_df.columns.get_loc("Comments")] = row[2]
            except Exception:
                logger.exception("Operation failed")
                return JsonResponse({"success": False, "error": "An unexpected error occurred"})
        if delete_row_index is not None:
            try:
                idx = int(delete_row_index)
                merged_df = merged_df.drop(idx).reset_index(drop=True)
            except Exception:
                logger.exception("Operation failed")
                return JsonResponse({"success": False, "error": "An unexpected error occurred"})
        if delete_field:
            for field in delete_field.split(","):
                field = field.strip()
                if field in merged_df.columns:
                    merged_df = merged_df.drop(columns=[field])
        # Save the updated DataFrame back to JSON (safe serialization)
        merged_df.to_json(pickle_path, orient="records", indent=2)
        preview_html = merged_df.head(40).to_html(index=False, escape=True)
        return JsonResponse({"success": True, "html": preview_html})
    except Exception:
        logger.exception("Error merging Excel files")
        return JsonResponse({"success": False, "error": "Failed to merge files"})


from django.conf import settings  # noqa: E402, F811
from django.http import JsonResponse  # noqa: E402, F811

from ..models import LegacyExcelUpload  # noqa: E402, F811


@login_required
@require_GET
def preview_legacy_excel_by_filename(request, filename):
    """Preview legacy excel by filename."""
    folder = os.path.join(settings.BASE_DIR, "product_images", "legacy_excels")
    # Security: prevent path traversal — only use the base filename
    safe_filename = os.path.basename(filename)
    file_path = os.path.join(folder, safe_filename)
    # Verify the resolved path is still inside the intended directory
    if not os.path.realpath(file_path).startswith(os.path.realpath(folder)):
        return JsonResponse({"success": False, "error": "Invalid filename"}, status=400)
    if not os.path.exists(file_path):
        return JsonResponse({"success": False, "error": "File not found"})
    try:
        df = pd.read_excel(file_path)
        preview_html = df.head(40).to_html(index=False, escape=True)
        return JsonResponse({"success": True, "html": preview_html})
    except Exception:
        logger.exception("Error previewing legacy Excel file")
        return JsonResponse({"success": False, "error": "Failed to preview file"})


@login_required
@require_GET
def fetch_legacy_excels(request):
    """Fetch legacy excels."""
    ist = dt_tz(timedelta(hours=5, minutes=30))  # type: ignore[name-defined]  # noqa: F821
    # Return LegacyExcelUpload objects scoped to current BU
    bu = get_current_bu(request)
    if bu:
        excels = (
            LegacyExcelUpload.objects.filter(zenition_product__isnull=True)
            .filter(Q(business_unit=bu) | Q(business_unit__isnull=True))
            .order_by("-uploaded_at")
        )
    else:
        excels = LegacyExcelUpload.objects.filter(zenition_product__isnull=True).order_by("-uploaded_at")
    # Deduplicate: keep only the latest upload per base filename
    seen_bases = {}
    for excel in excels:
        raw_name = excel.file.name.split("/")[-1]
        # Strip Django's random suffix (e.g. _EREAuQm) before extension
        base = re.sub(r"_[A-Za-z0-9]{7}(\.[^.]+)$", r"\1", raw_name)
        if base not in seen_bases:
            seen_bases[base] = {
                "file_name": raw_name,
                "base_name": base,
                "id": excel.id,
                "uploaded_at": (
                    excel.uploaded_at.astimezone(ist).strftime("%d-%m-%Y %I:%M %p") if excel.uploaded_at else ""
                ),
                "uploaded_by": str(excel.uploaded_by) if excel.uploaded_by else "",
            }
    return JsonResponse({"success": True, "excels": list(seen_bases.values())})


@login_required
@require_POST
def delete_legacy_excel(request, upload_id):
    """Delete a legacy excel upload and all duplicates with the same base filename."""
    if not is_super_admin(request.user):
        return JsonResponse({"success": False, "error": "Only Super Admins can delete files"}, status=403)
    try:
        upload = LegacyExcelUpload.objects.get(id=upload_id)
    except LegacyExcelUpload.DoesNotExist:
        return JsonResponse({"success": False, "error": "File not found"})
    # Determine base filename (strip Django's random suffix)
    raw_name = upload.file.name.split("/")[-1]
    base = re.sub(r"_[A-Za-z0-9]{7}(\.[^.]+)$", r"\1", raw_name)
    # Find all uploads with the same base filename, same type (legacy vs zenition)
    bu = get_current_bu(request)
    zen_product = upload.zenition_product
    qs = LegacyExcelUpload.objects.filter(zenition_product=zen_product)
    if bu:
        all_uploads = qs.filter(Q(business_unit=bu) | Q(business_unit__isnull=True))
    else:
        all_uploads = qs
    deleted_count = 0
    for upl in all_uploads:
        upl_raw = upl.file.name.split("/")[-1]
        upl_base = re.sub(r"_[A-Za-z0-9]{7}(\.[^.]+)$", r"\1", upl_raw)
        if upl_base == base:
            if upl.file and os.path.isfile(upl.file.path):
                os.remove(upl.file.path)
            upl.delete()
            deleted_count += 1
    return JsonResponse({"success": True, "deleted_count": deleted_count})


from ..models import TestEnvironment  # noqa: E402, F811


@login_required
def save_test_environment(request):
    """Save test environment."""
    if request.method == "POST":
        data = request.POST
        env = TestEnvironment.objects.create(
            mvs_binaries=data.get("mvs_binaries", ""),
            mvs_os=data.get("mvs_os", ""),
            stand_binaries=data.get("stand_binaries", ""),
            stand_os=data.get("stand_os", ""),
            apps_pc_binaries=data.get("apps_pc_binaries", ""),
            apps_pc_os=data.get("apps_pc_os", ""),
            test_environment=data.get("test_environment", ""),
        )
        return JsonResponse({"success": True, "id": env.id})
    return JsonResponse({"success": False, "error": "Invalid request method."}, status=400)


from io import BytesIO  # noqa: E402, F811

from openpyxl import Workbook  # noqa: E402, F811
from openpyxl.utils import get_column_letter  # noqa: E402, F811

from django.http import HttpResponse  # noqa: E402, F811


@login_required
def export_data(request):
    """Export data."""
    # pylint: disable=too-many-locals
    legacy_id = request.GET.get("legacy_id")
    zenition_id = request.GET.get("zenition_id")
    pickle_path = get_merge_pickle_path(legacy_id, zenition_id)
    if os.path.exists(pickle_path):
        merged_df = pd.read_json(pickle_path, orient="records")
    else:
        merged_df = pd.DataFrame()
    test_env = TestEnvironment.objects.last()
    zenition_product_id = request.GET.get("zenition_id")
    try:
        zenition_product = ZenitionProduct.objects.get(id=zenition_product_id)
        zenition_name = zenition_product.name
    except ZenitionProduct.DoesNotExist:
        zenition_name = "TestRepoZenitionProduct"

    wb = Workbook()
    ws = wb.active
    ws.title = "Exported Data"

    # Row 1: TestRepoZenitionProduct name
    ws.append([zenition_name])
    # Row 2: Empty
    ws.append([])
    # Rows 3-8: Test Environment key-value pairs
    env_pairs = [
        ("MVS Binaries", getattr(test_env, "mvs_binaries", "")),
        ("MVS OS", getattr(test_env, "mvs_os", "")),
        ("Stand Binaries", getattr(test_env, "stand_binaries", "")),
        ("Stand OS", getattr(test_env, "stand_os", "")),
        ("APPS PC Binaries", getattr(test_env, "apps_pc_binaries", "")),
        ("APPS PC OS", getattr(test_env, "apps_pc_os", "")),
    ]
    for key, value in env_pairs:
        ws.append([key, value])
    # Row 9: Empty
    ws.append([])
    # Row 10+: Merged Excel Preview
    if not merged_df.empty:
        ws.append(list(merged_df.columns))
        for row in merged_df.itertuples(index=False):
            ws.append(list(row))
    else:
        ws.append(["No merged data available"])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=exported_data.xlsx"
    return response


@login_required
@require_GET
def fetch_zenition_excels(request):  # noqa: CCR001
    """Fetch zenition excels."""
    ist = dt_tz(timedelta(hours=5, minutes=30))  # type: ignore[name-defined]  # noqa: F821
    bu = get_current_bu(request)
    if bu:
        zenition_excels = (
            LegacyExcelUpload.objects.filter(zenition_product__isnull=False)
            .filter(Q(business_unit=bu) | Q(business_unit__isnull=True))
            .select_related("zenition_product")
            .order_by("-uploaded_at")
        )
    else:
        zenition_excels = (
            LegacyExcelUpload.objects.filter(zenition_product__isnull=False)
            .select_related("zenition_product")
            .order_by("-uploaded_at")
        )
    # Group by zenition product, then deduplicate by base filename within each product
    products_dict: dict[str, dict] = {}  # product_name -> { base_name -> file_info }
    for excel in zenition_excels:
        product_name = excel.zenition_product.name
        raw_name = excel.file.name.split("/")[-1]
        base = re.sub(r"_[A-Za-z0-9]{7}(\.[^.]+)$", r"\1", raw_name)
        if product_name not in products_dict:
            products_dict[product_name] = {}
        if base not in products_dict[product_name]:
            products_dict[product_name][base] = {
                "file_name": raw_name,
                "base_name": base,
                "id": excel.id,
                "product_name": product_name,
                "uploaded_at": (
                    excel.uploaded_at.astimezone(ist).strftime("%d-%m-%Y %I:%M %p") if excel.uploaded_at else ""
                ),
                "uploaded_by": str(excel.uploaded_by) if excel.uploaded_by else "",
            }
    # Flatten: list of { product_name, files: [...] }
    folders = []
    for pname, files in sorted(products_dict.items()):
        folders.append({"product_name": pname, "files": list(files.values())})
    return JsonResponse({"success": True, "folders": folders})


from django.contrib.auth.decorators import login_required  # noqa: E402
from django.http import JsonResponse  # noqa: E402, F811
from django.shortcuts import render  # noqa: E402, F811
