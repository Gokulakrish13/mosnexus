"""Products app - Resource Import, Heatmap, Export, Lookup, Config, Components, Year views."""

# pylint: disable=broad-exception-caught,chained-comparison,import-error,invalid-name,logging-too-many-args
# pylint: disable=no-else-return,relative-beyond-top-level,too-many-lines

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..models import (
    ResourceAllocation,
    ResourceAllocationYear,
    ResourceComponentType,
    ResourceLocation,
    ResourceManager,
    ResourcePerson,
    ResourceRole,
)
from ._helpers import (
    AuditLog,
    HttpResponse,
    JsonResponse,
    Project,
    Workbook,
    date,
    get_bu_streams,
    get_column_letter,
    is_admin,
    is_app_admin,
    is_super_admin,
    json,
    logger,
    login_required,
    messages,
    openpyxl,
    redirect,
    render,
    transaction,
)

__all__ = [
    "resource_allocation_import_api",
    "resource_allocation_heatmap_api",
    "resource_allocation_export",
    "resource_person_assign_api",
    "resource_lookup_api",
    "resource_allocation_config",
    "resource_component_api",
    "resource_year_api",
]


@login_required
def resource_allocation_import_api(request):  # noqa: C901, CCR001
    """POST → upload Excel file and import allocations."""
    # pylint: disable=too-complex,too-many-branches,too-many-locals,too-many-return-statements,too-many-statements
    bu_streams = get_bu_streams(request)
    if not bu_streams.exists():
        return JsonResponse({"success": False, "error": "No streams"}, status=400)
    stream_obj = bu_streams.first()

    if request.method != "POST":
        return JsonResponse({"success": False, "error": "POST only"}, status=405)

    try:
        f = request.FILES.get("file")
        if not f:
            return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

        year = int(request.POST.get("year", date.today().year))
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        headers = [str(c.value or "").strip().lower() for c in ws[1]]

        def find_col(names):
            for n in names:
                for i, h in enumerate(headers):
                    if n in h:
                        return i
            return -1

        col_name = find_col(["name"])
        col_empid = find_col(["emp id", "emp_id", "employee id"])
        col_comp = find_col(["component"])
        col_type = find_col(["employee type", "fte type", "emp type"])
        col_proj = find_col(["project"])

        month_cols = {}
        mn_list = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
        for m_idx, mn in enumerate(mn_list, 1):
            for i, h in enumerate(headers):
                if mn in h:
                    month_cols[m_idx] = i
                    break

        if col_name < 0:
            return JsonResponse({"success": False, "error": 'Missing "Name" column'}, status=400)
        if col_proj < 0:
            return JsonResponse({"success": False, "error": 'Missing "Project" column'}, status=400)
        if not month_cols:
            return JsonResponse({"success": False, "error": "Missing month columns (Jan-Dec)"}, status=400)

        person_cache = {}
        for p in ResourcePerson.objects.filter(stream__in=bu_streams, is_active=True):
            person_cache[p.name.strip().lower()] = p
            if p.emp_id:
                person_cache[p.emp_id.strip().lower()] = p

        project_cache = {pr.name.strip().lower(): pr for pr in Project.objects.filter(stream__in=bu_streams)}
        comp_cache = {
            c.name.strip().lower(): c
            for c in ResourceComponentType.objects.filter(stream__in=bu_streams, is_active=True)
        }

        imported = skipped = created_persons = 0
        errors = []

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row = list(row)
                name = str(row[col_name] or "").strip()
                if not name:
                    continue

                person = None
                if col_empid >= 0 and len(row) > col_empid and row[col_empid]:
                    person = person_cache.get(str(row[col_empid]).strip().lower())
                if not person:
                    person = person_cache.get(name.lower())

                if not person:
                    comp = None
                    fte_type = "FTE"
                    if col_comp >= 0 and len(row) > col_comp and row[col_comp]:
                        comp = comp_cache.get(str(row[col_comp]).strip().lower())
                    if col_type >= 0 and len(row) > col_type and row[col_type]:
                        fte_type = str(row[col_type]).strip()
                    eid = (
                        str(row[col_empid]).strip()
                        if col_empid >= 0 and len(row) > col_empid and row[col_empid]
                        else ""
                    )
                    person = ResourcePerson.objects.create(
                        stream=stream_obj,
                        name=name,
                        emp_id=eid,
                        fte_type=fte_type,
                        component=comp,
                        is_active=True,
                        show_in_allocation=True,
                        created_by=request.user,
                    )
                    person_cache[name.lower()] = person
                    if eid:
                        person_cache[eid.lower()] = person
                    created_persons += 1

                proj_name = str(row[col_proj] or "").strip()
                if not proj_name:
                    skipped += 1
                    continue
                project = project_cache.get(proj_name.lower())
                if not project:
                    errors.append(f"Row {row_num}: Project '{proj_name}' not found")
                    skipped += 1
                    continue

                if not person.show_in_allocation:
                    person.show_in_allocation = True
                    person.save(update_fields=["show_in_allocation"])

                for m, ci in month_cols.items():
                    val = row[ci] if ci < len(row) else None
                    try:
                        alloc_val = float(val) if val is not None else 0
                    except (ValueError, TypeError):
                        alloc_val = 0
                    ResourceAllocation.objects.update_or_create(
                        person=person,
                        project=project,
                        year=year,
                        month=m,
                        defaults={"allocation": alloc_val, "created_by": request.user},
                    )
                    imported += 1

        msg = f"Import complete: {imported} cells imported"
        if created_persons:
            msg += f", {created_persons} new persons auto-created"
        if skipped:
            msg += f", {skipped} rows skipped"

        return JsonResponse(
            {
                "success": True,
                "message": msg,
                "imported": imported,
                "created_persons": created_persons,
                "skipped": skipped,
                "errors": errors[:20],
            }
        )
    except Exception as exc:
        logger.exception("resource_allocation_import error: %s", exc)
        return JsonResponse({"success": False, "error": str(exc)}, status=500)


# ── Heatmap API ───────────────────────────────────────────────────────────────


@login_required
def resource_allocation_heatmap_api(request):  # noqa: CCR001
    """GET → demand vs supply heatmap data by component and month."""
    # pylint: disable=too-many-locals
    bu_streams = get_bu_streams(request)
    if not bu_streams.exists():
        return JsonResponse({"success": False, "error": "No streams"}, status=400)

    year = int(request.GET.get("year", date.today().year))

    components = ResourceComponentType.objects.filter(stream__in=bu_streams, is_active=True).order_by(
        "sort_order", "name"
    )
    persons = ResourcePerson.objects.filter(stream__in=bu_streams, is_active=True).select_related("component")
    allocations = ResourceAllocation.objects.filter(person__stream__in=bu_streams, year=year).select_related("person")

    supply: dict[str, int] = {}
    for p in persons:
        cn = p.component.name if p.component else "Unassigned"
        supply[cn] = supply.get(cn, 0) + 1

    demand: dict[str, dict] = {}
    for a in allocations:
        cn = a.person.component.name if a.person.component else "Unassigned"
        if cn not in demand:
            demand[cn] = {}
        demand[cn][a.month] = demand[cn].get(a.month, 0) + float(a.allocation)

    comp_names = [c.name for c in components]
    if "Unassigned" in supply or "Unassigned" in demand:
        comp_names.append("Unassigned")

    heatmap = []
    for cn in comp_names:
        row = {"component": cn, "supply": supply.get(cn, 0), "months": {}}
        for m in range(1, 13):
            d = round(demand.get(cn, {}).get(m, 0), 2)
            s = supply.get(cn, 0)
            row["months"][str(m)] = {"demand": d, "supply": s, "gap": round(s - d, 2)}
        heatmap.append(row)

    return JsonResponse({"success": True, "year": year, "heatmap": heatmap})


@login_required
def resource_allocation_export(request):  # noqa: CCR001
    """Export resource allocation grid as Excel."""
    # pylint: disable=too-many-locals
    bu_streams = get_bu_streams(request)
    year = int(request.GET.get("year", date.today().year))

    persons = (
        ResourcePerson.objects.filter(stream__in=bu_streams, is_active=True)
        .select_related("component")
        .order_by("name")
    )

    allocations = ResourceAllocation.objects.filter(
        person__stream__in=bu_streams,
        year=year,
    ).select_related("person", "project")

    alloc_map = {}
    for a in allocations:
        alloc_map[(a.person_id, a.project_id, a.month)] = float(a.allocation)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Resource Allocation {year}"

    # Header row
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers = ["Component", "FTE Type", "Name", "Emp ID", "Manager", "Location", "Role", "Project"]
    for mn in month_names:
        headers.append(f"{str(year)[2:]}-{mn}")
    ws.append(headers)

    # Bold header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_num = 2
    for p in persons:
        person_projects = (
            ResourceAllocation.objects.filter(person=p, year=year).values_list("project_id", flat=True).distinct()
        )

        if not person_projects:
            comp_name = p.component.name if p.component else ""
            row = [comp_name, p.fte_type, p.name, p.emp_id, p.manager, p.location, p.role, ""]
            for _m in range(1, 13):
                row.append("-")
            ws.append(row)
            row_num += 1
        else:
            projects = Project.objects.filter(id__in=person_projects)
            for proj in projects:
                comp_name = p.component.name if p.component else ""
                row = [comp_name, p.fte_type, p.name, p.emp_id, p.manager, p.location, p.role, proj.name]
                for m in range(1, 13):
                    val = alloc_map.get((p.id, proj.id, m), 0)
                    row.append(f" {val:.1f} " if val else " -   ")
                ws.append(row)
                row_num += 1

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Resource_Allocation_{year}.xlsx"'
    wb.save(response)
    return response


@login_required
def resource_person_assign_api(request):
    """POST → assign a person to the allocation grid (set show_in_allocation=True).

    and create placeholder allocation records for the given year so the person
    appears in that year's grid immediately.
    """
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "POST required"}, status=405)

    bu_streams = get_bu_streams(request)
    if not bu_streams.exists():
        return JsonResponse({"success": False, "error": "No streams"}, status=400)

    try:
        data = json.loads(request.body)
        person_id = data.get("person_id")
        assign_year = int(data.get("year", date.today().year))
        person = ResourcePerson.objects.get(id=person_id, stream__in=bu_streams, is_active=True)
        person.show_in_allocation = True
        person.save()

        # We no longer auto-create placeholder allocation records.
        # The person will appear in the grid as an "unassigned" row
        # (the GET handler shows persons that have show_in_allocation=True
        #  AND have allocation records for the year, OR are newly added).
        # To make the person visible immediately, we mark them with a
        # year-level flag via a single allocation record with a null-like
        # sentinel. Instead, let the GET handler also include persons
        # that have show_in_allocation=True even if they have no records
        # for the selected year — as long as they were just assigned.
        # Simplest: always create 12 placeholder records with first project.
        has_allocs = ResourceAllocation.objects.filter(person=person, year=assign_year).exists()
        if not has_allocs:
            first_project = Project.objects.filter(stream__in=bu_streams).first()
            if first_project:
                for m in range(1, 13):
                    ResourceAllocation.objects.get_or_create(
                        person=person,
                        project=first_project,
                        year=assign_year,
                        month=m,
                        defaults={"allocation": 0, "created_by": request.user},
                    )
        AuditLog.log(
            action="update",
            title=f"Added {person.name} to allocation grid",
            user=request.user,
            request=request,
            obj=person,
            module="projects",
            severity="info",
            stream=bu_streams.first(),
        )
        return JsonResponse({"success": True, "message": f"{person.name} added to allocation grid"})
    except ResourcePerson.DoesNotExist:
        return JsonResponse({"success": False, "error": "Person not found or inactive"}, status=404)
    except Exception as exc:
        logger.exception("resource_person_assign error: %s", exc)
        return JsonResponse({"success": False, "error": str(exc)}, status=500)


@login_required
def resource_lookup_api(request, lookup_type):  # noqa: C901, CCR001
    """Generic CRUD API for resource lookup tables: manager, location, role.

    URL: /projects/resource-allocation/lookup/<type>/
    GET  → list items
    POST → create or update
    DELETE → deactivate
    """
    # pylint: disable=too-complex,too-many-branches,too-many-return-statements
    model_map = {
        "manager": ResourceManager,
        "location": ResourceLocation,
        "role": ResourceRole,
    }

    model_cls = model_map.get(lookup_type)
    if not model_cls:
        return JsonResponse({"success": False, "error": "Invalid lookup type"}, status=400)

    bu_streams = get_bu_streams(request)
    if not bu_streams.exists():
        return JsonResponse({"success": False, "error": "No streams available"}, status=400)
    stream_obj = bu_streams.first()

    if request.method == "GET":
        items = model_cls.objects.filter(stream__in=bu_streams, is_active=True).order_by("sort_order", "name")
        data = [{"id": i.id, "name": i.name, "sort_order": i.sort_order} for i in items]
        return JsonResponse({"success": True, "items": data})

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            item_id = data.get("id")
            name = (data.get("name") or "").strip()
            if not name:
                return JsonResponse({"success": False, "error": "Name is required"}, status=400)

            sort_order = int(data.get("sort_order", 0))

            if item_id:
                item = model_cls.objects.get(id=item_id, stream__in=bu_streams)
                item.name = name
                item.sort_order = sort_order
                item.save()
                return JsonResponse({"success": True, "id": item.id, "message": f"{lookup_type.title()} updated"})
            else:
                item, created = model_cls.objects.get_or_create(
                    stream=stream_obj, name=name, defaults={"sort_order": sort_order}
                )
                if not created:
                    if not item.is_active:
                        item.is_active = True
                        item.sort_order = sort_order
                        item.save()
                        return JsonResponse(
                            {"success": True, "id": item.id, "message": f"{lookup_type.title()} re-activated"}
                        )
                    return JsonResponse({"success": False, "error": f'"{name}" already exists'}, status=400)
                return JsonResponse({"success": True, "id": item.id, "message": f"{lookup_type.title()} created"})
        except model_cls.DoesNotExist:
            return JsonResponse({"success": False, "error": "Item not found"}, status=404)
        except Exception as exc:
            logger.exception("resource_lookup_api error: %s", exc)
            return JsonResponse({"success": False, "error": str(exc)}, status=500)

    if request.method == "DELETE":
        try:
            data = json.loads(request.body)
            item = model_cls.objects.get(id=data.get("id"), stream__in=bu_streams)
            item.is_active = False
            item.save()
            return JsonResponse({"success": True, "message": f"{lookup_type.title()} removed"})
        except model_cls.DoesNotExist:
            return JsonResponse({"success": False, "error": "Item not found"}, status=404)
        except Exception as exc:
            logger.exception("resource_lookup_api delete error: %s", exc)
            return JsonResponse({"success": False, "error": str(exc)}, status=500)

    return JsonResponse({"success": False, "error": "Invalid method"}, status=405)


@login_required
def resource_allocation_config(request):
    """Configuration page for Resource Allocation — manage component types and persons.

    Only accessible by app admins.
    """
    # Permission check
    if not (is_admin(request.user) or is_super_admin(request.user) or is_app_admin(request.user)):
        messages.error(request, "Permission denied. Only administrators can access this page.")
        return redirect("dashboard")

    bu_streams = get_bu_streams(request)
    if not bu_streams.exists():
        messages.error(request, "No streams available.")
        return redirect("dashboard")
    stream_obj = bu_streams.first()

    components = ResourceComponentType.objects.filter(stream__in=bu_streams).order_by("sort_order", "name")
    persons = (
        ResourcePerson.objects.filter(stream__in=bu_streams, is_active=True)
        .select_related("component")
        .order_by("name")
    )
    years = (
        ResourceAllocationYear.objects.filter(stream__in=bu_streams, is_active=True)
        .order_by("year")
        .values_list("year", flat=True)
        .distinct()
    )
    managers = ResourceManager.objects.filter(stream__in=bu_streams, is_active=True).order_by("sort_order", "name")
    locations = ResourceLocation.objects.filter(stream__in=bu_streams, is_active=True).order_by("sort_order", "name")
    roles = ResourceRole.objects.filter(stream__in=bu_streams, is_active=True).order_by("sort_order", "name")

    context = {
        "selected_stream": stream_obj.name,
        "stream": stream_obj,
        "components": components,
        "persons": persons,
        "years": years,
        "managers": managers,
        "locations": locations,
        "roles": roles,
        "all_streams": bu_streams.order_by("name"),
    }
    return render(request, "products/resource_allocation_config.html", context)


@login_required
def resource_component_api(request):  # noqa: C901, CCR001
    """CRUD API for ResourceComponentType (configurable component/user types like FE, CBE)."""
    # pylint: disable=too-complex,too-many-return-statements
    bu_streams = get_bu_streams(request)
    if not bu_streams.exists():
        return JsonResponse({"success": False, "error": "No streams available"}, status=400)
    stream_obj = bu_streams.first()

    if request.method == "GET":
        components = ResourceComponentType.objects.filter(stream__in=bu_streams, is_active=True).order_by(
            "sort_order", "name"
        )
        data = [
            {"id": c.id, "name": c.name, "description": c.description, "sort_order": c.sort_order} for c in components
        ]
        return JsonResponse({"success": True, "components": data})

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            comp_id = data.get("id")
            if comp_id:
                comp = ResourceComponentType.objects.get(id=comp_id, stream__in=bu_streams)
                comp.name = data.get("name", comp.name).strip()
                comp.description = data.get("description", comp.description)
                comp.sort_order = int(data.get("sort_order", comp.sort_order))
                comp.save()
                return JsonResponse({"success": True, "id": comp.id, "message": f'Component "{comp.name}" updated'})
            else:
                comp = ResourceComponentType.objects.create(
                    stream=stream_obj,
                    name=data.get("name", "").strip(),
                    description=data.get("description", ""),
                    sort_order=int(data.get("sort_order", 0)),
                )
                return JsonResponse({"success": True, "id": comp.id, "message": f'Component "{comp.name}" created'})
        except Exception:
            logger.exception("resource_component_api error")
            return JsonResponse({"success": False, "error": "An error occurred"}, status=500)

    if request.method == "DELETE":
        try:
            data = json.loads(request.body)
            comp_id = data.get("id")
            comp = ResourceComponentType.objects.get(id=comp_id, stream__in=bu_streams)
            comp_name = comp.name
            comp.is_active = False
            comp.save()
            return JsonResponse({"success": True, "message": f'Component "{comp_name}" removed'})
        except ResourceComponentType.DoesNotExist:
            return JsonResponse({"success": False, "error": "Component not found"}, status=404)
        except Exception:
            logger.exception("resource_component_api error")
            return JsonResponse({"success": False, "error": "An error occurred"}, status=500)

    return JsonResponse({"success": False, "error": "Invalid method"}, status=405)


@login_required
def resource_year_api(request):  # noqa: C901, CCR001
    """GET  → list all configured years for the stream.

    POST → add a new year
    DELETE → remove a year
    """
    # pylint: disable=too-complex,too-many-return-statements
    bu_streams = get_bu_streams(request)
    if not bu_streams.exists():
        return JsonResponse({"success": False, "error": "No streams available"}, status=400)
    _stream_obj = bu_streams.first()  # noqa: F841

    if request.method == "GET":
        years = sorted(
            set(
                ResourceAllocationYear.objects.filter(stream__in=bu_streams, is_active=True).values_list(
                    "year", flat=True
                )
            )
        )
        data = [{"year": y} for y in years]
        return JsonResponse({"success": True, "years": data})

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            year_val = int(data.get("year", 0))
            if year_val < 2000 or year_val > 2100:
                return JsonResponse({"success": False, "error": "Year must be between 2000 and 2100"}, status=400)
            # Create for all streams in this BU
            created_any = False
            for s in bu_streams:
                obj, created = ResourceAllocationYear.objects.get_or_create(
                    stream=s, year=year_val, defaults={"is_active": True}
                )
                if not created and not obj.is_active:
                    obj.is_active = True
                    obj.save()
                if created:
                    created_any = True
            return JsonResponse(
                {
                    "success": True,
                    "year": year_val,
                    "message": f"Year {year_val} added" if created_any else f"Year {year_val} already exists",
                }
            )
        except Exception:
            logger.exception("resource_year_api error")
            return JsonResponse({"success": False, "error": "An error occurred"}, status=500)

    if request.method == "DELETE":
        try:
            data = json.loads(request.body)
            year_val = int(data.get("year", 0))
            updated = ResourceAllocationYear.objects.filter(
                stream__in=bu_streams, year=year_val, is_active=True
            ).update(is_active=False)
            if updated:
                return JsonResponse({"success": True, "message": f"Year {year_val} removed"})
            else:
                return JsonResponse({"success": False, "error": "Year not found"}, status=404)
        except Exception:
            logger.exception("resource_year_api error")
            return JsonResponse({"success": False, "error": "An error occurred"}, status=500)

    return JsonResponse({"success": False, "error": "Invalid method"}, status=405)


# =============================================
# HOLISTIC DASHBOARD VIEWS
# =============================================
