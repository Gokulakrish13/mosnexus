"""Products app - Build Servers, Floor, and OS Management views."""

# pylint: disable=too-many-lines,broad-exception-caught

from openpyxl.styles import Alignment, Font, PatternFill

from ._helpers import (
    AuditLog,
    BuildServer,
    BuildServerHistory,
    BuildServerMaintenanceLog,
    Floor,
    HttpResponse,
    JsonResponse,
    OperatingSystem,
    Q,
    can_delete_products,
    can_edit_products,
    datetime,
    get_bu_streams,
    get_object_or_404,
    get_stream_or_404,
    io,
    logger,
    login_required,
    messages,
    openpyxl,
    redirect,
    render,
    require_http_methods,
    reverse,
    timezone,
)
from ..approval_triggers import check_approval_required, fire_approval_trigger

__all__ = [
    "build_servers_dashboard",
    "reservations_hub",
    "calibration_hub",
    "compliance_hub",
    "build_servers_list",
    "build_server_create",
    "build_server_detail",
    "build_server_edit",
    "build_server_delete",
    "build_servers_api",
    "build_servers_export",
]


@login_required
def build_servers_dashboard(request):
    """Main Build Servers dashboard with stream selection."""
    streams = get_bu_streams(request).order_by("name")

    selected_stream = streams.first() if streams.exists() else None

    context = {
        "streams": streams,
        "selected_stream": selected_stream,
        "stream": selected_stream,  # For base template compatibility
    }

    return render(request, "products/build_servers_dashboard.html", context)


# ================================
# HUB PAGES (Stream Selection)
# ================================


@login_required
def reservations_hub(request):
    """Reservations Hub - Select stream to access recurring reservations, waitlist, or utilization."""
    streams = get_bu_streams(request).order_by("name")

    context = {
        "streams": streams,
    }

    return render(request, "products/reservations_hub.html", context)


@login_required
def calibration_hub(request):
    """Calibration Hub - Select stream to access calibration dashboard, schedules, or records."""
    streams = get_bu_streams(request).order_by("name")

    context = {
        "streams": streams,
    }

    return render(request, "products/calibration_hub.html", context)


@login_required
def compliance_hub(request):
    """Compliance Hub - Select stream to access compliance dashboard, documents, alerts, or checklists."""
    streams = get_bu_streams(request).order_by("name")

    context = {
        "streams": streams,
    }

    return render(request, "products/compliance_hub.html", context)


@login_required
def build_servers_list(request, stream=None):
    """List all build servers for a specific stream."""
    stream_obj = get_stream_or_404(stream, request=request)

    if hasattr(request.user, "custom_profile") and not request.user.custom_profile.can_access_stream(stream_obj.name):
        messages.error(request, f"Access denied. You do not have permission to view {stream_obj.name} build servers.")
        return redirect("build_servers_dashboard")

    servers = BuildServer.objects.filter(stream=stream_obj).order_by("hostname")

    # Calculate status counts for all servers in this stream (before filtering)
    all_servers = BuildServer.objects.filter(stream=stream_obj)
    status_counts = {}
    for status, _ in BuildServer.STATUS_CHOICES:
        status_counts[status] = all_servers.filter(status=status).count()

    status_filter = request.GET.get("status")
    if status_filter:
        servers = servers.filter(status=status_filter)

    stream_type_filter = request.GET.get("stream_type")
    if stream_type_filter:
        servers = servers.filter(stream_type=stream_type_filter)

    floor_filter = request.GET.get("floor")
    if floor_filter:
        servers = servers.filter(floor=floor_filter)

    search = request.GET.get("search")
    if search:
        servers = servers.filter(
            Q(hostname__icontains=search)
            | Q(ip_address__icontains=search)
            | Q(location__icontains=search)
            | Q(owner__icontains=search)
            | Q(purpose__icontains=search)
        )

    context = {
        "stream": stream,
        "stream_obj": stream_obj,
        "selected_stream": stream_obj,  # Add for base template compatibility
        "servers": servers,
        "status_counts": status_counts,
        "status_choices": BuildServer.STATUS_CHOICES,
        "stream_type_choices": BuildServer.SERVER_TYPES,
        "floor_choices": [(floor.id, floor.name) for floor in Floor.objects.filter(stream=stream_obj, is_active=True)],
        "current_filters": {
            "status": status_filter,
            "stream_type": stream_type_filter,
            "floor": floor_filter,
            "search": search,
        },
    }

    return render(request, "products/build_servers_list.html", context)


@login_required
def build_server_create(request, stream=None):  # noqa: C901, CCR001
    # pylint: disable=too-many-locals,too-complex
    """Create a new build server."""
    stream_obj = get_stream_or_404(stream, request=request)

    if not can_edit_products(request.user):
        messages.error(request, "Access denied. You do not have permission to create build servers.")
        return redirect("build_servers_list", stream=stream)

    if request.method == "POST":
        try:
            floor_id = request.POST.get("floor")
            floor_instance = None
            if floor_id:
                try:
                    floor_instance = Floor.objects.get(id=floor_id)
                except Floor.DoesNotExist:
                    floor_instance = None
            os_id = request.POST.get("operating_system")
            os_instance = None
            if os_id:
                try:
                    os_instance = OperatingSystem.objects.get(id=os_id)
                except OperatingSystem.DoesNotExist:
                    os_instance = None
            server = BuildServer(
                hostname=request.POST.get("hostname"),
                ip_address=request.POST.get("ip_address"),
                location=request.POST.get("location"),
                floor=floor_instance,
                owner=request.POST.get("owner"),
                stream_type=(stream_obj.name if stream_obj.name in ["PIC", "HIC"] else "Other"),
                stream=stream_obj,
                status=request.POST.get("status", "Active"),
                operating_system_ref=os_instance,
                cpu_cores=request.POST.get("cpu_cores") or None,
                ram_gb=request.POST.get("ram_gb") or None,
                storage_gb=request.POST.get("storage_gb") or None,
                mac_address=request.POST.get("mac_address", ""),
                domain=request.POST.get("domain", ""),
                ssh_port=request.POST.get("ssh_port", 22),
                purpose=request.POST.get("purpose", ""),
                project_allocation=request.POST.get("project_allocation", ""),
                cost_center=request.POST.get("cost_center", ""),
                primary_contact=request.POST.get("primary_contact", ""),
                secondary_contact=request.POST.get("secondary_contact", ""),
                contact_email=request.POST.get("contact_email", ""),
                notes=request.POST.get("notes", ""),
                tags=request.POST.get("tags", ""),
                created_by=request.user,
                updated_by=request.user,
            )

            procurement_date = request.POST.get("procurement_date")
            if procurement_date:
                server.procurement_date = datetime.strptime(procurement_date, "%Y-%m-%d").date()

            warranty_expiry = request.POST.get("warranty_expiry")
            if warranty_expiry:
                server.warranty_expiry = datetime.strptime(warranty_expiry, "%Y-%m-%d").date()

            last_maintenance = request.POST.get("last_maintenance")
            if last_maintenance:
                server.last_maintenance = datetime.strptime(last_maintenance, "%Y-%m-%d").date()

            next_maintenance = request.POST.get("next_maintenance")
            if next_maintenance:
                server.next_maintenance = datetime.strptime(next_maintenance, "%Y-%m-%d").date()

            server.save()

            AuditLog.log(
                "create",
                f"Created build server: {server.hostname}",
                user=request.user,
                request=request,
                obj=server,
                module="build_servers",
                severity="info",
                stream=stream_obj,
            )

            BuildServerHistory.objects.create(
                build_server=server,
                action="created",
                user=request.user,
                details=f"Build server {server.hostname} created",
            )

            messages.success(request, f"Build server {server.hostname} created successfully!")
            return redirect("build_servers_list", stream=stream)

        except Exception:
            logger.exception("Error creating build server")
            form_error = "An error occurred. Please try again."
    else:
        form_error = None

    context = {
        "stream": stream,
        "stream_obj": stream_obj,
        "selected_stream": stream_obj,
        "status_choices": BuildServer.STATUS_CHOICES,
        "stream_type_choices": BuildServer.SERVER_TYPES,
        "floor_choices": [(floor.id, floor.name) for floor in Floor.objects.filter(stream=stream_obj, is_active=True)],
        "operating_system_choices": [
            (os.id, str(os)) for os in OperatingSystem.objects.filter(stream=stream_obj, is_active=True)
        ],
        "form_error": form_error,
    }

    return render(request, "products/build_server_form.html", context)


@login_required
def build_server_detail(request, stream, server_id):
    """View detailed information about a build server."""
    stream_obj = get_stream_or_404(stream, request=request)

    server = get_object_or_404(BuildServer, id=server_id, stream=stream_obj)

    history = BuildServerHistory.objects.filter(build_server=server).order_by("-timestamp")[:20]

    maintenance_logs = BuildServerMaintenanceLog.objects.filter(build_server=server).order_by("-scheduled_date")[:10]

    context = {
        "stream": stream,
        "stream_obj": stream_obj,
        "selected_stream": stream_obj,
        "server": server,
        "history": history,
        "maintenance_logs": maintenance_logs,
    }

    return render(request, "products/build_server_detail.html", context)


@login_required
def build_server_edit(request, stream, server_id):  # noqa: C901, CCR001
    # pylint: disable=too-many-locals,too-many-branches,too-many-statements,too-complex
    """Edit a build server."""
    stream_obj = get_stream_or_404(stream, request=request)

    server = get_object_or_404(BuildServer, id=server_id, stream=stream_obj)

    if not can_edit_products(request.user):
        messages.error(request, "Access denied. You do not have permission to edit build servers.")
        return redirect("build_server_detail", stream=stream, server_id=server_id)

    if request.method == "POST":
        try:
            old_values = {
                "hostname": server.hostname,
                "ip_address": server.ip_address,
                "location": server.location,
                "floor": server.floor,
                "owner": server.owner,
                "status": server.status,
            }

            server.hostname = request.POST.get("hostname")
            server.ip_address = request.POST.get("ip_address")
            server.location = request.POST.get("location")
            floor_id = request.POST.get("floor")
            if floor_id:
                try:
                    server.floor = Floor.objects.get(id=floor_id)
                except Floor.DoesNotExist:
                    server.floor = None
            else:
                server.floor = None
            server.owner = request.POST.get("owner")
            # Stream type is now static, not editable from form
            # server.stream_type remains unchanged

            # Handle status change with explicit validation
            new_status = request.POST.get("status")
            _approval_block = None
            if new_status and new_status in dict(BuildServer.STATUS_CHOICES):
                # ── Pre-action enforcement: block decommission if approval required ──
                if old_values["status"] != new_status and new_status in ("Inactive", "Offline"):
                    _approval_block = check_approval_required(
                        "server_decommission",
                        stream_obj.business_unit,
                        request.user,
                        entity_obj=server,
                        stream=stream_obj,
                        title=f"Server '{server.hostname}' \u2192 {new_status}",
                        description=(
                            f"Build server {server.hostname} ({server.ip_address}) status change "
                            f"from {old_values['status']} to {new_status}"
                        ),
                        intended_changes={
                            "action_type": "status_change",
                            "model_label": "products.BuildServer",
                            "pk": server.pk,
                            "changes": {"status": new_status},
                            "revert": {"status": old_values["status"]},
                            "metadata": {"entity_name": server.hostname, "stream_name": stream},
                        },
                    )
                elif old_values["status"] != new_status and new_status == "Maintenance":
                    _approval_block = check_approval_required(
                        "server_maintenance",
                        stream_obj.business_unit,
                        request.user,
                        entity_obj=server,
                        stream=stream_obj,
                        title=f"Server '{server.hostname}' \u2192 Under Maintenance",
                        description=(
                            f"Build server {server.hostname} ({server.ip_address}) maintenance mode "
                            f"from {old_values['status']} to Maintenance"
                        ),
                        intended_changes={
                            "action_type": "status_change",
                            "model_label": "products.BuildServer",
                            "pk": server.pk,
                            "changes": {"status": "Maintenance"},
                            "revert": {"status": old_values["status"]},
                            "metadata": {"entity_name": server.hostname, "stream_name": stream},
                        },
                    )
                server.status = old_values["status"] if _approval_block else new_status
            else:
                # Keep existing status if no valid status provided
                pass
            os_id = request.POST.get("operating_system")
            if os_id:
                try:
                    server.operating_system_ref = OperatingSystem.objects.get(id=os_id)
                except OperatingSystem.DoesNotExist:
                    server.operating_system_ref = None
            else:
                server.operating_system_ref = None
            server.cpu_cores = request.POST.get("cpu_cores") or None
            server.ram_gb = request.POST.get("ram_gb") or None
            server.storage_gb = request.POST.get("storage_gb") or None
            server.mac_address = request.POST.get("mac_address", "")
            server.domain = request.POST.get("domain", "")
            server.ssh_port = request.POST.get("ssh_port", 22)
            server.purpose = request.POST.get("purpose", "")
            server.project_allocation = request.POST.get("project_allocation", "")
            server.cost_center = request.POST.get("cost_center", "")
            server.primary_contact = request.POST.get("primary_contact", "")
            server.secondary_contact = request.POST.get("secondary_contact", "")
            server.contact_email = request.POST.get("contact_email", "")
            server.notes = request.POST.get("notes", "")
            server.tags = request.POST.get("tags", "")
            server.updated_by = request.user

            procurement_date = request.POST.get("procurement_date")
            if procurement_date:
                server.procurement_date = datetime.strptime(procurement_date, "%Y-%m-%d").date()
            else:
                server.procurement_date = None

            warranty_expiry = request.POST.get("warranty_expiry")
            if warranty_expiry:
                server.warranty_expiry = datetime.strptime(warranty_expiry, "%Y-%m-%d").date()
            else:
                server.warranty_expiry = None

            last_maintenance = request.POST.get("last_maintenance")
            if last_maintenance:
                server.last_maintenance = datetime.strptime(last_maintenance, "%Y-%m-%d").date()
            else:
                server.last_maintenance = None

            next_maintenance = request.POST.get("next_maintenance")
            if next_maintenance:
                server.next_maintenance = datetime.strptime(next_maintenance, "%Y-%m-%d").date()
            else:
                server.next_maintenance = None

            server.save()

            new_values = {
                "hostname": server.hostname,
                "ip_address": server.ip_address,
                "location": server.location,
                "floor": server.floor,
                "owner": server.owner,
                "status": server.status,
            }

            BuildServerHistory.objects.create(
                build_server=server,
                action="updated",
                user=request.user,
                details=f"Build server {server.hostname} updated",
                old_values=old_values,
                new_values=new_values,
            )

            AuditLog.log(
                "update",
                f"Updated build server: {server.hostname}",
                user=request.user,
                request=request,
                obj=server,
                module="build_servers",
                severity="info",
                stream=stream_obj,
            )

            # ── Auto-trigger approval for server decommission ──
            if _approval_block:
                messages.warning(
                    request,
                    f'\u23f3 Status change to "{new_status}" requires approval. '
                    f'Request #{_approval_block.id} submitted.',
                )
            else:
                messages.success(request, f"Build server {server.hostname} updated successfully!")
            return redirect("build_server_detail", stream=stream, server_id=server_id)

        except Exception:
            form_error = "An error occurred. Please try again."
    else:
        form_error = None

    context = {
        "stream": stream,
        "stream_obj": stream_obj,
        "selected_stream": stream_obj,
        "server": server,
        "status_choices": BuildServer.STATUS_CHOICES,
        "stream_type_choices": BuildServer.SERVER_TYPES,
        "floor_choices": [(floor.id, floor.name) for floor in Floor.objects.filter(stream=stream_obj, is_active=True)],
        "operating_system_choices": [
            (os.id, str(os)) for os in OperatingSystem.objects.filter(stream=stream_obj, is_active=True)
        ],
        "is_edit": True,
        "form_error": form_error,
    }

    return render(request, "products/build_server_form.html", context)


@login_required
def build_server_delete(request, stream, server_id):
    """Delete a build server."""
    stream_obj = get_stream_or_404(stream, request=request)

    server = get_object_or_404(BuildServer, id=server_id, stream=stream_obj)

    if not can_delete_products(request.user):
        messages.error(request, "Access denied. You do not have permission to delete build servers.")
        return redirect("build_server_detail", stream=stream, server_id=server_id)

    if request.method == "POST":
        hostname = server.hostname
        _bu = stream_obj.business_unit

        # ── Pre-action enforcement: block delete if approval required ──
        _approval = check_approval_required(
            "server_deleted",
            _bu,
            request.user,
            entity_obj=server,
            stream=stream_obj,
            title=f"Build server '{hostname}' deletion",
            description=f"Build server {hostname} delete requested from stream {stream}",
            intended_changes={
                "action_type": "delete",
                "model_label": "products.BuildServer",
                "pk": server.pk,
                "metadata": {"entity_name": hostname, "stream_name": stream},
            },
        )
        if _approval:
            messages.warning(
                request,
                f'\u23f3 Deleting server "{hostname}" requires approval. Request #{_approval.id} submitted.',
            )
            return redirect("build_servers_list", stream=stream)

        AuditLog.log(
            "delete",
            f"Deleted build server: {server.hostname}",
            user=request.user,
            request=request,
            obj=server,
            module="build_servers",
            severity="warning",
            stream=stream_obj,
        )
        server.delete()
        messages.success(request, f"Build server {hostname} deleted successfully!")
        return redirect("build_servers_list", stream=stream)

    context = {
        "stream": stream,
        "stream_obj": stream_obj,
        "selected_stream": stream_obj,
        "server": server,
    }

    return render(request, "products/build_server_confirm_delete.html", context)


@login_required
@require_http_methods(["GET"])
def build_servers_api(request, stream=None):
    """API endpoint to get build servers data for AJAX requests."""
    stream_obj = get_stream_or_404(stream, request=request)

    if hasattr(request.user, "custom_profile") and not request.user.custom_profile.can_access_stream(stream_obj.name):
        return JsonResponse({"error": "Access denied"}, status=403)

    try:
        servers = BuildServer.objects.filter(stream=stream_obj)

        status_filter = request.GET.get("status")
        if status_filter:
            servers = servers.filter(status=status_filter)

        stream_type_filter = request.GET.get("stream_type")
        if stream_type_filter:
            servers = servers.filter(stream_type=stream_type_filter)

        servers_data = []
        for server in servers:
            servers_data.append(
                {
                    "id": server.id,
                    "hostname": server.hostname,
                    "ip_address": server.ip_address,
                    "location": server.location,
                    "floor": server.floor.name if server.floor else "",
                    "owner": server.owner,
                    "stream_type": server.stream_type,
                    "status": server.status,
                    "status_display": server.get_status_display(),
                    "operating_system": str(server.operating_system_ref) if server.operating_system_ref else "",
                    "cpu_cores": server.cpu_cores,
                    "ram_gb": server.ram_gb,
                    "storage_gb": server.storage_gb,
                    "uptime_percentage": float(server.uptime_percentage),
                    "warranty_expiring_soon": server.is_warranty_expiring_soon(),
                    "days_until_warranty_expiry": server.days_until_warranty_expiry(),
                    "url": reverse("build_server_detail", kwargs={"stream": stream, "server_id": server.id}),
                }
            )

        return JsonResponse({"servers": servers_data, "total_count": len(servers_data)})

    except Exception:
        logger.exception("Operation failed")
        return JsonResponse({"error": "An unexpected error occurred"}, status=500)


@login_required
def build_servers_export(request, stream=None):  # noqa: CCR001
    # pylint: disable=too-many-locals
    """Export build servers data to Excel."""
    stream_obj = get_stream_or_404(stream, request=request)

    if hasattr(request.user, "custom_profile") and not request.user.custom_profile.can_access_stream(stream_obj.name):
        messages.error(request, f"Access denied. You do not have permission to export {stream_obj.name} build servers.")
        return redirect("build_servers_list", stream=stream)

    servers = BuildServer.objects.filter(stream=stream_obj).order_by("hostname")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{stream_obj.name} Build Servers"

    headers = [
        "Hostname",
        "IP Address",
        "Location",
        "Floor",
        "Owner",
        "Stream Type",
        "Status",
        "Operating System",
        "CPU Cores",
        "RAM (GB)",
        "Storage (GB)",
        "MAC Address",
        "Domain",
        "SSH Port",
        "Purpose",
        "Project Allocation",
        "Primary Contact",
        "Secondary Contact",
        "Contact Email",
        "Procurement Date",
        "Warranty Expiry",
        "Last Maintenance",
        "Next Maintenance",
        "Uptime %",
        "Notes",
        "Tags",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, server in enumerate(servers, 2):
        data = [
            server.hostname,
            server.ip_address,
            server.location,
            server.floor.name if server.floor else "",
            server.owner,
            server.stream_type,
            server.status,
            str(server.operating_system_ref) if server.operating_system_ref else "",
            server.cpu_cores or "",
            server.ram_gb or "",
            server.storage_gb or "",
            server.mac_address or "",
            server.domain or "",
            server.ssh_port,
            server.purpose or "",
            server.project_allocation or "",
            server.primary_contact or "",
            server.secondary_contact or "",
            server.contact_email or "",
            server.procurement_date.strftime("%Y-%m-%d") if server.procurement_date else "",
            server.warranty_expiry.strftime("%Y-%m-%d") if server.warranty_expiry else "",
            server.last_maintenance.strftime("%Y-%m-%d") if server.last_maintenance else "",
            server.next_maintenance.strftime("%Y-%m-%d") if server.next_maintenance else "",
            float(server.uptime_percentage),
            server.notes or "",
            server.tags or "",
        ]

        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{stream_obj.name}_build_servers_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )

    return response
